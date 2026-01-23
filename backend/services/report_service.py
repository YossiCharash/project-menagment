from sqlalchemy import func, select, and_, or_
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import date, timedelta
from typing import List, Dict, Any
from dateutil.relativedelta import relativedelta
import hashlib

from backend.models.transaction import Transaction
from backend.models.category import Category
from backend.models.project import Project
from backend.models.budget import Budget
from backend.models.supplier import Supplier
from backend.services.budget_service import BudgetService
from backend.services.fund_service import FundService
from backend.services.project_service import calculate_start_date, calculate_monthly_income_amount
import io
from io import BytesIO
import zipfile
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

try:
    from openpyxl.chart import PieChart, BarChart, LineChart, Reference

    CHARTS_AVAILABLE = True
except ImportError:
    CHARTS_AVAILABLE = False
    print("אזהרה: openpyxl.chart לא זמין - גרפים יידלגו")

# Helper function for consistent date formatting
def format_date_hebrew(d) -> str:
    """Format date in Hebrew format (DD/MM/YYYY)"""
    if d is None:
        return ""
    if isinstance(d, str):
        # Try to parse ISO format
        try:
            if 'T' in d:
                d = date.fromisoformat(d.split('T')[0])
            else:
                d = date.fromisoformat(d)
        except (ValueError, AttributeError):
            return d  # Return as-is if parsing fails
    if isinstance(d, date):
        return d.strftime('%d/%m/%Y')
    return str(d)

# Hebrew Labels to avoid hardcoded strings in logic
REPORT_LABELS = {
    "project_report": "דוח פרויקט",
    "production_date": "תאריך הפקה",
    "financial_summary": "סיכום פיננסי",
    "details": "פירוט",
    "amount": "סכום",
    "total_income": "סה״כ הכנסות",
    "total_expenses": "סה״כ הוצאות",
    "balance_profit": "יתרה / רווח",
    "fund_status": "מצב קופה",
    "current_balance": "יתרה נוכחית",
    "monthly_deposit": "הפקדה חודשית",
    "budget_vs_actual": "תקציב מול ביצוע",
    "category": "קטגוריה",
    "budget": "תקציב",
    "used": "נוצל",
    "remaining": "נותר",
    "status": "סטטוס",
    "general": "כללי",
    "transaction_details": "פירוט תנועות",
    "date": "תאריך",
    "type": "סוג",
    "description": "תיאור",
    "income": "הכנסה",
    "expense": "הוצאה",
    "expenses": "הוצאות",
    "payment_method": "אמצעי תשלום",
    "notes": "הערות",
    "file": "קובץ",
    "yes": "כן",
    "no": "לא",
    "exception": "חריגה",
    "ok": "תקין",
    "profit": "רווח",
    "monthly_budget": "תקציב (חודשי)",
    "annual_budget": "תקציב (שנתי)",
    "categories": "קטגוריות",
    "supplier": "ספק"
}

# Payment method translation - maps enum names to Hebrew values
PAYMENT_METHOD_TRANSLATIONS = {
    "STANDING_ORDER": "הוראת קבע",
    "CREDIT": "אשראי",
    "CHECK": "שיק",
    "CASH": "מזומן",
    "BANK_TRANSFER": "העברה בנקאית",
    "CENTRALIZED_YEAR_END": "גבייה מרוכזת סוף שנה",
    # Also include the PaymentMethod enum format (e.g. PaymentMethod.BANK_TRANSFER)
    "PaymentMethod.STANDING_ORDER": "הוראת קבע",
    "PaymentMethod.CREDIT": "אשראי",
    "PaymentMethod.CHECK": "שיק",
    "PaymentMethod.CASH": "מזומן",
    "PaymentMethod.BANK_TRANSFER": "העברה בנקאית",
    "PaymentMethod.CENTRALIZED_YEAR_END": "גבייה מרוכזת סוף שנה",
}

def translate_payment_method(payment_method) -> str:
    """Translate payment method enum to Hebrew"""
    if not payment_method:
        return ""
    pm_str = str(payment_method)
    # Check if it's already in Hebrew (one of the values)
    if pm_str in PAYMENT_METHOD_TRANSLATIONS.values():
        return pm_str
    # Look up in translations
    return PAYMENT_METHOD_TRANSLATIONS.get(pm_str, pm_str)


class ReportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def _calculate_expenses_with_period(
            self,
            project_id: int | None,
            start_date: date,
            end_date: date,
            from_fund: bool = False
    ) -> float:
        """
        Calculate total expenses for a period, handling both regular transactions (sum)
        and period-based transactions (pro-rated split).
        If project_id is None, calculates for all projects.
        """
        # 1. Regular expenses
        query_regular = select(func.coalesce(func.sum(Transaction.amount), 0)).where(
            and_(
                Transaction.type == "Expense",
                Transaction.from_fund == from_fund,
                Transaction.tx_date >= start_date,
                Transaction.tx_date <= end_date,
                or_(
                    Transaction.period_start_date.is_(None),
                    Transaction.period_end_date.is_(None)
                )
            )
        )

        # 2. Period expenses
        query_period = select(Transaction).options(
            selectinload(Transaction.category)
        ).where(
            and_(
                Transaction.type == "Expense",
                Transaction.from_fund == from_fund,
                Transaction.period_start_date.is_not(None),
                Transaction.period_end_date.is_not(None),
                Transaction.period_start_date <= end_date,
                Transaction.period_end_date >= start_date
            )
        )

        if project_id is not None:
            query_regular = query_regular.where(Transaction.project_id == project_id)
            query_period = query_period.where(Transaction.project_id == project_id)

        regular_expense = float((await self.db.execute(query_regular)).scalar_one())
        period_txs = (await self.db.execute(query_period)).scalars().all()

        period_expense = 0.0
        for tx in period_txs:
            total_days = (tx.period_end_date - tx.period_start_date).days + 1
            if total_days <= 0: continue

            daily_rate = float(tx.amount) / total_days
            overlap_start = max(tx.period_start_date, start_date)
            overlap_end = min(tx.period_end_date, end_date)
            overlap_days = (overlap_end - overlap_start).days + 1

            if overlap_days > 0:
                period_expense += daily_rate * overlap_days

        return regular_expense + period_expense

    async def _calculate_category_expenses_with_period(
            self,
            project_id: int | None,
            start_date: date,
            end_date: date,
            from_fund: bool = False
    ) -> Dict[str, float]:
        """
        Calculate expenses per category for a period, handling splitting.
        Returns {category_name: amount}
        If project_id is None, calculates for all projects.
        """
        # 1. Regular expenses grouped by category
        query_regular = select(
            Category.name.label('category'),
            func.coalesce(func.sum(Transaction.amount), 0).label('total_amount')
        ).outerjoin(
            Category, Transaction.category_id == Category.id
        ).where(
            and_(
                Transaction.type == "Expense",
                Transaction.from_fund == from_fund,
                Transaction.tx_date >= start_date,
                Transaction.tx_date <= end_date,
                or_(
                    Transaction.period_start_date.is_(None),
                    Transaction.period_end_date.is_(None)
                )
            )
        ).group_by(Category.name)

        # 2. Period expenses
        query_period = select(Transaction).options(selectinload(Transaction.category)).where(
            and_(
                Transaction.type == "Expense",
                Transaction.from_fund == from_fund,
                Transaction.period_start_date.is_not(None),
                Transaction.period_end_date.is_not(None),
                Transaction.period_start_date <= end_date,
                Transaction.period_end_date >= start_date
            )
        )

        if project_id is not None:
            query_regular = query_regular.where(Transaction.project_id == project_id)
            query_period = query_period.where(Transaction.project_id == project_id)

        regular_results = await self.db.execute(query_regular)
        category_expenses = {}
        for row in regular_results:
            cat_name = row.category or REPORT_LABELS["general"]
            category_expenses[cat_name] = float(row.total_amount)

        period_txs = (await self.db.execute(query_period)).scalars().all()

        for tx in period_txs:
            total_days = (tx.period_end_date - tx.period_start_date).days + 1
            if total_days <= 0: continue

            daily_rate = float(tx.amount) / total_days
            overlap_start = max(tx.period_start_date, start_date)
            overlap_end = min(tx.period_end_date, end_date)
            overlap_days = (overlap_end - overlap_start).days + 1

            if overlap_days > 0:
                amount = daily_rate * overlap_days
                cat_name = tx.category.name if tx.category else REPORT_LABELS["general"]
                category_expenses[cat_name] = category_expenses.get(cat_name, 0.0) + amount

        return category_expenses

    async def _calculate_monthly_category_supplier_expenses(
            self,
            project_id: int | None,
            start_date: date,
            end_date: date,
            from_fund: bool = False
    ) -> List[Dict[str, Any]]:
        """
        Calculate expenses per month, category, and supplier for a period.
        Returns list of dicts with: month, category, supplier, amount
        If there are multiple suppliers in the same category/month, each gets a separate row.
        If project_id is None, calculates for all projects.
        """
        from backend.models.supplier import Supplier
        from collections import defaultdict
        
        # Structure: {(month_key, category_name, supplier_name): amount}
        monthly_data = defaultdict(float)
        
        # 1. Regular expenses (no period dates)
        query_regular = select(
            Transaction.tx_date,
            Category.name.label('category'),
            Supplier.name.label('supplier'),
            Transaction.amount
        ).outerjoin(
            Category, Transaction.category_id == Category.id
        ).outerjoin(
            Supplier, Transaction.supplier_id == Supplier.id
        ).where(
            and_(
                Transaction.type == "Expense",
                Transaction.from_fund == from_fund,
                Transaction.tx_date >= start_date,
                Transaction.tx_date <= end_date,
                or_(
                    Transaction.period_start_date.is_(None),
                    Transaction.period_end_date.is_(None)
                )
            )
        )
        
        if project_id is not None:
            query_regular = query_regular.where(Transaction.project_id == project_id)
        
        regular_results = await self.db.execute(query_regular)
        for row in regular_results:
            month_key = row.tx_date.strftime('%Y-%m')
            cat_name = row.category or REPORT_LABELS["general"]
            supplier_name = row.supplier or "ללא ספק"
            amount = float(row.amount)
            monthly_data[(month_key, cat_name, supplier_name)] += amount
        
        # 2. Period expenses - split by month
        query_period = select(Transaction).options(
            selectinload(Transaction.category),
            selectinload(Transaction.supplier)
        ).where(
            and_(
                Transaction.type == "Expense",
                Transaction.from_fund == from_fund,
                Transaction.period_start_date.is_not(None),
                Transaction.period_end_date.is_not(None),
                Transaction.period_start_date <= end_date,
                Transaction.period_end_date >= start_date
            )
        )
        
        if project_id is not None:
            query_period = query_period.where(Transaction.project_id == project_id)
        
        period_txs = (await self.db.execute(query_period)).scalars().all()
        
        for tx in period_txs:
            cat_name = tx.category.name if tx.category else REPORT_LABELS["general"]
            supplier_name = tx.supplier.name if tx.supplier else "ללא ספק"
            
            # Split period transaction by month
            # Start from the later of: transaction start or filter start_date
            # End at the earlier of: transaction end or filter end_date
            effective_start = max(tx.period_start_date, start_date)
            effective_end = min(tx.period_end_date, end_date)
            
            if effective_start > effective_end:
                continue
            
            total_days = (tx.period_end_date - tx.period_start_date).days + 1
            if total_days <= 0:
                continue
            
            daily_rate = float(tx.amount) / total_days
            
            # Start from the first day of the month containing effective_start
            current_date = date(effective_start.year, effective_start.month, 1)
            
            while current_date <= effective_end:
                month_key = current_date.strftime('%Y-%m')
                
                # Calculate days in this month for this transaction
                if current_date.month == 12:
                    month_end = date(current_date.year + 1, 1, 1)
                else:
                    month_end = date(current_date.year, current_date.month + 1, 1)
                
                # Calculate the overlap between the transaction period and this month
                # within the filtered date range
                month_period_start = max(effective_start, current_date)
                month_period_end = min(effective_end, month_end - timedelta(days=1))
                
                # Calculate days from month_period_start to month_period_end (inclusive)
                days_in_month = (month_period_end - month_period_start).days + 1
                
                if days_in_month > 0:
                    amount = daily_rate * days_in_month
                    monthly_data[(month_key, cat_name, supplier_name)] += amount
                
                # Move to next month (start of next month)
                current_date = month_end
        
        # Convert to list of dicts
        result = []
        for (month_key, cat_name, supplier_name), amount in monthly_data.items():
            if amount > 0:  # Only include rows with expenses
                result.append({
                    'month': month_key,
                    'category': cat_name,
                    'supplier': supplier_name,
                    'amount': amount
                })
        
        # Sort by month, then category, then supplier
        result.sort(key=lambda x: (x['month'], x['category'], x['supplier']))
        
        return result

    async def project_profitability(self, project_id: int, start_date: date | None = None, end_date: date | None = None) -> dict:
        from sqlalchemy import or_
        from datetime import timedelta
        
        # For period transactions, we need to calculate proportional amounts
        # So we'll fetch transactions and calculate manually
        
        # 1. Regular income (no period dates) - filter by tx_date
        income_conditions_regular = [
            Transaction.project_id == project_id,
            Transaction.type == "Income",
            Transaction.from_fund == False,
            or_(
                Transaction.period_start_date.is_(None),
                Transaction.period_end_date.is_(None)
            )
        ]
        if start_date:
            income_conditions_regular.append(Transaction.tx_date >= start_date)
        if end_date:
            income_conditions_regular.append(Transaction.tx_date <= end_date)
        
        regular_income_q = select(func.coalesce(func.sum(Transaction.amount), 0)).where(
            and_(*income_conditions_regular)
        )
        regular_income = float((await self.db.execute(regular_income_q)).scalar_one())
        
        # 2. Period income - check if period overlaps with date range
        period_income_conditions = [
            Transaction.project_id == project_id,
            Transaction.type == "Income",
            Transaction.from_fund == False,
            Transaction.period_start_date.is_not(None),
            Transaction.period_end_date.is_not(None)
        ]
        if start_date or end_date:
            if start_date and end_date:
                # Period overlaps if: period_start <= end_date AND period_end >= start_date
                period_income_conditions.append(
                    and_(
                        Transaction.period_start_date <= end_date,
                        Transaction.period_end_date >= start_date
                    )
                )
            elif start_date:
                period_income_conditions.append(Transaction.period_end_date >= start_date)
            elif end_date:
                period_income_conditions.append(Transaction.period_start_date <= end_date)
        
        period_income_q = select(Transaction).where(and_(*period_income_conditions))
        period_income_txs = (await self.db.execute(period_income_q)).scalars().all()
        
        period_income = 0.0
        for tx in period_income_txs:
            if start_date or end_date:
                # Calculate proportional amount for the overlap
                overlap_start = max(tx.period_start_date, start_date if start_date else tx.period_start_date)
                overlap_end = min(tx.period_end_date, end_date if end_date else tx.period_end_date)
                total_days = (tx.period_end_date - tx.period_start_date).days + 1
                overlap_days = (overlap_end - overlap_start).days + 1
                if total_days > 0 and overlap_days > 0:
                    period_income += float(tx.amount) * (overlap_days / total_days)
            else:
                # No date filter, include full amount
                period_income += float(tx.amount)
        
        total_income = regular_income + period_income
        
        # 3. Regular expenses (no period dates) - filter by tx_date
        expense_conditions_regular = [
            Transaction.project_id == project_id,
            Transaction.type == "Expense",
            Transaction.from_fund == False,
            or_(
                Transaction.period_start_date.is_(None),
                Transaction.period_end_date.is_(None)
            )
        ]
        if start_date:
            expense_conditions_regular.append(Transaction.tx_date >= start_date)
        if end_date:
            expense_conditions_regular.append(Transaction.tx_date <= end_date)
        
        regular_expense_q = select(func.coalesce(func.sum(Transaction.amount), 0)).where(
            and_(*expense_conditions_regular)
        )
        regular_expense = float((await self.db.execute(regular_expense_q)).scalar_one())
        
        # 4. Period expenses - check if period overlaps with date range
        period_expense_conditions = [
            Transaction.project_id == project_id,
            Transaction.type == "Expense",
            Transaction.from_fund == False,
            Transaction.period_start_date.is_not(None),
            Transaction.period_end_date.is_not(None)
        ]
        if start_date or end_date:
            if start_date and end_date:
                # Period overlaps if: period_start <= end_date AND period_end >= start_date
                period_expense_conditions.append(
                    and_(
                        Transaction.period_start_date <= end_date,
                        Transaction.period_end_date >= start_date
                    )
                )
            elif start_date:
                period_expense_conditions.append(Transaction.period_end_date >= start_date)
            elif end_date:
                period_expense_conditions.append(Transaction.period_start_date <= end_date)
        
        period_expense_q = select(Transaction).where(and_(*period_expense_conditions))
        period_expense_txs = (await self.db.execute(period_expense_q)).scalars().all()
        
        period_expense = 0.0
        for tx in period_expense_txs:
            if start_date or end_date:
                # Calculate proportional amount for the overlap
                overlap_start = max(tx.period_start_date, start_date if start_date else tx.period_start_date)
                overlap_end = min(tx.period_end_date, end_date if end_date else tx.period_end_date)
                total_days = (tx.period_end_date - tx.period_start_date).days + 1
                overlap_days = (overlap_end - overlap_start).days + 1
                if total_days > 0 and overlap_days > 0:
                    period_expense += float(tx.amount) * (overlap_days / total_days)
            else:
                # No date filter, include full amount
                period_expense += float(tx.amount)
        
        total_expense = regular_expense + period_expense

        proj = (await self.db.execute(select(Project).where(Project.id == project_id))).scalar_one()

        income = total_income
        expenses = total_expense
        profit = income - expenses

        # Check if project has budgets or funds
        has_budget = proj.budget_monthly > 0 or proj.budget_annual > 0
        # Check fund exists
        fund_service = FundService(self.db)
        fund = await fund_service.get_fund_by_project(project_id)
        has_fund = fund is not None

        return {
            "project_id": project_id,
            "income": income,
            "expenses": expenses,
            "profit": profit,
            "budget_monthly": float(proj.budget_monthly or 0),
            "budget_annual": float(proj.budget_annual or 0),
            "has_budget": has_budget,
            "has_fund": has_fund
        }

    async def _calculate_summary_with_filters(
            self,
            project_id: int,
            start_date: date | None,
            end_date: date | None,
            transactions: List[Dict] | None = None
    ) -> Dict[str, Any]:
        """
        Calculate financial summary based on date filters.
        Uses already-fetched transactions if available to avoid duplicate queries.
        """
        # If transactions were already fetched, calculate from them
        if transactions:
            income = sum(
                float(tx.get('amount', 0)) 
                for tx in transactions 
                if tx.get('type') == 'Income'
            )
            expenses = sum(
                float(tx.get('amount', 0)) 
                for tx in transactions 
                if tx.get('type') == 'Expense'
            )
        else:
            # Query database with date filters, handling period transactions
            from sqlalchemy import or_
            
            # Regular income
            regular_income_conditions = [
                Transaction.project_id == project_id,
                Transaction.type == "Income",
                Transaction.from_fund == False,
                or_(
                    Transaction.period_start_date.is_(None),
                    Transaction.period_end_date.is_(None)
                )
            ]
            if start_date:
                regular_income_conditions.append(Transaction.tx_date >= start_date)
            if end_date:
                regular_income_conditions.append(Transaction.tx_date <= end_date)
            
            regular_income_q = select(func.coalesce(func.sum(Transaction.amount), 0)).where(
                and_(*regular_income_conditions)
            )
            regular_income = float((await self.db.execute(regular_income_q)).scalar_one())
            
            # Period income
            period_income_conditions = [
                Transaction.project_id == project_id,
                Transaction.type == "Income",
                Transaction.from_fund == False,
                Transaction.period_start_date.is_not(None),
                Transaction.period_end_date.is_not(None)
            ]
            if start_date or end_date:
                if start_date and end_date:
                    period_income_conditions.append(
                        and_(
                            Transaction.period_start_date <= end_date,
                            Transaction.period_end_date >= start_date
                        )
                    )
                elif start_date:
                    period_income_conditions.append(Transaction.period_end_date >= start_date)
                elif end_date:
                    period_income_conditions.append(Transaction.period_start_date <= end_date)
            
            period_income_q = select(Transaction).where(and_(*period_income_conditions))
            period_income_txs = (await self.db.execute(period_income_q)).scalars().all()
            
            period_income = 0.0
            for tx in period_income_txs:
                if start_date or end_date:
                    overlap_start = max(tx.period_start_date, start_date if start_date else tx.period_start_date)
                    overlap_end = min(tx.period_end_date, end_date if end_date else tx.period_end_date)
                    total_days = (tx.period_end_date - tx.period_start_date).days + 1
                    overlap_days = (overlap_end - overlap_start).days + 1
                    if total_days > 0 and overlap_days > 0:
                        period_income += float(tx.amount) * (overlap_days / total_days)
                else:
                    period_income += float(tx.amount)
            
            income = regular_income + period_income
            
            # Regular expenses
            regular_expense_conditions = [
                Transaction.project_id == project_id,
                Transaction.type == "Expense",
                Transaction.from_fund == False,
                or_(
                    Transaction.period_start_date.is_(None),
                    Transaction.period_end_date.is_(None)
                )
            ]
            if start_date:
                regular_expense_conditions.append(Transaction.tx_date >= start_date)
            if end_date:
                regular_expense_conditions.append(Transaction.tx_date <= end_date)
            
            regular_expense_q = select(func.coalesce(func.sum(Transaction.amount), 0)).where(
                and_(*regular_expense_conditions)
            )
            regular_expense = float((await self.db.execute(regular_expense_q)).scalar_one())
            
            # Period expenses
            period_expense_conditions = [
                Transaction.project_id == project_id,
                Transaction.type == "Expense",
                Transaction.from_fund == False,
                Transaction.period_start_date.is_not(None),
                Transaction.period_end_date.is_not(None)
            ]
            if start_date or end_date:
                if start_date and end_date:
                    period_expense_conditions.append(
                        and_(
                            Transaction.period_start_date <= end_date,
                            Transaction.period_end_date >= start_date
                        )
                    )
                elif start_date:
                    period_expense_conditions.append(Transaction.period_end_date >= start_date)
                elif end_date:
                    period_expense_conditions.append(Transaction.period_start_date <= end_date)
            
            period_expense_q = select(Transaction).where(and_(*period_expense_conditions))
            period_expense_txs = (await self.db.execute(period_expense_q)).scalars().all()
            
            period_expense = 0.0
            for tx in period_expense_txs:
                if start_date or end_date:
                    overlap_start = max(tx.period_start_date, start_date if start_date else tx.period_start_date)
                    overlap_end = min(tx.period_end_date, end_date if end_date else tx.period_end_date)
                    total_days = (tx.period_end_date - tx.period_start_date).days + 1
                    overlap_days = (overlap_end - overlap_start).days + 1
                    if total_days > 0 and overlap_days > 0:
                        period_expense += float(tx.amount) * (overlap_days / total_days)
                else:
                    period_expense += float(tx.amount)
            
            expenses = regular_expense + period_expense

        profit = income - expenses

        # Get project info for budget data
        proj = (await self.db.execute(select(Project).where(Project.id == project_id))).scalar_one()
        has_budget = proj.budget_monthly > 0 or proj.budget_annual > 0
        
        fund_service = FundService(self.db)
        fund = await fund_service.get_fund_by_project(project_id)
        has_fund = fund is not None

        return {
            "project_id": project_id,
            "income": income,
            "expenses": expenses,
            "profit": profit,
            "budget_monthly": float(proj.budget_monthly or 0),
            "budget_annual": float(proj.budget_annual or 0),
            "has_budget": has_budget,
            "has_fund": has_fund
        }

    # ==================== OPTIMIZED HELPER FUNCTIONS ====================
    # These functions calculate from already-fetched data to avoid duplicate DB queries
    
    def _calculate_summary_from_transactions_sync(
            self,
            transactions: List[Dict],
            project: Any = None,
            fund: Any = None
    ) -> Dict[str, Any]:
        """
        Calculate financial summary from already-fetched transactions.
        This is a synchronous function that works on data already in memory.
        Handles period-based transactions with pro-rata calculation.
        """
        income = 0.0
        expenses = 0.0
        
        for tx in transactions:
            # Skip fund transactions
            if tx.get('from_fund', False):
                continue
                
            tx_type = tx.get('type')
            amount = float(tx.get('amount', 0) or 0)
            
            if tx_type == 'Income':
                income += amount
            elif tx_type == 'Expense':
                expenses += amount
        
        profit = income - expenses
        
        # Get budget info from project if provided
        budget_monthly = 0.0
        budget_annual = 0.0
        has_budget = False
        
        if project:
            if hasattr(project, 'budget_monthly'):
                budget_monthly = float(project.budget_monthly or 0)
            elif isinstance(project, dict):
                budget_monthly = float(project.get('budget_monthly', 0) or 0)
                
            if hasattr(project, 'budget_annual'):
                budget_annual = float(project.budget_annual or 0)
            elif isinstance(project, dict):
                budget_annual = float(project.get('budget_annual', 0) or 0)
                
            has_budget = budget_monthly > 0 or budget_annual > 0
        
        has_fund = fund is not None
        
        return {
            "income": income,
            "expenses": expenses,
            "profit": profit,
            "budget_monthly": budget_monthly,
            "budget_annual": budget_annual,
            "has_budget": has_budget,
            "has_fund": has_fund
        }

    def _calculate_expenses_from_transactions_sync(
            self,
            transactions: List[Dict],
            start_date: date | None = None,
            end_date: date | None = None,
            from_fund: bool = False
    ) -> float:
        """
        Calculate total expenses from already-fetched transactions.
        Handles both regular transactions and period-based transactions (pro-rated).
        """
        if end_date is None:
            end_date = date.today()
        if start_date is None:
            start_date = date(2000, 1, 1)  # Very old date to include all
            
        total_expense = 0.0
        
        for tx in transactions:
            # Filter by from_fund
            if tx.get('from_fund', False) != from_fund:
                continue
                
            # Only expenses
            if tx.get('type') != 'Expense':
                continue
            
            amount = float(tx.get('amount', 0) or 0)
            tx_date = tx.get('tx_date')
            period_start = tx.get('period_start_date')
            period_end = tx.get('period_end_date')
            
            # Convert string dates if needed
            if isinstance(tx_date, str):
                try:
                    tx_date = date.fromisoformat(tx_date.split('T')[0])
                except:
                    continue
            if isinstance(period_start, str):
                try:
                    period_start = date.fromisoformat(period_start.split('T')[0])
                except:
                    period_start = None
            if isinstance(period_end, str):
                try:
                    period_end = date.fromisoformat(period_end.split('T')[0])
                except:
                    period_end = None
            
            # Check if this is a period-based transaction
            if period_start and period_end:
                # Pro-rata calculation for period expenses
                total_days = (period_end - period_start).days + 1
                if total_days <= 0:
                    continue
                
                daily_rate = amount / total_days
                overlap_start = max(period_start, start_date)
                overlap_end = min(period_end, end_date)
                overlap_days = (overlap_end - overlap_start).days + 1
                
                if overlap_days > 0:
                    total_expense += daily_rate * overlap_days
            else:
                # Regular transaction - check if in date range
                if tx_date and start_date <= tx_date <= end_date:
                    total_expense += amount
        
        return total_expense

    def _calculate_category_expenses_from_transactions_sync(
            self,
            transactions: List[Dict],
            start_date: date | None = None,
            end_date: date | None = None,
            from_fund: bool = False
    ) -> Dict[str, float]:
        """
        Calculate expenses per category from already-fetched transactions.
        Returns {category_name: amount}
        """
        if end_date is None:
            end_date = date.today()
        if start_date is None:
            start_date = date(2000, 1, 1)
            
        category_expenses = {}
        
        for tx in transactions:
            # Filter by from_fund
            if tx.get('from_fund', False) != from_fund:
                continue
                
            # Only expenses
            if tx.get('type') != 'Expense':
                continue
            
            cat_name = tx.get('category') or REPORT_LABELS["general"]
            amount = float(tx.get('amount', 0) or 0)
            tx_date = tx.get('tx_date')
            period_start = tx.get('period_start_date')
            period_end = tx.get('period_end_date')
            
            # Convert string dates if needed
            if isinstance(tx_date, str):
                try:
                    tx_date = date.fromisoformat(tx_date.split('T')[0])
                except:
                    continue
            if isinstance(period_start, str):
                try:
                    period_start = date.fromisoformat(period_start.split('T')[0])
                except:
                    period_start = None
            if isinstance(period_end, str):
                try:
                    period_end = date.fromisoformat(period_end.split('T')[0])
                except:
                    period_end = None
            
            # Check if this is a period-based transaction
            if period_start and period_end:
                # Pro-rata calculation
                total_days = (period_end - period_start).days + 1
                if total_days <= 0:
                    continue
                
                daily_rate = amount / total_days
                overlap_start = max(period_start, start_date)
                overlap_end = min(period_end, end_date)
                overlap_days = (overlap_end - overlap_start).days + 1
                
                if overlap_days > 0:
                    prorated_amount = daily_rate * overlap_days
                    category_expenses[cat_name] = category_expenses.get(cat_name, 0.0) + prorated_amount
            else:
                # Regular transaction - check if in date range
                if tx_date and start_date <= tx_date <= end_date:
                    category_expenses[cat_name] = category_expenses.get(cat_name, 0.0) + amount
        
        return category_expenses

    def _calculate_budgets_from_transactions_sync(
            self,
            budgets: List[Any],
            transactions: List[Dict],
            start_date: date | None = None,
            end_date: date | None = None
    ) -> List[Dict[str, Any]]:
        """
        Calculate budget spending from already-fetched transactions.
        This replaces get_project_budgets_for_period queries with in-memory calculation.
        """
        if end_date is None:
            end_date = date.today()
        if start_date is None:
            start_date = date(end_date.year, 1, 1)
        
        result = []
        
        # Group transactions by category for efficient lookup
        expenses_by_category = {}
        for tx in transactions:
            if tx.get('from_fund', False):
                continue
            if tx.get('type') != 'Expense':
                continue
                
            cat_name = tx.get('category') or REPORT_LABELS["general"]
            amount = float(tx.get('amount', 0) or 0)
            tx_date = tx.get('tx_date')
            period_start = tx.get('period_start_date')
            period_end = tx.get('period_end_date')
            
            # Convert string dates if needed
            if isinstance(tx_date, str):
                try:
                    tx_date = date.fromisoformat(tx_date.split('T')[0])
                except:
                    continue
            if isinstance(period_start, str):
                try:
                    period_start = date.fromisoformat(period_start.split('T')[0])
                except:
                    period_start = None
            if isinstance(period_end, str):
                try:
                    period_end = date.fromisoformat(period_end.split('T')[0])
                except:
                    period_end = None
            
            # Calculate amount to attribute to this category
            if period_start and period_end:
                # Pro-rata calculation
                total_days = (period_end - period_start).days + 1
                if total_days <= 0:
                    continue
                
                daily_rate = amount / total_days
                overlap_start = max(period_start, start_date)
                overlap_end = min(period_end, end_date)
                overlap_days = (overlap_end - overlap_start).days + 1
                
                if overlap_days > 0:
                    amount = daily_rate * overlap_days
                else:
                    continue
            else:
                # Regular transaction - check if in date range
                if not tx_date or not (start_date <= tx_date <= end_date):
                    continue
            
            if cat_name not in expenses_by_category:
                expenses_by_category[cat_name] = 0.0
            expenses_by_category[cat_name] += amount
        
        # Now calculate budget data for each budget
        for budget in budgets:
            # Get budget category name
            if hasattr(budget, 'category'):
                budget_category = budget.category
            elif isinstance(budget, dict):
                budget_category = budget.get('category')
            else:
                continue
            
            if not budget_category:
                continue
            
            # Get expenses for this category
            total_expenses = expenses_by_category.get(budget_category, 0.0)
            
            # Get budget amount
            if hasattr(budget, 'amount'):
                base_amount = float(budget.amount or 0)
            elif isinstance(budget, dict):
                base_amount = float(budget.get('amount', 0) or 0)
            else:
                base_amount = 0.0
            
            # Get period type
            if hasattr(budget, 'period_type'):
                period_type = budget.period_type
            elif isinstance(budget, dict):
                period_type = budget.get('period_type', 'Annual')
            else:
                period_type = 'Annual'
            
            # Calculate budget amount for the period (pro-rata if needed)
            budget_amount = base_amount
            
            if period_type == "Annual":
                # Get budget dates
                if hasattr(budget, 'start_date'):
                    budget_start = budget.start_date
                    budget_end = budget.end_date
                elif isinstance(budget, dict):
                    budget_start = budget.get('start_date')
                    budget_end = budget.get('end_date')
                else:
                    budget_start = None
                    budget_end = None
                
                if budget_start and budget_end:
                    budget_total_days = (budget_end - budget_start).days + 1
                    period_days = (end_date - start_date).days + 1
                    if budget_total_days > 0:
                        budget_amount = base_amount * (period_days / budget_total_days)
            elif period_type == "Monthly":
                # Calculate number of months in the period
                months_in_period = ((end_date.year - start_date.year) * 12 + 
                                   (end_date.month - start_date.month) + 1)
                budget_amount = base_amount * months_in_period
            
            remaining_amount = budget_amount - total_expenses
            spent_percentage = (total_expenses / budget_amount * 100) if budget_amount > 0 else 0
            
            # Get budget id
            if hasattr(budget, 'id'):
                budget_id = budget.id
            elif isinstance(budget, dict):
                budget_id = budget.get('id')
            else:
                budget_id = None
            
            # Get project_id
            if hasattr(budget, 'project_id'):
                project_id = budget.project_id
            elif isinstance(budget, dict):
                project_id = budget.get('project_id')
            else:
                project_id = None
            
            result.append({
                "id": budget_id,
                "project_id": project_id,
                "category": budget_category,
                "amount": round(budget_amount, 2),
                "base_amount": base_amount,
                "period_type": period_type,
                "spent_amount": round(total_expenses, 2),
                "remaining_amount": round(remaining_amount, 2),
                "spent_percentage": round(spent_percentage, 2),
                "is_over_budget": total_expenses > budget_amount,
                "period_start": start_date.isoformat(),
                "period_end": end_date.isoformat()
            })
        
        return result

    def _is_date_in_range(self, tx_date, start_date: date, end_date: date) -> bool:
        """Helper to check if a transaction date is within range"""
        if tx_date is None:
            return False
        if isinstance(tx_date, str):
            try:
                tx_date = date.fromisoformat(tx_date.split('T')[0])
            except:
                return False
        return start_date <= tx_date <= end_date

    def _is_date_before(self, tx_date, target_date: date) -> bool:
        """Helper to check if a transaction date is before a target date"""
        if tx_date is None:
            return False
        if isinstance(tx_date, str):
            try:
                tx_date = date.fromisoformat(tx_date.split('T')[0])
            except:
                return False
        return tx_date < target_date

    def _check_budget_alerts_from_transactions_sync(
            self,
            budgets: List[Any],
            transactions: List[Dict],
            as_of_date: date
    ) -> List[Dict[str, Any]]:
        """
        Check for budget alerts using already-fetched transactions.
        This replaces check_category_budget_alerts DB queries with in-memory calculation.
        """
        alerts = []
        
        # Group expenses by category
        expenses_by_category = {}
        for tx in transactions:
            if tx.get('from_fund', False):
                continue
            if tx.get('type') != 'Expense':
                continue
            
            cat_name = tx.get('category') or REPORT_LABELS["general"]
            amount = float(tx.get('amount', 0) or 0)
            expenses_by_category[cat_name] = expenses_by_category.get(cat_name, 0.0) + amount
        
        for budget in budgets:
            # Get budget info
            if hasattr(budget, 'category'):
                budget_category = budget.category
                budget_amount = float(budget.amount or 0)
                budget_id = budget.id
                project_id = budget.project_id
                budget_start = budget.start_date
                budget_end = budget.end_date
                period_type = budget.period_type
            else:
                continue
            
            if not budget_category or budget_amount <= 0:
                continue
            
            # Get spending for this category
            spent_amount = expenses_by_category.get(budget_category, 0.0)
            
            # Calculate expected spending percentage based on time elapsed
            expected_spent_percentage = 0
            if period_type == "Annual" and budget_end:
                total_days = (budget_end - budget_start).days + 1
                days_elapsed = max(0, (as_of_date - budget_start).days + 1)
                if total_days > 0:
                    expected_spent_percentage = min((days_elapsed / total_days) * 100, 100)
            elif period_type == "Monthly":
                total_days = 30
                days_elapsed = max(0, (as_of_date - budget_start).days + 1)
                if total_days > 0:
                    expected_spent_percentage = min((days_elapsed / total_days) * 100, 100)
            
            # Calculate actual spent percentage
            spent_percentage = (spent_amount / budget_amount * 100) if budget_amount > 0 else 0
            
            # Check alerts
            is_over_budget = spent_amount > budget_amount
            is_spending_too_fast = spent_percentage > (expected_spent_percentage + 10)
            
            if is_over_budget or is_spending_too_fast:
                alerts.append({
                    "project_id": project_id,
                    "budget_id": budget_id,
                    "category": budget_category,
                    "amount": budget_amount,
                    "spent_amount": spent_amount,
                    "spent_percentage": round(spent_percentage, 2),
                    "expected_spent_percentage": round(expected_spent_percentage, 2),
                    "is_over_budget": is_over_budget,
                    "is_spending_too_fast": is_spending_too_fast,
                    "alert_type": "over_budget" if is_over_budget else "spending_too_fast"
                })
        
        return alerts

    # ==================== END OPTIMIZED HELPER FUNCTIONS ====================

    async def get_dashboard_snapshot(self) -> Dict[str, Any]:
        """Get comprehensive dashboard snapshot with real-time financial data
        
        OPTIMIZED: Fetches all transactions in ONE query and calculates everything in memory.
        Before optimization: ~70+ queries for 10 projects
        After optimization: ~5 queries total regardless of project count
        """
        from sqlalchemy.orm import selectinload
        from backend.models.fund import Fund
        from backend.models.budget import Budget
        
        # Rollback any failed transaction before starting
        try:
            await self.db.rollback()
        except Exception:
            pass  # Ignore if there's no transaction to rollback

        # Get all active projects
        projects_query = select(Project).where(Project.is_active == True)
        projects_result = await self.db.execute(projects_query)
        projects = list(projects_result.scalars().all())
        print(f"📋 Found {len(projects)} active projects")

        if not projects:
            return {
                "projects": [],
                "alerts": {
                    "budget_overrun": [],
                    "budget_warning": [],
                    "missing_proof": [],
                    "unpaid_recurring": [],
                    "negative_fund_balance": [],
                    "category_budget_alerts": []
                },
                "summary": {
                    "total_income": 0,
                    "total_expense": 0,
                    "total_profit": 0
                },
                "expense_categories": []
            }

        # Get current date
        current_date = date.today()
        
        # Pre-load ALL project data immediately to avoid lazy loading issues
        projects_data = []
        project_ids = []
        for project in projects:
            try:
                project_dict = {
                    "id": project.id,
                    "name": project.name,
                    "description": project.description,
                    "start_date": project.start_date.isoformat() if project.start_date else None,
                    "end_date": project.end_date.isoformat() if project.end_date else None,
                    "budget_monthly": project.budget_monthly,
                    "budget_annual": project.budget_annual,
                    "num_residents": project.num_residents,
                    "monthly_price_per_apartment": project.monthly_price_per_apartment,
                    "address": project.address,
                    "city": project.city,
                    "relation_project": project.relation_project,
                    "is_parent_project": project.is_parent_project,
                    "image_url": project.image_url,
                    "is_active": project.is_active,
                    "manager_id": project.manager_id,
                    "created_at": project.created_at
                }
                projects_data.append(project_dict)
                project_ids.append(project.id)
            except Exception as e:
                print(f"אזהרה: שגיאה בטעינת נתוני פרויקט: {e}")
                continue

        # ==================== OPTIMIZED: BATCH QUERIES ====================
        # Query 1: Fetch ALL transactions for ALL projects at once (single query!)
        all_transactions_query = select(Transaction).options(
            selectinload(Transaction.category)
        ).where(
            Transaction.project_id.in_(project_ids)
        ).order_by(Transaction.tx_date.desc())
        
        all_transactions_result = await self.db.execute(all_transactions_query)
        all_transaction_objects = list(all_transactions_result.scalars().all())
        
        # Convert to dicts and group by project_id
        transactions_by_project = {}
        all_transactions_list = []  # For global calculations
        
        for tx in all_transaction_objects:
            tx_dict = {
                "id": tx.id,
                "project_id": tx.project_id,
                "tx_date": tx.tx_date,
                "type": tx.type,
                "amount": float(tx.amount) if tx.amount else 0.0,
                "category": tx.category.name if tx.category else None,
                "from_fund": getattr(tx, 'from_fund', False) or False,
                "file_path": getattr(tx, 'file_path', None),
                "is_exceptional": getattr(tx, 'is_exceptional', False),
                "period_start_date": getattr(tx, 'period_start_date', None),
                "period_end_date": getattr(tx, 'period_end_date', None),
            }
            
            if tx.project_id not in transactions_by_project:
                transactions_by_project[tx.project_id] = []
            transactions_by_project[tx.project_id].append(tx_dict)
            all_transactions_list.append(tx_dict)
        
        # Query 2: Fetch ALL funds for ALL projects at once
        funds_query = select(Fund).where(Fund.project_id.in_(project_ids))
        funds_result = await self.db.execute(funds_query)
        funds_list = list(funds_result.scalars().all())
        funds_by_project = {f.project_id: f for f in funds_list}
        
        # Query 3: Fetch ALL budgets for ALL projects at once
        budgets_query = select(Budget).where(
            and_(
                Budget.project_id.in_(project_ids),
                Budget.is_active == True
            )
        )
        budgets_result = await self.db.execute(budgets_query)
        budgets_list = list(budgets_result.scalars().all())
        budgets_by_project = {}
        for b in budgets_list:
            if b.project_id not in budgets_by_project:
                budgets_by_project[b.project_id] = []
            budgets_by_project[b.project_id].append(b)
        
        # ==================== END BATCH QUERIES ====================

        # Calculate financial data for each project FROM MEMORY (no more DB queries!)
        projects_with_finance = []
        total_income = 0
        total_expense = 0
        budget_overrun_projects = []
        budget_warning_projects = []
        missing_proof_projects = []
        unpaid_recurring_projects = []
        negative_fund_balance_projects = []
        category_budget_alerts = []

        # Process each project using pre-loaded data
        for proj_data in projects_data:
            project_id = proj_data["id"]
            
            # Convert string dates to date objects if needed
            project_start_date_raw = proj_data["start_date"]
            if project_start_date_raw:
                if isinstance(project_start_date_raw, str):
                    project_start_date = date.fromisoformat(project_start_date_raw)
                elif isinstance(project_start_date_raw, date):
                    project_start_date = project_start_date_raw
                else:
                    project_start_date = None
            else:
                project_start_date = None

            # Calculate start date
            if project_start_date:
                calculation_start_date = project_start_date
            else:
                calculation_start_date = current_date - relativedelta(years=1)

            # Get transactions for this project from memory
            project_transactions = transactions_by_project.get(project_id, [])
            
            # Calculate income and expenses FROM MEMORY using optimized helper
            yearly_income = 0.0
            yearly_expense = 0.0
            
            for tx in project_transactions:
                # Skip fund transactions
                if tx.get('from_fund', False):
                    continue
                
                tx_date = tx.get('tx_date')
                period_start = tx.get('period_start_date')
                period_end = tx.get('period_end_date')
                amount = float(tx.get('amount', 0) or 0)
                
                # Convert string dates if needed
                if isinstance(tx_date, str):
                    try:
                        tx_date = date.fromisoformat(tx_date.split('T')[0])
                    except:
                        continue
                
                if isinstance(period_start, str):
                    try:
                        period_start = date.fromisoformat(period_start.split('T')[0])
                    except:
                        period_start = None
                if isinstance(period_end, str):
                    try:
                        period_end = date.fromisoformat(period_end.split('T')[0])
                    except:
                        period_end = None
                
                # Check date range
                in_range = False
                calculated_amount = amount
                
                if period_start and period_end:
                    # Pro-rata calculation for period transactions
                    total_days = (period_end - period_start).days + 1
                    if total_days > 0:
                        daily_rate = amount / total_days
                        overlap_start = max(period_start, calculation_start_date)
                        overlap_end = min(period_end, current_date)
                        overlap_days = (overlap_end - overlap_start).days + 1
                        if overlap_days > 0:
                            calculated_amount = daily_rate * overlap_days
                            in_range = True
                else:
                    # Regular transaction
                    if tx_date and calculation_start_date <= tx_date <= current_date:
                        in_range = True
                
                if in_range:
                    if tx.get('type') == 'Income':
                        yearly_income += calculated_amount
                    elif tx.get('type') == 'Expense':
                        yearly_expense += calculated_amount

            # Budget calculations (from memory)
            try:
                budget_annual = float(proj_data["budget_annual"] if proj_data["budget_annual"] is not None else 0)
                budget_monthly = float(proj_data["budget_monthly"] if proj_data["budget_monthly"] is not None else 0)
            except (AttributeError, ValueError):
                budget_annual = 0.0
                budget_monthly = 0.0

            # Calculate income from the monthly budget
            project_income = 0.0
            monthly_income = float(proj_data["budget_monthly"] or 0)
            if monthly_income > 0:
                if proj_data["start_date"]:
                    start_date_val = proj_data["start_date"]
                    if isinstance(start_date_val, str):
                        income_calculation_start = date.fromisoformat(start_date_val)
                    elif isinstance(start_date_val, date):
                        income_calculation_start = start_date_val
                    else:
                        income_calculation_start = calculation_start_date
                elif proj_data.get("created_at"):
                    created_at_val = proj_data["created_at"]
                    if hasattr(created_at_val, 'date'):
                        income_calculation_start = created_at_val.date()
                    elif isinstance(created_at_val, str):
                        income_calculation_start = date.fromisoformat(created_at_val.split('T')[0])
                    elif isinstance(created_at_val, date):
                        income_calculation_start = created_at_val
                    else:
                        income_calculation_start = calculation_start_date
                else:
                    income_calculation_start = calculation_start_date
                project_income = calculate_monthly_income_amount(monthly_income, income_calculation_start, current_date)
                yearly_income = 0.0

            project_total_income = yearly_income + project_income
            profit = project_total_income - yearly_expense

            # Calculate profit percentage
            if project_total_income > 0:
                profit_percent = (profit / project_total_income * 100)
            else:
                profit_percent = 0

            # Determine status color
            if profit_percent >= 10:
                status_color = "green"
            elif profit_percent <= -10:
                status_color = "red"
            else:
                status_color = "yellow"

            # Budget overrun check (from memory)
            yearly_budget = 0.0
            if budget_monthly > 0:
                if proj_data["start_date"]:
                    start_date_val = proj_data["start_date"]
                    if isinstance(start_date_val, str):
                        start_date_parsed = date.fromisoformat(start_date_val)
                    elif isinstance(start_date_val, date):
                        start_date_parsed = start_date_val
                    else:
                        start_date_parsed = calculation_start_date
                    start_month = date(start_date_parsed.year, start_date_parsed.month, 1)
                else:
                    start_month = date(calculation_start_date.year, calculation_start_date.month, 1)
                end_month = date(current_date.year, current_date.month, 1)
                month_count = 0
                temp_month = start_month
                while temp_month <= end_month:
                    month_count += 1
                    if temp_month.month == 12:
                        temp_month = date(temp_month.year + 1, 1, 1)
                    else:
                        temp_month = date(temp_month.year, temp_month.month + 1, 1)
                yearly_budget = budget_monthly * month_count
            elif budget_annual > 0:
                days_in_period = (current_date - calculation_start_date).days + 1
                days_in_year = 365
                yearly_budget = (budget_annual / days_in_year) * days_in_period
            
            if yearly_budget > 0:
                budget_percentage = (yearly_expense / yearly_budget) * 100
                if yearly_expense > yearly_budget:
                    budget_overrun_projects.append(project_id)
                elif budget_percentage >= 70:
                    budget_warning_projects.append(project_id)

            # Check alerts FROM MEMORY (no DB queries!)
            # Missing proof check
            missing_proof_count = sum(
                1 for tx in project_transactions
                if not tx.get('from_fund', False)
                and not tx.get('file_path')
                and self._is_date_in_range(tx.get('tx_date'), calculation_start_date, current_date)
            )
            if missing_proof_count > 0:
                missing_proof_projects.append(project_id)

            # Unpaid recurring check
            unpaid_recurring_count = sum(
                1 for tx in project_transactions
                if not tx.get('from_fund', False)
                and tx.get('type') == 'Expense'
                and not tx.get('is_exceptional', False)
                and not tx.get('file_path')
                and self._is_date_before(tx.get('tx_date'), current_date)
            )
            if unpaid_recurring_count > 0:
                unpaid_recurring_projects.append(project_id)

            # Category budget alerts FROM MEMORY
            project_budgets = budgets_by_project.get(project_id, [])
            if project_budgets:
                budget_alerts = self._check_budget_alerts_from_transactions_sync(
                    project_budgets,
                    project_transactions,
                    current_date
                )
                category_budget_alerts.extend(budget_alerts)

            # Negative fund balance check FROM MEMORY
            fund = funds_by_project.get(project_id)
            if fund and float(fund.current_balance) < 0:
                negative_fund_balance_projects.append(project_id)

            # Build project data
            start_date_str = proj_data["start_date"] if isinstance(proj_data["start_date"], str) else (proj_data["start_date"].isoformat() if proj_data["start_date"] else None)
            end_date_str = proj_data["end_date"] if isinstance(proj_data["end_date"], str) else (proj_data["end_date"].isoformat() if proj_data["end_date"] else None)
            created_at_str = proj_data["created_at"] if isinstance(proj_data["created_at"], str) else (proj_data["created_at"].isoformat() if proj_data["created_at"] else None)
            
            project_data = {
                "id": project_id,
                "name": proj_data["name"],
                "description": proj_data["description"],
                "start_date": start_date_str,
                "end_date": end_date_str,
                "budget_monthly": float(proj_data["budget_monthly"] or 0),
                "budget_annual": float(proj_data["budget_annual"] or 0),
                "num_residents": proj_data["num_residents"],
                "monthly_price_per_apartment": float(proj_data["monthly_price_per_apartment"] or 0),
                "address": proj_data["address"],
                "city": proj_data["city"],
                "relation_project": proj_data["relation_project"],
                "is_parent_project": proj_data["is_parent_project"],
                "image_url": proj_data["image_url"],
                "is_active": proj_data["is_active"],
                "manager_id": proj_data["manager_id"],
                "created_at": created_at_str,
                "income_month_to_date": project_total_income,
                "expense_month_to_date": yearly_expense,
                "profit_percent": round(profit_percent, 1),
                "status_color": status_color,
                "children": []
            }

            projects_with_finance.append(project_data)
            total_income += project_total_income
            total_expense += yearly_expense

        # Build project hierarchy
        project_map = {p["id"]: p for p in projects_with_finance}
        root_projects = []

        for project_data in projects_with_finance:
            if project_data["relation_project"] and project_data["relation_project"] in project_map:
                parent = project_map[project_data["relation_project"]]
                parent["children"].append(project_data)
            else:
                root_projects.append(project_data)

        # Calculate total profit
        total_profit = total_income - total_expense

        # Calculate expense categories FROM MEMORY
        earliest_start = date.today() - relativedelta(years=1)
        for proj_data in projects_data:
            proj_start_date_raw = proj_data["start_date"]
            if proj_start_date_raw:
                if isinstance(proj_start_date_raw, str):
                    proj_start_date = date.fromisoformat(proj_start_date_raw)
                elif isinstance(proj_start_date_raw, date):
                    proj_start_date = proj_start_date_raw
                else:
                    proj_start_date = None
            else:
                proj_start_date = None
            project_start = calculate_start_date(proj_start_date)
            if project_start < earliest_start:
                earliest_start = project_start

        # Calculate expense categories FROM MEMORY using optimized helper
        cat_expenses_map = self._calculate_category_expenses_from_transactions_sync(
            all_transactions_list,
            earliest_start,
            current_date,
            from_fund=False
        )
        
        expense_categories = []
        for cat_name, amount in cat_expenses_map.items():
            if amount > 0:
                expense_categories.append({
                    "category": cat_name,
                    "amount": amount,
                    "color": self._get_category_color(cat_name)
                })

        return {
            "projects": projects_with_finance,
            "alerts": {
                "budget_overrun": budget_overrun_projects,
                "budget_warning": budget_warning_projects,
                "missing_proof": missing_proof_projects,
                "unpaid_recurring": unpaid_recurring_projects,
                "negative_fund_balance": negative_fund_balance_projects,
                "category_budget_alerts": category_budget_alerts
            },
            "summary": {
                "total_income": round(total_income, 2),
                "total_expense": round(total_expense, 2),
                "total_profit": round(total_profit, 2)
            },
            "expense_categories": expense_categories
        }

    import hashlib

    def _get_category_color(self, category: str) -> str:
        """Get random-like, but consistent color for any category"""
        hash_object = hashlib.md5(category.encode())
        hex_color = "#" + hash_object.hexdigest()[:6]
        return hex_color

    async def get_project_expense_categories(self, project_id: int) -> List[Dict[str, Any]]:
        """Get expense categories breakdown for a specific project"""
        # Calculate for all time (wide range)
        start_date = date(2000, 1, 1)
        end_date = date(2100, 1, 1)

        cat_expenses_map = await self._calculate_category_expenses_with_period(
            project_id,
            start_date,
            end_date,
            from_fund=False
        )

        expense_categories = []
        for cat_name, amount in cat_expenses_map.items():
            if amount > 0:
                expense_categories.append({
                    "category": cat_name,
                    "amount": amount,
                    "color": self._get_category_color(cat_name)
                })

        return expense_categories

    async def get_project_transactions(self, project_id: int) -> List[Dict[str, Any]]:
        """Get all transactions for a specific project (including recurring ones)"""
        transactions_query = select(Transaction).options(
            selectinload(Transaction.category)
        ).where(Transaction.project_id == project_id).order_by(Transaction.tx_date.desc())
        transactions_result = await self.db.execute(transactions_query)
        transactions = list(transactions_result.scalars().all())

        return [
            {
                "id": tx.id,
                "project_id": tx.project_id,
                "tx_date": tx.tx_date.isoformat(),
                "type": tx.type,
                "amount": float(tx.amount),
                "description": tx.description,
                "category": tx.category.name if tx.category else None,
                "notes": tx.notes,
                "is_exceptional": tx.is_exceptional,
                "is_generated": getattr(tx, 'is_generated', False),
                "recurring_template_id": getattr(tx, 'recurring_template_id', None),
                "created_at": tx.created_at.isoformat() if hasattr(tx, 'created_at') and tx.created_at else None,
                "period_start_date": tx.period_start_date.isoformat() if hasattr(tx,
                                                                                 'period_start_date') and tx.period_start_date else None,
                "period_end_date": tx.period_end_date.isoformat() if hasattr(tx,
                                                                             'period_end_date') and tx.period_end_date else None
            }
            for tx in transactions
        ]

    async def get_expenses_by_transaction_date(
            self,
            project_id: int | None = None,
            start_date: date | None = None,
            end_date: date | None = None
    ) -> Dict[str, Any]:
        """
        Get expenses aggregated by transaction date for dashboard.
        Shows expenses related to specific transaction dates with aggregation.
        Handles both regular transactions (by tx_date) and period transactions (split by month).
        """
        from sqlalchemy import or_
        from collections import defaultdict
        from datetime import timedelta
        
        # 1. Regular expenses (no period dates) - filter by tx_date
        regular_conditions = [
            Transaction.type == 'Expense',
            Transaction.from_fund == False,
            or_(
                Transaction.period_start_date.is_(None),
                Transaction.period_end_date.is_(None)
            )
        ]
        
        if project_id:
            regular_conditions.append(Transaction.project_id == project_id)
        if start_date:
            regular_conditions.append(Transaction.tx_date >= start_date)
        if end_date:
            regular_conditions.append(Transaction.tx_date <= end_date)
        
        regular_query = select(
            Transaction.tx_date,
            func.sum(Transaction.amount).label('total_expense'),
            func.count(Transaction.id).label('transaction_count')
        ).where(
            and_(*regular_conditions)
        ).group_by(Transaction.tx_date).order_by(Transaction.tx_date.desc())
        
        regular_result = await self.db.execute(regular_query)
        regular_rows = regular_result.all()
        
        # Aggregate expenses by date
        expenses_by_date_dict = defaultdict(lambda: {'expense': 0.0, 'transaction_count': 0})
        
        for row in regular_rows:
            date_key = row.tx_date.isoformat()
            expenses_by_date_dict[date_key]['expense'] += float(row.total_expense)
            expenses_by_date_dict[date_key]['transaction_count'] += row.transaction_count
        
        # 2. Period expenses - split by month
        period_conditions = [
            Transaction.type == 'Expense',
            Transaction.from_fund == False,
            Transaction.period_start_date.is_not(None),
            Transaction.period_end_date.is_not(None)
        ]
        
        if project_id:
            period_conditions.append(Transaction.project_id == project_id)
        if start_date or end_date:
            if start_date and end_date:
                period_conditions.append(
                    and_(
                        Transaction.period_start_date <= end_date,
                        Transaction.period_end_date >= start_date
                    )
                )
            elif start_date:
                period_conditions.append(Transaction.period_end_date >= start_date)
            elif end_date:
                period_conditions.append(Transaction.period_start_date <= end_date)
        
        period_query = select(Transaction).where(and_(*period_conditions))
        period_txs = (await self.db.execute(period_query)).scalars().all()
        
        for tx in period_txs:
            # Calculate the effective date range for this period transaction
            effective_start = max(tx.period_start_date, start_date if start_date else tx.period_start_date)
            effective_end = min(tx.period_end_date, end_date if end_date else tx.period_end_date)
            
            if effective_start > effective_end:
                continue
            
            total_days = (tx.period_end_date - tx.period_start_date).days + 1
            if total_days <= 0:
                continue
            
            daily_rate = float(tx.amount) / total_days
            
            # Split period transaction by month
            current_date = effective_start
            while current_date <= effective_end:
                # Calculate month end
                if current_date.month == 12:
                    month_end = date(current_date.year + 1, 1, 1)
                else:
                    month_end = date(current_date.year, current_date.month + 1, 1)
                
                # The period end in this month
                period_end_in_month = min(effective_end, month_end - timedelta(days=1))
                days_in_month = (period_end_in_month - current_date).days + 1
                
                if days_in_month > 0:
                    amount = daily_rate * days_in_month
                    # Use first day of month as the date key for aggregation
                    month_start = date(current_date.year, current_date.month, 1)
                    date_key = month_start.isoformat()
                    expenses_by_date_dict[date_key]['expense'] += amount
                    expenses_by_date_dict[date_key]['transaction_count'] += 1
                
                # Move to next month
                current_date = month_end
        
        # Convert to list and sort
        expenses_by_date = []
        total_expense = 0.0
        total_count = 0
        
        for date_key, data in sorted(expenses_by_date_dict.items(), reverse=True):
            expenses_by_date.append({
                'date': date_key,
                'expense': data['expense'],
                'transaction_count': data['transaction_count']
            })
            total_expense += data['expense']
            total_count += data['transaction_count']

        return {
            'expenses_by_date': expenses_by_date,
            'total_expense': total_expense,
            'total_transaction_count': total_count,
            'period_start': start_date.isoformat() if start_date else None,
            'period_end': end_date.isoformat() if end_date else None
        }

    async def generate_custom_report(self, options, chart_images: Dict[str, bytes] = None) -> bytes:
        """Generate a custom report (PDF, Excel, or ZIP) based on options"""
        # options is expected to be ReportOptions instance, but using dynamic typing to avoid circular import at module level
        from backend.schemas.report import ReportOptions
        from sqlalchemy.orm import selectinload
        from backend.repositories.budget_repository import BudgetRepository

        # 1. Fetch data based on options
        project_id = options.project_id

        # Fetch basic project info
        proj = (await self.db.execute(select(Project).where(Project.id == project_id))).scalar_one()

        # --- OPTIMIZED: Fetch all transactions at once (single query) ---
        transactions = []
        query = select(Transaction).options(
            selectinload(Transaction.category),
            selectinload(Transaction.supplier),
            selectinload(Transaction.project)
        ).where(Transaction.project_id == project_id)
        
        # Apply date filters at DB level for efficiency
        if options.start_date:
            query = query.where(Transaction.tx_date >= options.start_date)
        if options.end_date:
            query = query.where(Transaction.tx_date <= options.end_date)
        if options.transaction_types:
            query = query.where(Transaction.type.in_(options.transaction_types))
        if options.only_recurring:
            query = query.where(Transaction.recurring_template_id.isnot(None))

        # Filter by Categories (list of category names)
        if options.categories and len(options.categories) > 0:
            query = query.join(Category, Transaction.category_id == Category.id).where(
                Category.name.in_(options.categories))

        # Filter by Suppliers (list of supplier IDs)
        if options.suppliers and len(options.suppliers) > 0:
            query = query.where(Transaction.supplier_id.in_(options.suppliers))

        query = query.order_by(Transaction.tx_date.desc())
        result = await self.db.execute(query)
        transaction_objects = list(result.scalars().all())

        # Convert to dictionaries IMMEDIATELY while session is active to avoid lazy loading issues
        for tx in transaction_objects:
            tx_dict = {
                "id": tx.id,
                "project_id": tx.project_id,
                "tx_date": tx.tx_date,
                "type": tx.type,
                "amount": float(tx.amount),
                "description": tx.description,
                "category": tx.category.name if tx.category else None,
                "category_obj": tx.category,  # Keep object reference if needed
                "notes": tx.notes,
                "is_exceptional": tx.is_exceptional,
                "is_generated": getattr(tx, 'is_generated', False),
                "recurring_template_id": getattr(tx, 'recurring_template_id', None),
                "created_at": getattr(tx, 'created_at', None),
                "period_start_date": getattr(tx, 'period_start_date', None),
                "period_end_date": getattr(tx, 'period_end_date', None),
                "file_path": getattr(tx, 'file_path', None),
                "payment_method": getattr(tx, 'payment_method', None),
                "from_fund": getattr(tx, 'from_fund', False),
                "project_name": tx.project.name if tx.project else "",
                "supplier_name": tx.supplier.name if tx.supplier else None,
            }
            transactions.append(tx_dict)

        # --- OPTIMIZED: Budgets - Calculate from already fetched transactions ---
        budgets_data = []
        if options.include_budgets:
            # Fetch budgets list (single query) - no spending calculation yet
            budget_repo = BudgetRepository(self.db)
            budgets_list = await budget_repo.get_active_budgets_for_project(project_id)
            
            # Calculate spending from transactions in memory (no additional queries!)
            budgets_data = self._calculate_budgets_from_transactions_sync(
                budgets_list,
                transactions,
                options.start_date,
                options.end_date
            )

        # --- Funds (single query) ---
        fund_data = None
        if options.include_funds:
            fund_service = FundService(self.db)
            fund_data = await fund_service.get_fund_by_project(project_id)

        # --- OPTIMIZED: Summary - Calculate from already fetched transactions ---
        summary_data = {}
        if options.include_summary:
            # Calculate summary from transactions in memory (no additional queries!)
            summary_data = self._calculate_summary_from_transactions_sync(
                transactions,
                project=proj,
                fund=fund_data
            )
            # Add project_id to match expected format
            summary_data["project_id"] = project_id
        
        # If transactions not requested but we needed them for calculations, clear them
        if not options.include_transactions:
            transactions = []

        # --- Monthly Breakdown Data ---
        monthly_breakdown = []
        # Always calculate monthly breakdown if we have transactions (even if include_transactions is False, we still want the breakdown)
        # Calculate monthly breakdown by category and supplier
        # If filtering by year, start from the first month of the contract in that year, not January
        start_date = options.start_date or date(2000, 1, 1)
        end_date = options.end_date or date.today()
        
        # If filtering by a specific year and project has a start_date, adjust start_date 
        # to begin from the contract's first month in that year (not January)
        if options.start_date and options.end_date and proj.start_date:
            start_year = options.start_date.year
            end_year = options.end_date.year
            # If filtering by a single year
            if start_year == end_year:
                # Use the first day of the contract start month in the filtered year
                # This works even if contract started in a different year
                contract_start_month = proj.start_date.month
                contract_start_day = proj.start_date.day
                adjusted_start = date(start_year, contract_start_month, contract_start_day)
                # If contract start month is within the filtered range, use it
                if adjusted_start >= options.start_date and adjusted_start <= options.end_date:
                    # Contract start month is within the filter range - use it as start
                    start_date = adjusted_start
                elif adjusted_start < options.start_date:
                    # Contract start month is before the filter start - use filter start
                    start_date = options.start_date
                else:
                    # Contract start month is after the filter end - use filter start
                    start_date = options.start_date
            else:
                # Multi-year filter, use provided start_date
                start_date = options.start_date
        elif not options.start_date and not options.end_date and proj.start_date:
            # No date filter, use project start_date
            start_date = proj.start_date
        
        try:
            monthly_breakdown = await self._calculate_monthly_category_supplier_expenses(
                project_id,
                start_date,
                end_date,
                from_fund=False
            )
            # Apply category filter if specified
            if options.categories and len(options.categories) > 0:
                monthly_breakdown = [row for row in monthly_breakdown if row['category'] in options.categories]
            # Apply supplier filter if specified
            if options.suppliers and len(options.suppliers) > 0:
                # Need to get supplier names from IDs
                supplier_query = select(Supplier).where(Supplier.id.in_(options.suppliers))
                supplier_result = await self.db.execute(supplier_query)
                supplier_names = {s.name for s in supplier_result.scalars().all()}
                monthly_breakdown = [row for row in monthly_breakdown if row['supplier'] in supplier_names]
            print(f"INFO: Monthly breakdown calculated: {len(monthly_breakdown)} rows")
        except Exception as e:
            print(f"WARNING: Error calculating monthly breakdown: {e}")
            import traceback
            traceback.print_exc()
            monthly_breakdown = []

        # 2. Generate Output
        if options.format == "pdf":
            return await self._generate_pdf(proj, options, transactions, budgets_data, fund_data, summary_data,
                                            chart_images, monthly_breakdown)
        elif options.format == "excel":
            return await self._generate_excel(proj, options, transactions, budgets_data, fund_data, summary_data,
                                              chart_images, monthly_breakdown)
        elif options.format == "zip":
            # For ZIP, we generate the PDF/Excel report AND include documents
            report_content = await self._generate_excel(proj, options, transactions, budgets_data, fund_data,
                                                        summary_data, chart_images, monthly_breakdown)
            return await self._generate_zip(proj, options, report_content, transactions)

        raise ValueError("Invalid format")

    async def generate_supplier_report(self, options, chart_images: Dict[str, bytes] = None) -> bytes:
        """Generate a report for a specific supplier with all their transactions"""
        from backend.schemas.report import SupplierReportOptions
        from backend.models.supplier import Supplier
        from sqlalchemy.orm import selectinload

        # 1. Fetch supplier info
        supplier = (
            await self.db.execute(select(Supplier).where(Supplier.id == options.supplier_id))).scalar_one_or_none()
        if not supplier:
            raise ValueError(f"ספק עם מזהה {options.supplier_id} לא נמצא")

        # 2. Fetch transactions for this supplier
        transactions = []
        if options.include_transactions:
            query = select(Transaction).options(
                selectinload(Transaction.category),
                selectinload(Transaction.project),
                selectinload(Transaction.supplier)
            ).where(Transaction.supplier_id == options.supplier_id)

            if options.start_date:
                query = query.where(Transaction.tx_date >= options.start_date)
            if options.end_date:
                query = query.where(Transaction.tx_date <= options.end_date)
            if options.transaction_types:
                query = query.where(Transaction.type.in_(options.transaction_types))
            if options.only_recurring:
                query = query.where(Transaction.recurring_template_id.isnot(None))

            # Filter by Categories
            if options.categories and len(options.categories) > 0:
                query = query.join(Category, Transaction.category_id == Category.id).where(
                    Category.name.in_(options.categories))

            # Filter by Projects
            if options.project_ids and len(options.project_ids) > 0:
                query = query.where(Transaction.project_id.in_(options.project_ids))

            query = query.order_by(Transaction.tx_date.desc())
            result = await self.db.execute(query)
            transaction_objects = list(result.scalars().all())

            # Convert to dictionaries IMMEDIATELY while session is active to avoid lazy loading issues
            transactions = []
            for tx in transaction_objects:
                tx_dict = {
                    "id": tx.id,
                    "project_id": tx.project_id,
                    "tx_date": tx.tx_date,
                    "type": tx.type,
                    "amount": float(tx.amount),
                    "description": tx.description,
                    "category": tx.category.name if tx.category else None,
                    "notes": tx.notes,
                    "is_exceptional": tx.is_exceptional,
                    "is_generated": getattr(tx, 'is_generated', False),
                    "recurring_template_id": getattr(tx, 'recurring_template_id', None),
                    "created_at": getattr(tx, 'created_at', None),
                    "period_start_date": getattr(tx, 'period_start_date', None),
                    "period_end_date": getattr(tx, 'period_end_date', None),
                    "file_path": getattr(tx, 'file_path', None),
                    "payment_method": getattr(tx, 'payment_method', None),
                    "project_name": tx.project.name if tx.project else "",
                    "supplier_name": tx.supplier.name if tx.supplier else None,
                }
                transactions.append(tx_dict)

        # 3. Calculate summary for supplier
        summary_data = {
            "supplier_name": supplier.name,
            "total_income": 0.0,
            "total_expenses": 0.0,
            "total_amount": 0.0,
            "transaction_count": len(transactions)
        }

        for tx in transactions:
            tx_type = tx.get('type') if isinstance(tx, dict) else tx.type
            tx_amount = tx.get('amount') if isinstance(tx, dict) else float(tx.amount)
            if tx_type == "Income":
                summary_data["total_income"] += float(tx_amount)
            else:
                summary_data["total_expenses"] += float(tx_amount)

        summary_data["total_amount"] = summary_data["total_income"] - summary_data["total_expenses"]

        # 4. Generate output
        if options.format == "pdf":
            return await self._generate_supplier_pdf(supplier, options, transactions, summary_data, chart_images)
        elif options.format == "excel":
            return await self._generate_supplier_excel(supplier, options, transactions, summary_data, chart_images)
        elif options.format == "zip":
            report_content = await self._generate_supplier_excel(supplier, options, transactions, summary_data,
                                                                 chart_images)
            return await self._generate_supplier_zip(supplier, options, report_content, transactions)

        raise ValueError("פורמט לא תקין")

    async def _generate_supplier_pdf(self, supplier, options, transactions, summary, chart_images=None) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)

        # Register Hebrew Font (same logic as _generate_pdf)
        font_name = 'Helvetica'
        try:
            services_dir = os.path.dirname(__file__)
            backend_dir = os.path.dirname(services_dir)
            project_root = os.path.dirname(backend_dir)

            possible_paths = [
                os.path.join(backend_dir, 'static', 'fonts', 'Heebo-Regular.ttf'),
                os.path.join(project_root, 'backend', 'static', 'fonts', 'Heebo-Regular.ttf'),
                '/app/backend/static/fonts/Heebo-Regular.ttf',
                'backend/static/fonts/Heebo-Regular.ttf',
            ]

            font_path = None
            for path in possible_paths:
                if os.path.exists(path):
                    font_path = os.path.abspath(path)
                    break

            font_loaded = False
            if font_path and os.path.exists(font_path):
                try:
                    # Register with proper embedding
                    hebrew_font = TTFont('Hebrew', font_path, subfontIndex=0)
                    pdfmetrics.registerFont(hebrew_font)
                    font_name = 'Hebrew'
                    font_loaded = True
                    print(f"✓ גופן עברי נרשם בהצלחה לדוח ספק: {font_path}")
                except Exception as e:
                    print(f"✗ רישום גופן לדוח ספק נכשל: {e}")
                    font_path = None
            
            # Try Windows system fonts if Heebo not found
            if not font_loaded and os.name == 'nt':
                windows_fonts_dir = r'C:\Windows\Fonts'
                windows_fonts = [
                    ['arial.ttf', 'arial.ttc', 'Arial.ttf'],
                    ['tahoma.ttf', 'Tahoma.ttf'],
                    ['arialuni.ttf', 'Arial Unicode MS.ttf'],
                ]
                for font_variants in windows_fonts:
                    for font_file in font_variants:
                        win_font_path = os.path.join(windows_fonts_dir, font_file)
                        if os.path.exists(win_font_path):
                            try:
                                hebrew_font = TTFont('Hebrew', win_font_path, subfontIndex=0)
                                pdfmetrics.registerFont(hebrew_font)
                                font_name = 'Hebrew'
                                font_loaded = True
                                break
                            except Exception:
                                continue
                    if font_loaded:
                        break
        except Exception as e:
            print(f"✗ שגיאה בטעינת גופן לדוח ספק: {e}")

        styles = getSampleStyleSheet()
        style_normal = ParagraphStyle('HebrewNormal', parent=styles['Normal'], fontName=font_name, fontSize=10,
                                      alignment=1)
        style_title = ParagraphStyle('HebrewTitle', parent=styles['Heading1'], fontName=font_name, fontSize=16,
                                     alignment=1, textColor=colors.HexColor('#1E3A8A'))
        style_h2 = ParagraphStyle('HebrewHeading2', parent=styles['Heading2'], fontName=font_name, fontSize=12,
                                  alignment=1, textColor=colors.HexColor('#1F2937'))

        elements = []

        # Use arabic-reshaper and python-bidi for proper RTL support
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            bidi_available = True
        except ImportError:
            bidi_available = False

        def format_text(text):
            if not text: return ""
            if not isinstance(text, str): text = str(text)
            if font_loaded and bidi_available:
                try:
                    reshaped_text = arabic_reshaper.reshape(text)
                    bidi_text = get_display(reshaped_text)
                    return bidi_text
                except Exception:
                    return text
            return text

        elements.append(Paragraph(format_text(f"דוח ספק: {supplier.name}"), style_title))
        elements.append(
            Paragraph(format_text(f"{REPORT_LABELS['production_date']}: {date.today().strftime('%d/%m/%Y')}"),
                      style_normal))
        elements.append(Spacer(1, 20))

        # Summary
        elements.append(Paragraph(format_text("סיכום"), style_h2))
        elements.append(Spacer(1, 10))
        data = [
            [format_text("פרט"), format_text(REPORT_LABELS['amount'])],
            [format_text("סה״כ הכנסות"), f"{summary['total_income']:,.2f} ₪"],
            [format_text("סה״כ הוצאות"), f"{summary['total_expenses']:,.2f} ₪"],
            [format_text("סה״כ עסקאות"), f"{summary['transaction_count']}"],
        ]
        t = Table(data, colWidths=[200, 150])
        t.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), font_name),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#DBEAFE')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))

        # Transactions
        if transactions:
            elements.append(Paragraph(format_text(REPORT_LABELS['transaction_details']), style_h2))
            elements.append(Spacer(1, 10))
            
            # Check which columns have data
            has_project = any(tx.project.name if tx.project else None for tx in transactions)
            has_description = any(tx.description for tx in transactions)
            
            # Build dynamic column list
            columns = ['date']  # Always include date
            if has_project:
                columns.append('project')
            columns.extend(['type', 'amount'])  # Always include type and amount
            if has_description:
                columns.append('description')
            
            col_to_label = {
                'date': REPORT_LABELS['date'],
                'project': "פרויקט",
                'type': REPORT_LABELS['type'],
                'amount': REPORT_LABELS['amount'],
                'description': REPORT_LABELS['description']
            }
            
            col_widths_map = {
                'date': 80,
                'project': 100,
                'type': 60,
                'amount': 80,
                'description': 200
            }
            
            # Adjust description width if no project
            if has_description and not has_project:
                col_widths_map['description'] = 280
            
            col_widths = [col_widths_map[col] for col in columns]
            
            # Build headers
            tx_data = [[format_text(col_to_label[col]) for col in columns]]
            
            for tx in transactions:
                tx_type = REPORT_LABELS['income'] if tx.type == "Income" else REPORT_LABELS['expense']
                tx_desc = tx.description or ""
                if len(tx_desc) > 30:
                    tx_desc = tx_desc[:27] + "..."

                project_name = tx.project.name if tx.project else ""
                if len(project_name) > 20:
                    project_name = project_name[:17] + "..."

                col_to_value = {
                    'date': format_date_hebrew(tx.tx_date),
                    'project': format_text(project_name),
                    'type': format_text(tx_type),
                    'amount': f"{tx.amount:,.2f}",
                    'description': format_text(tx_desc)
                }
                
                tx_data.append([col_to_value[col] for col in columns])

            tx_table = Table(tx_data, repeatRows=1, colWidths=col_widths)
            tx_table.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), font_name),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFEDD5')),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('PADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(tx_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.read()

    async def _generate_supplier_excel(self, supplier, options, transactions, summary, chart_images=None) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        fill_blue = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        fill_orange = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")

        # Summary Sheet
        ws = wb.create_sheet("סיכום")
        ws.sheet_view.rightToLeft = True

        ws.append([f"דוח ספק: {supplier.name}"])
        ws.append([f"{REPORT_LABELS['production_date']}: {date.today().strftime('%d/%m/%Y')}"])
        ws.append([])

        ws.append(["פרט", "סכום"])
        ws.append(["סה״כ הכנסות", summary['total_income']])
        ws.append(["סה״כ הוצאות", summary['total_expenses']])
        ws.append(["סה״כ עסקאות", summary['transaction_count']])

        for cell in ws[4]:
            cell.font = header_font
            cell.fill = fill_blue
            cell.alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

        # Transactions Sheet
        if transactions:
            ws_tx = wb.create_sheet(REPORT_LABELS['transaction_details'][:30])
            ws_tx.sheet_view.rightToLeft = True
            
            # Check which columns have data
            has_project = any(
                (tx.get('project_name') if isinstance(tx, dict) else (tx.project.name if tx.project else None))
                for tx in transactions
            )
            has_category = any(
                (tx.get('category') if isinstance(tx, dict) else (tx.category.name if tx.category else None))
                for tx in transactions
            )
            has_description = any(
                (tx.get('description') if isinstance(tx, dict) else tx.description)
                for tx in transactions
            )
            has_payment_method = any(
                (tx.get('payment_method') if isinstance(tx, dict) else tx.payment_method)
                for tx in transactions
            )
            has_notes = any(
                (tx.get('notes') if isinstance(tx, dict) else tx.notes)
                for tx in transactions
            )
            has_file = any(
                (tx.get('file_path') if isinstance(tx, dict) else tx.file_path)
                for tx in transactions
            )
            
            # Build dynamic column list
            columns = ['date']  # Always include date
            if has_project:
                columns.append('project')
            columns.extend(['type', 'amount'])  # Always include type and amount
            if has_category:
                columns.append('category')
            if has_description:
                columns.append('description')
            if has_payment_method:
                columns.append('payment_method')
            if has_notes:
                columns.append('notes')
            if has_file:
                columns.append('file')
            
            col_to_label = {
                'date': REPORT_LABELS['date'],
                'project': "פרויקט",
                'type': REPORT_LABELS['type'],
                'amount': REPORT_LABELS['amount'],
                'category': REPORT_LABELS['category'],
                'description': REPORT_LABELS['description'],
                'payment_method': REPORT_LABELS['payment_method'],
                'notes': REPORT_LABELS['notes'],
                'file': REPORT_LABELS['file']
            }
            
            col_widths = {
                'date': 12,
                'project': 20,
                'type': 10,
                'amount': 12,
                'category': 15,
                'description': 30,
                'payment_method': 15,
                'notes': 20,
                'file': 8
            }
            
            # Add headers
            headers = [col_to_label[col] for col in columns]
            ws_tx.append(headers)

            for cell in ws_tx[1]:
                cell.font = header_font
                cell.fill = fill_orange
                cell.alignment = Alignment(horizontal='center')

            for tx in transactions:
                # Transactions are now dictionaries, access fields directly
                if isinstance(tx, dict):
                    cat_name = tx.get('category') or ""
                    tx_type = REPORT_LABELS['income'] if tx.get('type') == "Income" else REPORT_LABELS['expense']
                    project_name = tx.get('project_name') or ""

                    col_to_value = {
                        'date': format_date_hebrew(tx.get('tx_date')),
                        'project': project_name,
                        'type': tx_type,
                        'amount': tx.get('amount'),
                        'category': cat_name,
                        'description': tx.get('description') or "",
                        'payment_method': translate_payment_method(tx.get('payment_method')),
                        'notes': tx.get('notes') or "",
                        'file': REPORT_LABELS['yes'] if tx.get('file_path') else REPORT_LABELS['no']
                    }
                else:
                    # Fallback for Transaction objects
                    cat_name = tx.category.name if tx.category else ""
                    tx_type = REPORT_LABELS['income'] if tx.type == "Income" else REPORT_LABELS['expense']
                    project_name = tx.project.name if tx.project else ""

                    col_to_value = {
                        'date': format_date_hebrew(tx.tx_date),
                        'project': project_name,
                        'type': tx_type,
                        'amount': tx.amount,
                        'category': cat_name,
                        'description': tx.description or "",
                        'payment_method': translate_payment_method(tx.payment_method),
                        'notes': tx.notes or "",
                        'file': REPORT_LABELS['yes'] if tx.file_path else REPORT_LABELS['no']
                    }
                
                row = [col_to_value[col] for col in columns]
                ws_tx.append(row)

            # Set column widths
            col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
            for i, col in enumerate(columns):
                if i < len(col_letters):
                    ws_tx.column_dimensions[col_letters[i]].width = col_widths[col]

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    async def _generate_supplier_zip(self, supplier, options, report_content, transactions) -> bytes:
        from backend.services.s3_service import S3Service
        try:
            s3_service = S3Service()
            has_s3 = True
        except Exception:
            has_s3 = False

        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
            ext = "xlsx"
            zf.writestr(f"supplier_{supplier.id}_report.{ext}", report_content)

            if has_s3 and options.include_transactions:
                for tx in transactions:
                    if tx.file_path:
                        try:
                            content = s3_service.get_file_content(tx.file_path)
                            if content:
                                fname = f"{tx.tx_date}_{tx.id}.{tx.file_path.split('.')[-1]}"
                                zf.writestr(f"documents/{fname}", content)
                        except Exception:
                            pass

        output.seek(0)
        return output.read()

    def _create_chart_image(self, chart_type: str, data: Dict[str, Any], summary: Dict[str, Any] = None,
                            transactions: List[Dict] = None, budgets: List[Dict] = None) -> io.BytesIO | None:
        """Create a chart image using matplotlib and return as BytesIO"""
        
        print(f"INFO: Starting chart creation for: {chart_type}")

        # Helper function to fix Hebrew text for matplotlib (RTL support)
        def fix_hebrew(text: str) -> str:
            """Reverse Hebrew text for proper display in matplotlib"""
            if not text:
                return text
            # Check if text contains Hebrew characters
            has_hebrew = any('\u0590' <= c <= '\u05FF' for c in text)
            if has_hebrew:
                # Reverse the text for RTL display
                return text[::-1]
            return text

        # הגדרת labels - already reversed for matplotlib
        labels_dict = {
            'income': fix_hebrew('הכנסות'),
            'expenses': fix_hebrew('הוצאות'),
            'general': fix_hebrew('כללי'),
            'category': fix_hebrew('קטגוריה'),
            'amount': fix_hebrew('סכום'),
            'date': fix_hebrew('תאריך')
        }
        
        fig = None
        ax = None

        try:
            # ייבוא והגדרת matplotlib
            import os
            os.environ['MPLBACKEND'] = 'Agg'
            
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
            
            # סגירת כל הגרפים הקודמים
            plt.close('all')
            
            # יצירת figure ישירות (לא דרך pyplot) - גדול יותר ומקצועי יותר
            # DPI גבוה יותר לאיכות טובה יותר ב-PDF וב-Excel
            fig = Figure(figsize=(12, 8), dpi=150, facecolor='white')
            canvas = FigureCanvas(fig)
            ax = fig.add_subplot(111)
            ax.set_facecolor('#FAFAFA')  # רקע אפור בהיר מקצועי
            
            # Professional color palette
            COLOR_INCOME = '#059669'      # Emerald-600
            COLOR_EXPENSE = '#E11D48'     # Rose-600
            COLOR_BG = '#FAFAFA'          # Light gray background
            COLOR_TEXT = '#0F172A'        # Slate-900
            COLOR_TEXT_MUTED = '#64748B'  # Slate-500
            
            print(f"INFO: Figure created successfully for {chart_type}")
            
            chart_created = False  # דגל לבדוק שגרף נוצר

            # --- תרשים עוגה: הכנסות מול הוצאות ---
            if chart_type == "income_expense_pie" and summary:
                income = float(summary.get('income', 0) or 0)
                expenses = float(summary.get('expenses', 0) or 0)

                labels = []
                sizes = []
                colors_list = []

                if income > 0:
                    labels.append(labels_dict['income'])
                    sizes.append(income)
                    colors_list.append(COLOR_INCOME)  # ירוק מקצועי
                if expenses > 0:
                    labels.append(labels_dict['expenses'])
                    sizes.append(expenses)
                    colors_list.append(COLOR_EXPENSE)  # אדום מקצועי

                if not sizes or sum(sizes) == 0:
                    # No data - return None to skip this chart
                    print(f"INFO: No data for income_expense_pie chart (income={income}, expenses={expenses})")
                    return None
                else:
                    single_value = len(sizes) == 1
                    ax.set_aspect('equal')
                    
                    # צבעים מפורשים לכל מקרה
                    if single_value:
                        pie_colors = [COLOR_INCOME] if labels[0] == labels_dict['income'] else [COLOR_EXPENSE]
                    else:
                        pie_colors = colors_list
                    
                    print(f"INFO: Creating pie chart with {len(sizes)} segments, colors: {pie_colors}")
                    
                    # יצירת עוגה מקצועית עם shadow עדין
                    wedges, texts, autotexts = ax.pie(
                        sizes, 
                        colors=pie_colors, 
                        autopct=lambda p: f'{p:.1f}%' if p > 0 and not single_value else '',
                        startangle=90, 
                        textprops={'fontsize': 16, 'weight': 'bold', 'color': 'white'},
                        shadow=True,
                        explode=[0.05] * len(sizes)  # הפרדה עדינה בין החלקים
                    )
                    
                    # עיצוב מקצועי של החלקים
                    for wedge in wedges:
                        wedge.set_edgecolor('white')
                        wedge.set_linewidth(3)
                        wedge.set_alpha(0.9)
                    
                    # מקרא מקצועי עם סכומים
                    legend_labels = [f"₪ {s:,.0f}\n{l}" for l, s in zip(labels, sizes)]
                    legend = ax.legend(
                        wedges,
                        legend_labels,
                        loc="center left",
                        bbox_to_anchor=(1.15, 0.5),
                        fontsize=12,
                        frameon=True,
                        framealpha=0.95,
                        edgecolor='#E2E8F0',
                        facecolor='white',
                        shadow=True
                    )
                    legend.get_frame().set_linewidth(1.5)
                    
                    if single_value:
                        ax.text(0, 0, f"100%", ha='center', va='center', fontsize=24, 
                               color='white', fontweight='bold')
                    
                    ax.set_title(fix_hebrew('הכנסות מול הוצאות'), 
                                fontsize=18, fontweight='bold', color=COLOR_TEXT, pad=20)
                    chart_created = True

            # --- תרשים עוגה: הוצאות לפי קטגוריה ---
            elif chart_type == "expense_by_category_pie" and transactions:
                category_expenses = {}
                for tx in transactions:
                    if tx.get('type') == 'Expense':
                        cat = tx.get('category') or 'כללי'
                        category_expenses[cat] = category_expenses.get(cat, 0) + float(tx.get('amount', 0) or 0)

                category_expenses = {k: v for k, v in category_expenses.items() if v > 0}

                if not category_expenses:
                    plt.close(fig)
                    print("INFO: No expense data for expense_by_category_pie chart")
                    return None
                
                sorted_pairs = sorted(category_expenses.items(), key=lambda x: x[1], reverse=True)
                # Fix Hebrew for category names
                labels = [fix_hebrew(p[0]) for p in sorted_pairs]
                sizes = [p[1] for p in sorted_pairs]

                # צבעים מקצועיים - פלטת צבעים מגוונת
                color_palette = ['#E11D48', '#F97316', '#EAB308', '#22C55E', '#06B6D4', 
                                '#3B82F6', '#8B5CF6', '#EC4899', '#F43F5E', '#84CC16',
                                '#14B8A6', '#A855F7', '#F59E0B', '#EF4444', '#10B981']
                pie_colors = [color_palette[i % len(color_palette)] for i in range(len(labels))]
                
                ax.axis('equal')
                # הפרדה עדינה יותר לחלקים הקטנים
                explode_values = [0.05 if s < max(sizes) * 0.1 else 0.02 for s in sizes]
                wedges, texts, autotexts = ax.pie(
                    sizes, 
                    colors=pie_colors, 
                    autopct=lambda p: f'{p:.1f}%' if p > 3 else '',  # הצג אחוזים רק אם גדול מ-3%
                    startangle=90, 
                    textprops={'fontsize': 11, 'weight': 'bold', 'color': 'white'}, 
                    shadow=True,
                    explode=explode_values
                )
                for wedge in wedges:
                    wedge.set_edgecolor('white')
                    wedge.set_linewidth(2.5)
                    wedge.set_alpha(0.9)

                # מקרא מקצועי עם סכומים - format: amount first, then name
                legend_labels = [f"₪ {s:,.0f}\n{l}" for l, s in zip(labels, sizes)]
                legend = ax.legend(
                    wedges,
                    legend_labels,
                    loc="center left",
                    bbox_to_anchor=(1.2, 0.5),
                    fontsize=10,
                    frameon=True,
                    framealpha=0.95,
                    edgecolor='#E2E8F0',
                    facecolor='white',
                    shadow=True
                )
                legend.get_frame().set_linewidth(1.5)
                
                ax.set_title(fix_hebrew('הוצאות לפי קטגוריה'), 
                            fontsize=18, fontweight='bold', color=COLOR_TEXT, pad=20)
                chart_created = True
            # --- תרשים עמודות: הוצאות לפי קטגוריה ---
            elif chart_type == "expense_by_category_bar" and transactions:
                category_expenses = {}
                for tx in transactions:
                    if tx.get('type') == 'Expense':
                        cat = tx.get('category') or 'כללי'
                        category_expenses[cat] = category_expenses.get(cat, 0) + float(tx.get('amount', 0) or 0)

                category_expenses = {k: v for k, v in category_expenses.items() if v > 0}
                
                if not category_expenses:
                    print("INFO: No expense data for expense_by_category_bar chart")
                    return None

                sorted_cats = sorted(category_expenses.items(), key=lambda x: x[1], reverse=True)
                # Fix Hebrew for category names
                categories = [fix_hebrew(x[0]) for x in sorted_cats]
                amounts = [x[1] for x in sorted_cats]

                print(f"INFO: Creating bar chart with {len(categories)} categories")
                
                # צבע מקצועי - גרדיאנט לפי גובה
                max_amount = max(amounts)
                bar_colors = []
                base_color = COLOR_EXPENSE
                for amount in amounts:
                    # ככל שהסכום גדול יותר, הצבע כהה יותר
                    intensity = 0.6 + (amount / max_amount) * 0.4
                    bar_colors.append(base_color)
                
                bars = ax.bar(categories, amounts, color=bar_colors, edgecolor='white', 
                             linewidth=2, alpha=0.85)
                
                # רשת רקע עדינה
                ax.grid(axis='y', alpha=0.2, linestyle='--', linewidth=0.8)
                ax.set_ylabel(fix_hebrew('סכום (₪)'), fontsize=12, fontweight='bold', color=COLOR_TEXT)
                ax.tick_params(axis='x', rotation=45, labelsize=10)
                ax.tick_params(axis='y', labelsize=10)

                # תוויות על העמודות
                for bar in bars:
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width() / 2., height + (max(amounts) * 0.02), 
                           f'₪ {height:,.0f}', ha='center', va='bottom', 
                           fontsize=9, fontweight='bold', color=COLOR_TEXT)
                
                ax.set_title(fix_hebrew('הוצאות לפי קטגוריה'), 
                            fontsize=18, fontweight='bold', color=COLOR_TEXT, pad=20)
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
                ax.spines['left'].set_color('#E2E8F0')
                ax.spines['bottom'].set_color('#E2E8F0')
                chart_created = True

            # --- תרשים קו: מגמות לאורך זמן ---
            elif chart_type == "trends_line" and transactions:
                from collections import defaultdict
                daily_data = defaultdict(lambda: {'income': 0, 'expense': 0})

                for tx in transactions:
                    tx_date = tx.get('tx_date')
                    if tx_date:
                        if isinstance(tx_date, str):
                            tx_date = date.fromisoformat(tx_date.split('T')[0])
                        date_str = tx_date.strftime('%Y-%m-%d')

                        if tx.get('type') == 'Income':
                            daily_data[date_str]['income'] += float(tx.get('amount', 0) or 0)
                        else:
                            daily_data[date_str]['expense'] += float(tx.get('amount', 0) or 0)

                if not daily_data:
                    print("INFO: No transaction data for trends_line chart")
                    return None

                sorted_dates = sorted(daily_data.keys())
                incomes = [daily_data[d]['income'] for d in sorted_dates]
                expenses = [daily_data[d]['expense'] for d in sorted_dates]
                
                # Check if there's actual data (not all zeros)
                if sum(incomes) == 0 and sum(expenses) == 0:
                    print("INFO: All zeros in trends_line chart data")
                    return None

                print(f"INFO: Creating trends chart with {len(sorted_dates)} dates")
                
                # קו הכנסות מקצועי
                ax.plot(sorted_dates, incomes, marker='o', label=fix_hebrew('הכנסות'), 
                       color=COLOR_INCOME, linewidth=3, markersize=8, markerfacecolor=COLOR_INCOME,
                       markeredgecolor='white', markeredgewidth=2)
                # קו הוצאות מקצועי
                ax.plot(sorted_dates, expenses, marker='s', label=fix_hebrew('הוצאות'), 
                       color=COLOR_EXPENSE, linewidth=3, markersize=8, markerfacecolor=COLOR_EXPENSE,
                       markeredgecolor='white', markeredgewidth=2)
                # מילוי תחת הקווים - אפקט מקצועי
                ax.fill_between(sorted_dates, incomes, alpha=0.2, color=COLOR_INCOME)
                ax.fill_between(sorted_dates, expenses, alpha=0.2, color=COLOR_EXPENSE)
                
                ax.tick_params(axis='x', rotation=45, labelsize=10)
                ax.tick_params(axis='y', labelsize=10)
                ax.legend(loc='upper right', fontsize=12, framealpha=0.95, 
                         edgecolor='#E2E8F0', facecolor='white', shadow=True)
                ax.set_title(fix_hebrew('מגמות הכנסות והוצאות לאורך זמן'), 
                            fontsize=18, fontweight='bold', color=COLOR_TEXT, pad=20)
                ax.set_ylabel(fix_hebrew('סכום (₪)'), fontsize=12, fontweight='bold', color=COLOR_TEXT)
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
                ax.spines['left'].set_color('#E2E8F0')
                ax.spines['bottom'].set_color('#E2E8F0')
                ax.grid(axis='y', alpha=0.2, linestyle='--', linewidth=0.8)
                chart_created = True

            # --- תרשים עמודות: הוצאות לפי ספק ---
            elif chart_type == "expense_by_supplier_bar" and transactions:
                supplier_expenses = {}
                for tx in transactions:
                    if tx.get('type') == 'Expense':
                        supplier = tx.get('supplier') or 'ללא ספק'
                        supplier_expenses[supplier] = supplier_expenses.get(supplier, 0) + float(tx.get('amount', 0) or 0)

                supplier_expenses = {k: v for k, v in supplier_expenses.items() if v > 0}
                
                if not supplier_expenses:
                    print("INFO: No expense data for expense_by_supplier_bar chart")
                    return None

                sorted_suppliers = sorted(supplier_expenses.items(), key=lambda x: x[1], reverse=True)[:10]  # Top 10
                suppliers = [fix_hebrew(x[0]) for x in sorted_suppliers]
                amounts = [x[1] for x in sorted_suppliers]

                print(f"INFO: Creating supplier bar chart with {len(suppliers)} suppliers")
                
                bars = ax.barh(suppliers, amounts, color=COLOR_EXPENSE, edgecolor='white', 
                             linewidth=2, alpha=0.85)
                
                ax.grid(axis='x', alpha=0.2, linestyle='--', linewidth=0.8)
                ax.set_xlabel(fix_hebrew('סכום (₪)'), fontsize=12, fontweight='bold', color=COLOR_TEXT)
                ax.tick_params(axis='x', labelsize=10)
                ax.tick_params(axis='y', labelsize=9)

                for i, bar in enumerate(bars):
                    width = bar.get_width()
                    ax.text(width + (max(amounts) * 0.02), bar.get_y() + bar.get_height() / 2., 
                           f'₪ {width:,.0f}', ha='left', va='center', 
                           fontsize=9, fontweight='bold', color=COLOR_TEXT)
                
                ax.set_title(fix_hebrew('הוצאות לפי ספק (10 הגדולים)'), 
                            fontsize=18, fontweight='bold', color=COLOR_TEXT, pad=20)
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
                ax.spines['left'].set_color('#E2E8F0')
                ax.spines['bottom'].set_color('#E2E8F0')
                chart_created = True

            # --- תרשים קו: מגמות חודשיות ---
            elif chart_type == "monthly_trends_line" and transactions:
                from collections import defaultdict
                monthly_data = defaultdict(lambda: {'income': 0, 'expense': 0})

                for tx in transactions:
                    tx_date = tx.get('tx_date')
                    if tx_date:
                        if isinstance(tx_date, str):
                            tx_date = date.fromisoformat(tx_date.split('T')[0])
                        month_key = tx_date.strftime('%Y-%m')

                        if tx.get('type') == 'Income':
                            monthly_data[month_key]['income'] += float(tx.get('amount', 0) or 0)
                        else:
                            monthly_data[month_key]['expense'] += float(tx.get('amount', 0) or 0)

                if not monthly_data:
                    print("INFO: No transaction data for monthly_trends_line chart")
                    return None

                sorted_months = sorted(monthly_data.keys())
                incomes = [monthly_data[m]['income'] for m in sorted_months]
                expenses = [monthly_data[m]['expense'] for m in sorted_months]
                
                if sum(incomes) == 0 and sum(expenses) == 0:
                    print("INFO: All zeros in monthly_trends_line chart data")
                    return None

                print(f"INFO: Creating monthly trends chart with {len(sorted_months)} months")
                
                ax.plot(sorted_months, incomes, marker='o', label=fix_hebrew('הכנסות'), 
                       color=COLOR_INCOME, linewidth=3, markersize=8, markerfacecolor=COLOR_INCOME,
                       markeredgecolor='white', markeredgewidth=2)
                ax.plot(sorted_months, expenses, marker='s', label=fix_hebrew('הוצאות'), 
                       color=COLOR_EXPENSE, linewidth=3, markersize=8, markerfacecolor=COLOR_EXPENSE,
                       markeredgecolor='white', markeredgewidth=2)
                ax.fill_between(sorted_months, incomes, alpha=0.2, color=COLOR_INCOME)
                ax.fill_between(sorted_months, expenses, alpha=0.2, color=COLOR_EXPENSE)
                
                ax.tick_params(axis='x', rotation=45, labelsize=10)
                ax.tick_params(axis='y', labelsize=10)
                ax.legend(loc='upper right', fontsize=12, framealpha=0.95, 
                         edgecolor='#E2E8F0', facecolor='white', shadow=True)
                ax.set_title(fix_hebrew('מגמות חודשיות - הכנסות והוצאות'), 
                            fontsize=18, fontweight='bold', color=COLOR_TEXT, pad=20)
                ax.set_ylabel(fix_hebrew('סכום (₪)'), fontsize=12, fontweight='bold', color=COLOR_TEXT)
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
                ax.spines['left'].set_color('#E2E8F0')
                ax.spines['bottom'].set_color('#E2E8F0')
                ax.grid(axis='y', alpha=0.2, linestyle='--', linewidth=0.8)
                chart_created = True

            # --- תרשים עוגה: תקציב מול ביצוע ---
            elif chart_type == "budget_vs_actual" and budgets:
                budget_data = []
                actual_data = []
                labels = []
                
                for b in budgets:
                    cat_name = b.get('category') or 'כללי'
                    budget_amount = float(b.get('amount', 0) or 0)
                    spent_amount = float(b.get('spent_amount', 0) or 0)
                    
                    if budget_amount > 0:
                        labels.append(fix_hebrew(cat_name))
                        budget_data.append(budget_amount)
                        actual_data.append(spent_amount)

                if not budget_data:
                    print("INFO: No budget data for budget_vs_actual chart")
                    return None

                x = range(len(labels))
                width = 0.35

                bars1 = ax.bar([i - width/2 for i in x], budget_data, width, 
                              label=fix_hebrew('תקציב'), color=COLOR_INCOME, alpha=0.8, edgecolor='white', linewidth=2)
                bars2 = ax.bar([i + width/2 for i in x], actual_data, width, 
                              label=fix_hebrew('ביצוע'), color=COLOR_EXPENSE, alpha=0.8, edgecolor='white', linewidth=2)

                ax.set_xlabel(fix_hebrew('קטגוריה'), fontsize=12, fontweight='bold', color=COLOR_TEXT)
                ax.set_ylabel(fix_hebrew('סכום (₪)'), fontsize=12, fontweight='bold', color=COLOR_TEXT)
                ax.set_title(fix_hebrew('תקציב מול ביצוע לפי קטגוריה'), 
                            fontsize=18, fontweight='bold', color=COLOR_TEXT, pad=20)
                ax.set_xticks(x)
                ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=9)
                ax.legend(fontsize=12, framealpha=0.95, edgecolor='#E2E8F0', facecolor='white', shadow=True)
                ax.grid(axis='y', alpha=0.2, linestyle='--', linewidth=0.8)
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
                ax.spines['left'].set_color('#E2E8F0')
                ax.spines['bottom'].set_color('#E2E8F0')
                chart_created = True

            # --- תרשים קו: הוצאות מצטברות ---
            elif chart_type == "cumulative_expenses" and transactions:
                from collections import defaultdict
                daily_expenses = defaultdict(float)

                for tx in transactions:
                    if tx.get('type') == 'Expense':
                        tx_date = tx.get('tx_date')
                        if tx_date:
                            if isinstance(tx_date, str):
                                tx_date = date.fromisoformat(tx_date.split('T')[0])
                            date_str = tx_date.strftime('%Y-%m-%d')
                            daily_expenses[date_str] += float(tx.get('amount', 0) or 0)

                if not daily_expenses:
                    print("INFO: No expense data for cumulative_expenses chart")
                    return None

                sorted_dates = sorted(daily_expenses.keys())
                cumulative = []
                running_total = 0
                for d in sorted_dates:
                    running_total += daily_expenses[d]
                    cumulative.append(running_total)

                if max(cumulative) == 0:
                    print("INFO: All zeros in cumulative_expenses chart data")
                    return None

                print(f"INFO: Creating cumulative expenses chart with {len(sorted_dates)} dates")
                
                ax.plot(sorted_dates, cumulative, marker='o', color=COLOR_EXPENSE, 
                       linewidth=3, markersize=6, markerfacecolor=COLOR_EXPENSE,
                       markeredgecolor='white', markeredgewidth=2)
                ax.fill_between(sorted_dates, cumulative, alpha=0.3, color=COLOR_EXPENSE)
                
                ax.tick_params(axis='x', rotation=45, labelsize=10)
                ax.tick_params(axis='y', labelsize=10)
                ax.set_title(fix_hebrew('הוצאות מצטברות לאורך זמן'), 
                            fontsize=18, fontweight='bold', color=COLOR_TEXT, pad=20)
                ax.set_ylabel(fix_hebrew('סכום מצטבר (₪)'), fontsize=12, fontweight='bold', color=COLOR_TEXT)
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
                ax.spines['left'].set_color('#E2E8F0')
                ax.spines['bottom'].set_color('#E2E8F0')
                ax.grid(axis='y', alpha=0.2, linestyle='--', linewidth=0.8)
                chart_created = True

            else:
                # סוג גרף לא מוכר או חסרים נתונים
                print(f"INFO: Unknown chart type or missing data: {chart_type}")
                return None

            # בדיקה שגרף נוצר
            if not chart_created:
                print(f"WARNING: No chart was created for type: {chart_type}")
                return None

            # שמירת התוצאה עם רקע לבן מפורש
            try:
                fig.tight_layout()
            except Exception as layout_err:
                print(f"WARNING: tight_layout failed: {layout_err}")
            
            # שמירת התמונה באמצעות canvas
            img_buffer = io.BytesIO()
            canvas.print_png(img_buffer)
            
            img_buffer.seek(0)
            img_data = img_buffer.read()
            
            # בדיקה שהתמונה נוצרה כראוי
            if len(img_data) < 1000:
                print(f"WARNING: Chart image seems too small ({len(img_data)} bytes)")
                return None
            
            # בדיקת חתימת PNG
            if img_data[:8] != b'\x89PNG\r\n\x1a\n':
                print(f"WARNING: Generated image is not valid PNG")
                return None
            
            print(f"INFO: Chart '{chart_type}' created successfully, size: {len(img_data)} bytes")
            img_buffer.seek(0)
            return img_buffer

        except ImportError as ie:
            print(f"Matplotlib לא מותקן: {ie}")
            return None
        except Exception as e:
            print(f"שגיאה ביצירת גרף '{chart_type}': {e}")
            import traceback
            traceback.print_exc()
            return None

    def _add_native_excel_charts(
            self, 
            wb, 
            ws, 
            start_row: int, 
            summary: Dict[str, Any], 
            transactions: List[Dict],
            chart_types: List[str]
    ) -> int:
        """
        Add native Excel charts (PieChart, BarChart, LineChart) to the workbook.
        Returns the number of charts added.
        """
        if not CHARTS_AVAILABLE:
            return 0

        charts_added = 0
        current_row = start_row
        
        # Create a separate data sheet for chart data
        data_sheet_name = "_ChartData"
        if data_sheet_name in wb.sheetnames:
            ws_data = wb[data_sheet_name]
        else:
            ws_data = wb.create_sheet(data_sheet_name)
        
        data_row = 1

        try:
            # Income vs Expense Pie Chart
            if "income_expense_pie" in chart_types and summary:
                income = float(summary.get('income', 0))
                expenses = float(summary.get('expenses', 0))
                
                if income > 0 or expenses > 0:
                    # Write data to hidden sheet
                    ws_data[f'A{data_row}'] = REPORT_LABELS['income']
                    ws_data[f'B{data_row}'] = income
                    ws_data[f'A{data_row + 1}'] = REPORT_LABELS['expenses']
                    ws_data[f'B{data_row + 1}'] = expenses
                    
                    # Create pie chart
                    chart = PieChart()
                    chart.title = f"{REPORT_LABELS['income']} מול {REPORT_LABELS['expenses']}"
                    
                    labels = Reference(ws_data, min_col=1, min_row=data_row, max_row=data_row + 1)
                    data = Reference(ws_data, min_col=2, min_row=data_row, max_row=data_row + 1)
                    chart.add_data(data)
                    chart.set_categories(labels)
                    
                    # Style the chart
                    chart.width = 12
                    chart.height = 8
                    
                    # Add chart to main sheet
                    ws.add_chart(chart, f'A{current_row}')
                    current_row += 18
                    data_row += 3
                    charts_added += 1

            # Expense by Category Pie Chart
            if "expense_by_category_pie" in chart_types and transactions:
                category_expenses = {}
                for tx in transactions:
                    if tx.get('type') == 'Expense':
                        cat = tx.get('category') or REPORT_LABELS['general']
                        category_expenses[cat] = category_expenses.get(cat, 0) + float(tx.get('amount', 0))
                
                if category_expenses:
                    sorted_cats = sorted(category_expenses.items(), key=lambda x: x[1], reverse=True)
                    
                    # Write data
                    cat_start_row = data_row
                    for idx, (cat_name, amount) in enumerate(sorted_cats):
                        ws_data[f'A{data_row + idx}'] = cat_name
                        ws_data[f'B{data_row + idx}'] = amount
                    
                    # Create pie chart
                    chart = PieChart()
                    chart.title = f"{REPORT_LABELS['expenses']} לפי {REPORT_LABELS['category']}"
                    
                    labels = Reference(ws_data, min_col=1, min_row=cat_start_row, max_row=cat_start_row + len(sorted_cats) - 1)
                    data = Reference(ws_data, min_col=2, min_row=cat_start_row, max_row=cat_start_row + len(sorted_cats) - 1)
                    chart.add_data(data)
                    chart.set_categories(labels)
                    
                    chart.width = 12
                    chart.height = 8
                    
                    ws.add_chart(chart, f'A{current_row}')
                    current_row += 18
                    data_row += len(sorted_cats) + 2
                    charts_added += 1

            # Expense by Category Bar Chart
            if "expense_by_category_bar" in chart_types and transactions:
                category_expenses = {}
                for tx in transactions:
                    if tx.get('type') == 'Expense':
                        cat = tx.get('category') or REPORT_LABELS['general']
                        category_expenses[cat] = category_expenses.get(cat, 0) + float(tx.get('amount', 0))
                
                if category_expenses:
                    sorted_cats = sorted(category_expenses.items(), key=lambda x: x[1], reverse=True)
                    
                    # Write data with header
                    ws_data[f'A{data_row}'] = REPORT_LABELS['category']
                    ws_data[f'B{data_row}'] = REPORT_LABELS['amount']
                    
                    for idx, (cat_name, amount) in enumerate(sorted_cats):
                        ws_data[f'A{data_row + idx + 1}'] = cat_name
                        ws_data[f'B{data_row + idx + 1}'] = amount
                    
                    # Create bar chart
                    chart = BarChart()
                    chart.title = f"{REPORT_LABELS['expenses']} לפי {REPORT_LABELS['category']}"
                    chart.type = "col"
                    chart.style = 10
                    
                    data = Reference(ws_data, min_col=2, min_row=data_row, max_row=data_row + len(sorted_cats))
                    categories = Reference(ws_data, min_col=1, min_row=data_row + 1, max_row=data_row + len(sorted_cats))
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(categories)
                    
                    chart.width = 14
                    chart.height = 8
                    
                    ws.add_chart(chart, f'A{current_row}')
                    current_row += 18
                    data_row += len(sorted_cats) + 3
                    charts_added += 1

            # Trends Line Chart
            if "trends_line" in chart_types and transactions:
                from collections import defaultdict
                daily_data = defaultdict(lambda: {'income': 0, 'expense': 0})
                
                for tx in transactions:
                    tx_date = tx.get('tx_date')
                    if tx_date:
                        if isinstance(tx_date, str):
                            date_str = tx_date[:10]  # YYYY-MM-DD
                        else:
                            date_str = tx_date.strftime('%Y-%m-%d')
                        
                        if tx.get('type') == 'Income':
                            daily_data[date_str]['income'] += float(tx.get('amount', 0))
                        else:
                            daily_data[date_str]['expense'] += float(tx.get('amount', 0))
                
                if daily_data:
                    sorted_dates = sorted(daily_data.keys())
                    
                    # Write headers
                    ws_data[f'A{data_row}'] = REPORT_LABELS['date']
                    ws_data[f'B{data_row}'] = REPORT_LABELS['income']
                    ws_data[f'C{data_row}'] = REPORT_LABELS['expenses']
                    
                    # Write data
                    for idx, date_str in enumerate(sorted_dates):
                        ws_data[f'A{data_row + idx + 1}'] = date_str
                        ws_data[f'B{data_row + idx + 1}'] = daily_data[date_str]['income']
                        ws_data[f'C{data_row + idx + 1}'] = daily_data[date_str]['expense']
                    
                    # Create line chart
                    chart = LineChart()
                    chart.title = "מגמות לאורך זמן"
                    chart.style = 13
                    
                    data = Reference(ws_data, min_col=2, min_row=data_row, max_col=3, max_row=data_row + len(sorted_dates))
                    categories = Reference(ws_data, min_col=1, min_row=data_row + 1, max_row=data_row + len(sorted_dates))
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(categories)
                    
                    chart.width = 16
                    chart.height = 8
                    
                    ws.add_chart(chart, f'A{current_row}')
                    current_row += 18
                    charts_added += 1

            # Hide the data sheet
            ws_data.sheet_state = 'hidden'
            
        except Exception as e:
            import traceback
            print(f"אזהרה: שגיאה ביצירת גרפים נייטיביים ב-Excel: {e}")
            traceback.print_exc()
        
        return charts_added

    async def _generate_pdf(self, project, options, transactions, budgets, fund, summary, chart_images=None, monthly_breakdown=None) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)

        # Register Hebrew Font
        font_name = 'Helvetica'  # Default fallback
        try:
            # Get the base directory of the services folder
            services_dir = os.path.dirname(__file__)
            backend_dir = os.path.dirname(services_dir)
            project_root = os.path.dirname(backend_dir)

            # Paths to check for the font (in order of preference)
            possible_paths = [
                os.path.join(backend_dir, 'static', 'fonts', 'Heebo-Regular.ttf'),  # Relative from services/
                os.path.join(project_root, 'backend', 'static', 'fonts', 'Heebo-Regular.ttf'),  # From project root
                '/app/backend/static/fonts/Heebo-Regular.ttf',  # Docker absolute path
                'backend/static/fonts/Heebo-Regular.ttf',  # Run from root (string path)
            ]

            font_path = None
            print(f"🔍 Looking for Hebrew font in {len(possible_paths)} possible locations...")
            print(f"   Services dir: {services_dir}")
            print(f"   Backend dir: {backend_dir}")
            print(f"   Project root: {project_root}")

            for path in possible_paths:
                if os.path.exists(path):
                    font_path = os.path.abspath(path)  # Use absolute path
                    print(f"✓ Found font at: {font_path}")
                    break
                else:
                    print(f"✗ לא נמצא: {path}")

            # If font not found or corrupted, try to download it (Self-healing)
            if not font_path or (font_path and os.path.exists(font_path)):
                # Check if existing font is valid by trying to read it
                if font_path and os.path.exists(font_path):
                    try:
                        # Quick validation - try to open as TTFont
                        test_font = TTFont(font_path)
                        test_font.close()
                        print(f"✓ קובץ גופן קיים תקין")
                    except Exception:
                        print(f"אזהרה: קובץ גופן קיים נראה פגום, מנסה להוריד מחדש")
                        font_path = None  # Mark as not found so we try to download

                if not font_path:
                    try:
                        import urllib.request
                        print("גופן לא נמצא או פגום. מנסה להוריד Heebo-Regular.ttf...")

                        # Determine where to save
                        if os.path.exists('/app/backend/static'):
                            target_dir = '/app/backend/static/fonts'
                        else:
                            # Dev environment or fallback
                            target_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'fonts')

                        os.makedirs(target_dir, exist_ok=True)
                        target_path = os.path.join(target_dir, 'Heebo-Regular.ttf')

                        # Try multiple URLs
                        urls = [
                            "https://github.com/google/fonts/raw/main/ofl/heebo/static/Heebo-Regular.ttf",
                            "https://raw.githubusercontent.com/google/fonts/main/ofl/heebo/static/Heebo-Regular.ttf",
                        ]

                        downloaded = False
                        for url in urls:
                            try:
                                print(f"מנסה להוריד מ-: {url}")
                                urllib.request.urlretrieve(url, target_path)
                                # Validate the downloaded file
                                test_font = TTFont(target_path)
                                test_font.close()
                                font_path = target_path
                                downloaded = True
                                print(f"✓ גופן הורד ואומת בהצלחה מ-{url}")
                                break
                            except Exception as e:
                                print(f"✗ הורדה/אימות מ-{url} נכשל: {e}")
                                if os.path.exists(target_path):
                                    os.remove(target_path)
                                continue

                        if not downloaded:
                            print("אזהרה: לא ניתן להוריד קובץ גופן תקין")
                    except Exception as e:
                        print(f"הורדת גופן נכשלה: {e}")

            font_loaded = False
            if font_path and os.path.exists(font_path):
                try:
                    # Test if font can be loaded and supports Hebrew
                    test_font_obj = TTFont('TestHebrew', font_path, subfontIndex=0)
                    # Check if font has Hebrew character support by checking Unicode ranges
                    # This is a basic check - if font loads, it should work
                    test_font_obj.close()
                    
                    # Register font with proper embedding - subfontIndex=0 ensures subset embedding
                    # This embeds only the characters used, making PDFs smaller
                    hebrew_font = TTFont('Hebrew', font_path, subfontIndex=0)
                    pdfmetrics.registerFont(hebrew_font)
                    font_name = 'Hebrew'
                    font_loaded = True
                    print(f"✓ גופן עברי נרשם בהצלחה מ-{font_path}")
                except Exception as e:
                    print(f"✗ רישום גופן מ-{font_path} נכשל: {e}")
                    import traceback
                    print(traceback.format_exc())
                    font_path = None  # Mark as failed so we try system fonts

            # Try Windows system fonts with Hebrew support (if Heebo not found or failed)
            if not font_loaded and os.name == 'nt':  # Windows
                windows_fonts_dir = r'C:\Windows\Fonts'
                # Try multiple possible font file names for each font family
                windows_fonts = [
                    ['arial.ttf', 'arial.ttc', 'Arial.ttf', 'Arial Regular.ttf'],
                    ['tahoma.ttf', 'Tahoma.ttf', 'tahoma.ttc'],
                    ['arialuni.ttf', 'Arial Unicode MS.ttf', 'ARIALUNI.TTF'],
                    ['gadugi.ttf', 'Gadugi.ttf', 'GADUGI.TTF'],
                    ['calibri.ttf', 'Calibri.ttf', 'calibri.ttc', 'CALIBRI.TTF'],
                    ['segoeui.ttf', 'Segoe UI.ttf', 'segoeui.ttc'],
                ]
                print("🔍 מנסה גופני מערכת של Windows עם תמיכה בעברית...")
                for font_variants in windows_fonts:
                    for font_file in font_variants:
                        win_font_path = os.path.join(windows_fonts_dir, font_file)
                        if os.path.exists(win_font_path):
                            try:
                                # Register with proper embedding - use subfontIndex=0 for subset embedding
                                # This ensures the font is properly embedded in the PDF
                                hebrew_font = TTFont('Hebrew', win_font_path, subfontIndex=0)
                                pdfmetrics.registerFont(hebrew_font)
                                font_name = 'Hebrew'
                                font_loaded = True
                                print(f"✓ משתמש בהצלחה בגופן מערכת Windows: {font_file}")
                                break
                            except Exception as e3:
                                print(f"✗ טעינת {win_font_path} נכשלה: {e3}")
                                continue
                    if font_loaded:
                        break

            # Try Linux system font as last resort (only if not Windows)
            if not font_loaded and os.name != 'nt':
                linux_fonts = [
                    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                    '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                    '/usr/share/fonts/truetype/noto/NotoSansHebrew-Regular.ttf',
                ]
                for linux_font_path in linux_fonts:
                    if os.path.exists(linux_font_path):
                        try:
                            hebrew_font = TTFont('Hebrew', linux_font_path, subfontIndex=0)
                            pdfmetrics.registerFont(hebrew_font)
                            font_name = 'Hebrew'
                            font_loaded = True
                            print(f"✓ משתמש בגופן מערכת Linux: {linux_font_path}")
                            break
                        except Exception as e2:
                            print(f"✗ Failed to load {linux_font_path}: {e2}")
                            continue

        except Exception as e:
            print(f"✗ אזהרה: גופן עברי לא נמצא ({e}), משתמש ב-Helvetica ברירת מחדל")
            font_loaded = False

        if not font_loaded:
            print("❌ שגיאה קריטית: גופן עברי לא נטען! טקסט עברי לא יוצג כראוי ויופיע כקוביות שחורות.")
            print("   אנא ודא שאחד מהגופנים הבאים מותקן במערכת:")
            print("   - Heebo-Regular.ttf (מומלץ)")
            print("   - Arial, Tahoma, Arial Unicode MS, או Calibri (Windows)")
            # Don't fail completely, but warn that output will be broken
            # The font_name is still 'Helvetica' which won't work for Hebrew

        styles = getSampleStyleSheet()
        
        # ===== PROFESSIONAL COLOR PALETTE (matching Excel) =====
        COLOR_PRIMARY_DARK = '#0F172A'    # Slate-900
        COLOR_PRIMARY_MID = '#1E293B'     # Slate-800
        COLOR_ACCENT_TEAL = '#0D9488'     # Teal-600
        COLOR_ACCENT_EMERALD = '#059669'  # Emerald-600
        COLOR_ACCENT_ROSE = '#E11D48'     # Rose-600
        COLOR_ACCENT_AMBER = '#D97706'    # Amber-600
        COLOR_BG_LIGHT = '#F8FAFC'        # Slate-50
        COLOR_BG_ALT = '#F1F5F9'          # Slate-100
        COLOR_TEXT_DARK = '#0F172A'       # Dark text
        COLOR_TEXT_MUTED = '#64748B'      # Slate-500
        
        # Main title style - elegant dark with accent
        style_title = ParagraphStyle(
            'HebrewTitle', 
            parent=styles['Heading1'], 
            fontName=font_name, 
            fontSize=24, 
            alignment=1,
            textColor=colors.HexColor(COLOR_PRIMARY_DARK), 
            leading=32, 
            spaceAfter=8, 
            spaceBefore=15,
            borderWidth=0,
            borderPadding=12,
            borderColor=colors.HexColor(COLOR_ACCENT_TEAL),
            borderRadius=4
        )
        
        # Section header style - teal accent
        style_h2 = ParagraphStyle(
            'HebrewHeading2', 
            parent=styles['Heading2'], 
            fontName=font_name, 
            fontSize=14, 
            alignment=1,
            textColor=colors.white,
            backColor=colors.HexColor(COLOR_ACCENT_TEAL),
            leading=20, 
            spaceBefore=18, 
            spaceAfter=10,
            leftIndent=10,
            rightIndent=10,
            borderPadding=8
        )
        
        # Category header style - amber accent
        style_category = ParagraphStyle(
            'HebrewCategory', 
            parent=styles['Heading2'], 
            fontName=font_name, 
            fontSize=12, 
            alignment=1,
            textColor=colors.white,
            backColor=colors.HexColor(COLOR_ACCENT_AMBER),
            leading=18, 
            spaceBefore=12, 
            spaceAfter=6,
            leftIndent=8,
            rightIndent=8,
            borderPadding=6
        )
        
        # Normal text style
        style_normal = ParagraphStyle(
            'HebrewNormal', 
            parent=styles['Normal'], 
            fontName=font_name, 
            fontSize=10, 
            alignment=1,
            leading=14, 
            spaceAfter=8,
            textColor=colors.HexColor(COLOR_TEXT_DARK)
        )
        
        # Subtitle/date style
        style_subtitle = ParagraphStyle(
            'HebrewSubtitle', 
            parent=styles['Normal'], 
            fontName=font_name, 
            fontSize=10, 
            alignment=1,
            leading=14, 
            spaceAfter=15,
            textColor=colors.HexColor(COLOR_TEXT_MUTED),
            fontStyle='italic'
        )
        
        # Table cell style with text wrapping
        style_table_cell = ParagraphStyle(
            'HebrewTableCell', 
            parent=styles['Normal'], 
            fontName=font_name, 
            fontSize=9, 
            alignment=1, 
            leading=12, 
            wordWrap='CJK',
            textColor=colors.HexColor(COLOR_TEXT_DARK)
        )

        elements = []

        # Use arabic-reshaper and python-bidi for proper RTL support
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            bidi_available = True
        except ImportError:
            bidi_available = False
            print("⚠️ אזהרה: arabic-reshaper או python-bidi לא זמינים, משתמש בעיצוב טקסט פשוט")
            print("   התקן עם: pip install arabic-reshaper python-bidi")

        def format_text(text):
            if not text: return ""
            if not isinstance(text, str): text = str(text)

            # If font is loaded and bidi is available, use proper RTL shaping
            if font_loaded and bidi_available:
                try:
                    # Reshape Arabic/Hebrew text for proper display
                    reshaped_text = arabic_reshaper.reshape(text)
                    # Get bidirectional display
                    bidi_text = get_display(reshaped_text)
                    return bidi_text
                except Exception as e:
                    # Only log first error to avoid spam
                    if not hasattr(format_text, '_logged_error'):
                        print(f"⚠️ אזהרה: שגיאה בעיבוד bidi: {e}, משתמש בטקסט כפי שהוא")
                        format_text._logged_error = True
                    return text
            elif font_loaded:
                # Font loaded but bidi not available - still use the font, text might display backwards
                # but at least it won't be black squares
                return text
            else:
                # No font loaded - Hebrew will show as black squares
                if not hasattr(format_text, '_no_font_warned'):
                    print("⚠️ אזהרה: גופן עברי לא נטען - טקסט עברי יופיע כקוביות שחורות!")
                    format_text._no_font_warned = True
                return text

        # Add Logo at the top
        try:
            from reportlab.platypus import Image
            # Try to find logo in multiple locations
            logo_paths = [
                os.path.join(os.path.dirname(os.path.dirname(__file__)), '..', 'frontend', 'public', 'logo.png'),
                os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'frontend', 'public', 'logo.png'),
                'frontend/public/logo.png',
            ]
            logo_path = None
            for path in logo_paths:
                if os.path.exists(path):
                    logo_path = path
                    break

            if logo_path:
                logo = Image(logo_path, width=100, height=100)
                logo.hAlign = 'CENTER'
                elements.append(logo)
                elements.append(Spacer(1, 10))
        except Exception as e:
            print(f"לא ניתן לטעון לוגו: {e}")

        # ===== TITLE SECTION =====
        # Decorative header line
        elements.append(Table([[""]], colWidths=[520], rowHeights=[4], style=[
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(COLOR_ACCENT_TEAL))
        ]))
        elements.append(Spacer(1, 15))
        
        # Main title with project name
        elements.append(Paragraph(format_text(f"📊 {REPORT_LABELS['project_report']}"), style_title))
        elements.append(Paragraph(format_text(project.name), ParagraphStyle(
            'ProjectName', parent=styles['Heading1'], fontName=font_name, fontSize=20, alignment=1,
            textColor=colors.HexColor(COLOR_ACCENT_TEAL), leading=26, spaceAfter=8
        )))
        
        # Date range subtitle
        date_range_text = ""
        if options.start_date and options.end_date:
            date_range_text = f"{options.start_date.strftime('%d/%m/%Y')} - {options.end_date.strftime('%d/%m/%Y')}"
        elif options.start_date:
            date_range_text = f"מ-{options.start_date.strftime('%d/%m/%Y')}"
        elif options.end_date:
            date_range_text = f"עד {options.end_date.strftime('%d/%m/%Y')}"
        else:
            date_range_text = "כל התקופות"
        
        elements.append(Paragraph(
            format_text(f"תקופה: {date_range_text}"),
            style_subtitle
        ))
        
        # Production date
        elements.append(Paragraph(
            format_text(f"{REPORT_LABELS['production_date']}: {date.today().strftime('%d/%m/%Y')}"),
            style_subtitle
        ))
        
        # Decorative divider
        elements.append(Table([[""]], colWidths=[520], rowHeights=[2], style=[
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(COLOR_BG_ALT))
        ]))
        elements.append(Spacer(1, 20))

        # ========== עמוד ראשון: קופה ותקציבים ==========
        
        # Fund - פרטי קופה - Professional Card Style
        if options.include_funds and fund:
            elements.append(Paragraph(format_text(f"🏦 {REPORT_LABELS['fund_status']}"), style_h2))
            elements.append(Spacer(1, 8))
            
            # Reversed for Hebrew RTL: value on left, label on right
            fund_data = [
                [f"{fund.current_balance:,.2f} ₪", format_text(f"💵 {REPORT_LABELS['current_balance']}")],
                [f"{fund.monthly_amount:,.2f} ₪", format_text(f"📅 {REPORT_LABELS['monthly_deposit']}")]
            ]
            fund_table = Table(fund_data, colWidths=[160, 220])
            fund_table.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#CCFBF1')),  # Teal-100 for label
                ('BACKGROUND', (1, 1), (1, 1), colors.HexColor('#CCFBF1')),
                ('BACKGROUND', (0, 0), (0, 1), colors.HexColor(COLOR_BG_LIGHT)),
                ('TEXTCOLOR', (1, 0), (1, 1), colors.HexColor(COLOR_ACCENT_EMERALD)),
                # Align value (column 0) left, label (column 1) right
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 14),
                ('TOPPADDING', (0, 0), (-1, -1), 14),
            ]))
            elements.append(fund_table)
            elements.append(Spacer(1, 20))

        # Budgets - תקציבים - Professional Budget Table
        if options.include_budgets and budgets:
            elements.append(Paragraph(format_text(f"📋 {REPORT_LABELS['budget_vs_actual']}"), style_h2))
            elements.append(Spacer(1, 8))
            
            # Headers with usage percent column - reversed for Hebrew RTL: category on right, calculations on left
            budget_table_data = [[
                format_text("ניצול %"),
                format_text(REPORT_LABELS['remaining']),
                format_text(REPORT_LABELS['used']), 
                format_text(REPORT_LABELS['budget']),
                format_text(REPORT_LABELS['category'])
            ]]
            
            for b in budgets:
                cat_name = b['category'] if b['category'] else REPORT_LABELS['general']
                usage_percent = (b['spent_amount'] / b['amount'] * 100) if b['amount'] > 0 else 0
                budget_table_data.append([
                    f"{usage_percent:.1f}%",
                    f"{b['remaining_amount']:,.2f} ₪",
                    f"{b['spent_amount']:,.2f} ₪",
                    f"{b['amount']:,.2f} ₪",
                    format_text(cat_name)
                ])

            budget_table = Table(budget_table_data, colWidths=[70, 95, 95, 95, 110])
            
            # Build dynamic style for conditional coloring
            budget_style = [
                ('FONT', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLOR_PRIMARY_MID)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                # Align calculation columns (0-3) center, category column (4) right
                ('ALIGN', (0, 0), (3, -1), 'CENTER'),
                ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ]
            
            # Add alternating row colors and conditional formatting
            for row_idx, b in enumerate(budgets):
                row = row_idx + 1  # Account for header row
                usage_percent = (b['spent_amount'] / b['amount'] * 100) if b['amount'] > 0 else 0
                remaining = b['remaining_amount']
                
                # Alternating background
                if row_idx % 2 == 0:
                    budget_style.append(('BACKGROUND', (0, row), (-1, row), colors.HexColor(COLOR_BG_LIGHT)))
                else:
                    budget_style.append(('BACKGROUND', (0, row), (-1, row), colors.HexColor(COLOR_BG_ALT)))
                
                # Color the "used" column (now column 2) based on usage
                if usage_percent > 100:
                    budget_style.append(('TEXTCOLOR', (2, row), (2, row), colors.HexColor(COLOR_ACCENT_ROSE)))
                    budget_style.append(('BACKGROUND', (0, row), (-1, row), colors.HexColor('#FFE4E6')))  # Rose-100
                elif usage_percent > 80:
                    budget_style.append(('TEXTCOLOR', (2, row), (2, row), colors.HexColor(COLOR_ACCENT_AMBER)))
                
                # Color the "remaining" column (now column 1)
                if remaining < 0:
                    budget_style.append(('TEXTCOLOR', (1, row), (1, row), colors.HexColor(COLOR_ACCENT_ROSE)))
                else:
                    budget_style.append(('TEXTCOLOR', (1, row), (1, row), colors.HexColor(COLOR_ACCENT_EMERALD)))
            
            budget_table.setStyle(TableStyle(budget_style))
            elements.append(budget_table)
            elements.append(Spacer(1, 25))

        # ========== סיכום תקציבים ==========
        if options.include_budgets and budgets:
            elements.append(Paragraph(format_text("📊 סיכום תקציבים"), style_h2))
            elements.append(Spacer(1, 15))
            
            # טבלת סיכום תקציבים - reversed for Hebrew RTL: category on right, calculations on left
            budget_summary_data = [[
                format_text("אחוז ניצול"),
                format_text(REPORT_LABELS['remaining']),
                format_text(REPORT_LABELS['used']), 
                format_text(REPORT_LABELS['budget']),
                format_text(REPORT_LABELS['category'])
            ]]
            
            total_budget = 0
            total_spent = 0
            total_remaining = 0
            
            for b in budgets:
                cat_name = b['category'] if b['category'] else REPORT_LABELS['general']
                budget_amount = b['amount']
                spent_amount = b['spent_amount']
                remaining_amount = b['remaining_amount']
                usage_percent = (spent_amount / budget_amount * 100) if budget_amount > 0 else 0
                
                total_budget += budget_amount
                total_spent += spent_amount
                total_remaining += remaining_amount
                
                budget_summary_data.append([
                    f"{usage_percent:.1f}%",
                    f"{remaining_amount:,.2f} ₪",
                    f"{spent_amount:,.2f} ₪",
                    f"{budget_amount:,.2f} ₪",
                    format_text(cat_name)
                ])
            
            # שורת סיכום
            total_usage = (total_spent / total_budget * 100) if total_budget > 0 else 0
            budget_summary_data.append([
                f"{total_usage:.1f}%",
                f"{total_remaining:,.2f} ₪",
                f"{total_spent:,.2f} ₪",
                f"{total_budget:,.2f} ₪",
                format_text("סה\"כ")
            ])
            
            budget_summary_table = Table(budget_summary_data, colWidths=[70, 90, 90, 90, 110], style=[
                ('FONT', (0, 0), (-1, -1), font_name),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#4C1D95')),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4C1D95')),  # Purple-900 header
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EDE9FE')),  # Purple-100 for total row
                ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#F5F3FF')),  # Purple-50 for data rows
                # Align calculation columns (0-3) center, category column (4) right
                ('ALIGN', (0, 0), (3, -1), 'CENTER'),
                ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('PADDING', (0, 0), (-1, -1), 8),
                ('FONTSIZE', (0, -1), (-1, -1), 11),  # Larger font for total row
            ])
            elements.append(budget_summary_table)
            elements.append(Spacer(1, 20))

        # Summary - סיכום פיננסי - Professional Summary Card
        if options.include_summary and summary:
            elements.append(Paragraph(format_text(f"💰 {REPORT_LABELS['financial_summary']}"), style_h2))
            elements.append(Spacer(1, 8))
            
            # Reversed for Hebrew RTL: amount on left, details on right
            summary_data = [
                [format_text(REPORT_LABELS['amount']), format_text(REPORT_LABELS['details'])],
                [f"{summary['income']:,.2f} ₪", format_text(f"↗️ {REPORT_LABELS['total_income']}")],
                [f"{summary['expenses']:,.2f} ₪", format_text(f"↘️ {REPORT_LABELS['total_expenses']}")],
                [f"{summary['profit']:,.2f} ₪", format_text(f"📈 {REPORT_LABELS['balance_profit']}")],
            ]
            
            # Dynamic styling based on profit/loss
            profit_color = COLOR_ACCENT_EMERALD if summary['profit'] >= 0 else COLOR_ACCENT_ROSE
            profit_bg = '#D1FAE5' if summary['profit'] >= 0 else '#FFE4E6'  # Emerald-100 or Rose-100
            
            summary_table = Table(summary_data, colWidths=[160, 220])
            summary_table.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLOR_PRIMARY_DARK)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('TOPPADDING', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 14),
                # Income row - green
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#D1FAE5')),
                ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor(COLOR_ACCENT_EMERALD)),
                # Expense row - red
                ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#FFE4E6')),
                ('TEXTCOLOR', (1, 2), (1, 2), colors.HexColor(COLOR_ACCENT_ROSE)),
                # Profit row - conditional
                ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor(profit_bg)),
                ('TEXTCOLOR', (1, 3), (1, 3), colors.HexColor(profit_color)),
                # General styling - align amount (col 0) left, details (col 1) right
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 25))

        # Monthly Breakdown Table - דוח חודשי
        if monthly_breakdown and len(monthly_breakdown) > 0:
            elements.append(Paragraph(format_text(f"📅 דוח חודשי - הוצאות לפי קטגוריה וספק"), style_h2))
            elements.append(Spacer(1, 12))
            
            # Prepare monthly breakdown table data - reversed for Hebrew RTL: category/supplier on right, date/amount on left
            monthly_table_data = [
                [
                    format_text(REPORT_LABELS['amount']),
                    format_text(REPORT_LABELS['supplier']),
                    format_text(REPORT_LABELS['category']),
                    format_text(REPORT_LABELS.get('date', 'חודש'))
                ]
            ]
            
            for row_data in monthly_breakdown:
                # Format month as Hebrew date (YYYY-MM -> MM/YYYY)
                month_str = row_data['month']
                try:
                    year, month = month_str.split('-')
                    month_display = f"{month}/{year}"
                except:
                    month_display = month_str
                
                cat_name = row_data['category'] or REPORT_LABELS['general']
                supplier_name = row_data['supplier'] or "ללא ספק"
                amount = row_data['amount']
                
                monthly_table_data.append([
                    f"{amount:,.2f} ₪",
                    format_text(supplier_name),
                    format_text(cat_name),
                    format_text(month_display)
                ])
            
            monthly_table = Table(monthly_table_data, colWidths=[120, 150, 150, 100])
            monthly_table.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLOR_PRIMARY_DARK)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                # Data rows
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                # Align amount (col 0) left, supplier/category/date (cols 1-3) right
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (3, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(monthly_table)
            elements.append(Spacer(1, 25))

        # ========== עמוד שני והלאה: עסקאות ==========
        
        # Transactions - Group by category and create separate tables
        if options.include_transactions and transactions:
            # מעבר לעמוד חדש לפני העסקאות
            elements.append(PageBreak())
            
            # Section header with decorative line
            elements.append(Table([[""]], colWidths=[520], rowHeights=[3], style=[
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(COLOR_ACCENT_TEAL))
            ]))
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(format_text(f"📝 {REPORT_LABELS['transaction_details']}"), style_h2))
            elements.append(Spacer(1, 12))

            # Group transactions by year first, then by category
            from collections import defaultdict
            transactions_by_year_and_category = defaultdict(lambda: defaultdict(list))
            # Get selected categories if any
            selected_categories = set(options.categories) if options.categories and len(options.categories) > 0 else None

            for tx in transactions:
                if isinstance(tx, dict):
                    cat_name = tx.get('category') or REPORT_LABELS['general']
                    tx_date = tx.get('tx_date')
                else:
                    cat_name = tx.category.name if tx.category else REPORT_LABELS['general']
                    tx_date = tx.tx_date
                
                # Extract year from transaction date
                if isinstance(tx_date, str):
                    try:
                        if 'T' in tx_date:
                            tx_date = date.fromisoformat(tx_date.split('T')[0])
                        else:
                            tx_date = date.fromisoformat(tx_date)
                    except:
                        # If parsing fails, use current year
                        tx_date = date.today()
                elif not isinstance(tx_date, date):
                    tx_date = date.today()
                
                year = tx_date.year

                # Only include transactions from selected categories if categories were selected
                if selected_categories is None or cat_name in selected_categories:
                    transactions_by_year_and_category[year][cat_name].append(tx)

            # Sort years in descending order (newest first)
            sorted_years = sorted(transactions_by_year_and_category.keys(), reverse=True)

            # Create tables for each year, then each category within that year
            for year in sorted_years:
                # Year header
                elements.append(PageBreak() if year != sorted_years[0] else Spacer(1, 0))
                year_header_style = ParagraphStyle(
                    'YearHeader', 
                    parent=styles['Heading1'], 
                    fontName=font_name, 
                    fontSize=16, 
                    alignment=1,
                    textColor=colors.white,
                    backColor=colors.HexColor(COLOR_PRIMARY_DARK),
                    leading=24, 
                    spaceBefore=15, 
                    spaceAfter=10,
                    leftIndent=10,
                    rightIndent=10,
                    borderPadding=8
                )
                elements.append(Paragraph(format_text(f"שנת {year}"), year_header_style))
                elements.append(Spacer(1, 12))
                
                # Get categories for this year, sorted alphabetically
                year_categories = sorted(transactions_by_year_and_category[year].keys())
                
                for cat_name in year_categories:
                    cat_transactions = transactions_by_year_and_category[year][cat_name]
                # Category header with amber accent
                elements.append(Paragraph(format_text(f"📁 {REPORT_LABELS['category']}: {cat_name}"), style_category))
                elements.append(Spacer(1, 6))
                
                # Check which columns have data
                has_suppliers = any(
                    (tx.get('supplier_name') if isinstance(tx, dict) else (tx.supplier.name if tx.supplier else None))
                    for tx in cat_transactions
                ) if cat_transactions else False
                
                has_descriptions = any(
                    (tx.get('description') if isinstance(tx, dict) else tx.description)
                    for tx in cat_transactions
                ) if cat_transactions else False
                
                # Build dynamic columns list based on available data
                columns = ['date', 'type', 'amount']
                if has_suppliers:
                    columns.append('supplier')
                if has_descriptions:
                    columns.append('description')
                
                # Column widths based on which columns are shown
                col_width_map = {
                    'date': 70,
                    'type': 55,
                    'amount': 70,
                    'supplier': 120,
                    'description': 200
                }
                
                # Adjust description width if no supplier
                if has_descriptions and not has_suppliers:
                    col_width_map['description'] = 275
                
                col_widths = [col_width_map[col] for col in columns]
                
                # Build table headers dynamically
                col_to_label = {
                    'date': REPORT_LABELS['date'],
                    'type': REPORT_LABELS['type'],
                    'amount': REPORT_LABELS['amount'],
                    'supplier': REPORT_LABELS['supplier'],
                    'description': REPORT_LABELS['description']
                }
                
                tx_data = [[Paragraph(format_text(col_to_label[col]), style_table_cell) for col in columns]]
                
                # Track category totals
                cat_total_income = 0
                cat_total_expense = 0
                
                # Add transaction rows
                for tx in cat_transactions:
                    if isinstance(tx, dict):
                        is_income = tx.get('type') == "Income"
                        tx_type = REPORT_LABELS['income'] if is_income else REPORT_LABELS['expense']
                        tx_desc = tx.get('description') or ""
                        supplier_name = tx.get('supplier_name') or ""
                        tx_date = tx.get('tx_date')
                        tx_amount = tx.get('amount', 0)
                    else:
                        is_income = tx.type == "Income"
                        tx_type = REPORT_LABELS['income'] if is_income else REPORT_LABELS['expense']
                        tx_desc = tx.description or ""
                        supplier_name = tx.supplier.name if tx.supplier else ""
                        tx_date = tx.tx_date
                        tx_amount = tx.amount
                    
                    # Track totals
                    if is_income:
                        cat_total_income += tx_amount
                    else:
                        cat_total_expense += tx_amount
                    
                    col_to_value = {
                        'date': format_date_hebrew(tx_date),
                        'type': format_text(tx_type),
                        'amount': f"{tx_amount:,.2f} ₪",
                        'supplier': format_text(supplier_name),
                        'description': format_text(tx_desc)
                    }
                    
                    tx_data.append([Paragraph(col_to_value[col], style_table_cell) for col in columns])
                
                # Build table style with alternating rows
                tx_style = [
                    ('FONT', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                    # Header row
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLOR_PRIMARY_MID)),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('TOPPADDING', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    # General
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('PADDING', (0, 0), (-1, -1), 6),
                ]
                
                # Add alternating row colors for data rows
                for row_idx in range(1, len(tx_data)):
                    if row_idx % 2 == 1:
                        tx_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor(COLOR_BG_LIGHT)))
                    else:
                        tx_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor(COLOR_BG_ALT)))
                
                tx_table = Table(tx_data, repeatRows=1, colWidths=col_widths, style=tx_style)
                elements.append(tx_table)
                
                # Category summary row
                cat_net = cat_total_income - cat_total_expense
                summary_color = COLOR_ACCENT_EMERALD if cat_net >= 0 else COLOR_ACCENT_ROSE
                summary_bg = '#FEF3C7'  # Amber-100
                
                cat_summary_data = [[
                    format_text(f"סה״כ {cat_name}:"),
                    f"{cat_net:,.2f} ₪"
                ]]
                cat_summary_table = Table(cat_summary_data, colWidths=[300, 160])
                cat_summary_table.setStyle(TableStyle([
                    ('FONT', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(summary_bg)),
                    ('TEXTCOLOR', (1, 0), (1, 0), colors.HexColor(summary_color)),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('PADDING', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ]))
                elements.append(cat_summary_table)
                elements.append(Spacer(1, 18))  # Space between category tables

            # ========== COMPREHENSIVE PERIOD SUMMARY ==========
            # Add a comprehensive summary table at the end showing all categories
            elements.append(PageBreak())
            elements.append(Table([[""]], colWidths=[520], rowHeights=[3], style=[
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(COLOR_ACCENT_TEAL))
            ]))
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(format_text("📊 סיכום כולל לתקופה"), style_h2))
            elements.append(Spacer(1, 12))
            
            # Calculate totals by category across all years
            category_totals = {}
            for year in sorted_years:
                for cat_name in transactions_by_year_and_category[year].keys():
                    if cat_name not in category_totals:
                        category_totals[cat_name] = {'income': 0, 'expense': 0}
                    
                    for tx in transactions_by_year_and_category[year][cat_name]:
                        if isinstance(tx, dict):
                            tx_type = tx.get('type')
                            tx_amount = tx.get('amount', 0)
                        else:
                            tx_type = tx.type
                            tx_amount = tx.amount
                        
                        if tx_type == "Income":
                            category_totals[cat_name]['income'] += float(tx_amount)
                        else:
                            category_totals[cat_name]['expense'] += float(tx_amount)
            
            # Create category expense summary table
            elements.append(Paragraph(format_text("הוצאות לפי קטגוריה"), style_category))
            elements.append(Spacer(1, 6))
            
            # Reversed for Hebrew RTL: category on right, calculations on left
            cat_expense_data = [[
                format_text("נטו"),
                format_text(REPORT_LABELS['income']),
                format_text(REPORT_LABELS['expenses']),
                format_text(REPORT_LABELS['category'])
            ]]
            
            grand_total_income = 0
            grand_total_expense = 0
            
            for cat_name, totals in sorted(category_totals.items()):
                cat_income = totals['income']
                cat_expense = totals['expense']
                cat_net = cat_income - cat_expense
                
                grand_total_income += cat_income
                grand_total_expense += cat_expense
                
                cat_expense_data.append([
                    f"{cat_net:,.2f} ₪",
                    f"{cat_income:,.2f} ₪",
                    f"{cat_expense:,.2f} ₪",
                    format_text(cat_name)
                ])
            
            # Grand total row
            grand_net = grand_total_income - grand_total_expense
            cat_expense_data.append([
                f"{grand_net:,.2f} ₪",
                f"{grand_total_income:,.2f} ₪",
                f"{grand_total_expense:,.2f} ₪",
                format_text("סה״כ")
            ])
            
            cat_expense_table = Table(cat_expense_data, colWidths=[110, 110, 110, 150])
            
            cat_expense_style = [
                ('FONT', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLOR_PRIMARY_MID)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                # Align calculation columns (0-2) center, category column (3) right
                ('ALIGN', (0, 0), (2, -1), 'CENTER'),
                ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                # Total row styling
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#CCFBF1')),
                ('FONTSIZE', (0, -1), (-1, -1), 11),
            ]
            
            # Add alternating colors for data rows
            for row_idx in range(1, len(cat_expense_data) - 1):  # Exclude header and total
                if row_idx % 2 == 1:
                    cat_expense_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor(COLOR_BG_LIGHT)))
                else:
                    cat_expense_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor(COLOR_BG_ALT)))
            
            cat_expense_table.setStyle(TableStyle(cat_expense_style))
            elements.append(cat_expense_table)
            elements.append(Spacer(1, 25))
            
            # Add budget vs actual summary if budgets are available
            if budgets and len(budgets) > 0:
                elements.append(Paragraph(format_text("תקציב מול ביצוע לתקופה"), style_category))
                elements.append(Spacer(1, 6))
                
                budget_summary_data = [[
                    format_text(REPORT_LABELS['category']),
                    format_text(REPORT_LABELS['budget']),
                    format_text(REPORT_LABELS['used']),
                    format_text(REPORT_LABELS['remaining']),
                    format_text("ניצול %")
                ]]
                
                total_budget = 0
                total_spent = 0
                
                for b in budgets:
                    cat_name = b['category'] if b['category'] else REPORT_LABELS['general']
                    budget_amount = b['amount']
                    spent_amount = b['spent_amount']
                    remaining = b['remaining_amount']
                    usage_pct = (spent_amount / budget_amount * 100) if budget_amount > 0 else 0
                    
                    total_budget += budget_amount
                    total_spent += spent_amount
                    
                    budget_summary_data.append([
                        format_text(cat_name),
                        f"{budget_amount:,.2f} ₪",
                        f"{spent_amount:,.2f} ₪",
                        f"{remaining:,.2f} ₪",
                        f"{usage_pct:.1f}%"
                    ])
                
                # Total row
                total_remaining = total_budget - total_spent
                total_usage = (total_spent / total_budget * 100) if total_budget > 0 else 0
                budget_summary_data.append([
                    format_text("סה״כ"),
                    f"{total_budget:,.2f} ₪",
                    f"{total_spent:,.2f} ₪",
                    f"{total_remaining:,.2f} ₪",
                    f"{total_usage:.1f}%"
                ])
                
                budget_summary_table = Table(budget_summary_data, colWidths=[110, 95, 95, 95, 70])
                
                budget_summary_style = [
                    ('FONT', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLOR_PRIMARY_MID)),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('PADDING', (0, 0), (-1, -1), 10),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#CCFBF1')),
                ]
                
                # Add conditional coloring for budget rows
                for row_idx, b in enumerate(budgets):
                    row = row_idx + 1
                    usage_pct = (b['spent_amount'] / b['amount'] * 100) if b['amount'] > 0 else 0
                    remaining = b['remaining_amount']
                    
                    if row_idx % 2 == 0:
                        budget_summary_style.append(('BACKGROUND', (0, row), (-1, row), colors.HexColor(COLOR_BG_LIGHT)))
                    else:
                        budget_summary_style.append(('BACKGROUND', (0, row), (-1, row), colors.HexColor(COLOR_BG_ALT)))
                    
                    if usage_pct > 100:
                        budget_summary_style.append(('TEXTCOLOR', (2, row), (2, row), colors.HexColor(COLOR_ACCENT_ROSE)))
                        budget_summary_style.append(('BACKGROUND', (0, row), (-1, row), colors.HexColor('#FFE4E6')))
                    
                    if remaining < 0:
                        budget_summary_style.append(('TEXTCOLOR', (3, row), (3, row), colors.HexColor(COLOR_ACCENT_ROSE)))
                    else:
                        budget_summary_style.append(('TEXTCOLOR', (3, row), (3, row), colors.HexColor(COLOR_ACCENT_EMERALD)))
                
                budget_summary_table.setStyle(TableStyle(budget_summary_style))
                elements.append(budget_summary_table)
                elements.append(Spacer(1, 25))
            
            # Final overall summary
            elements.append(Paragraph(format_text("סיכום פיננסי כולל"), style_category))
            elements.append(Spacer(1, 6))
            
            final_summary_data = [
                [format_text("פרט"), format_text(REPORT_LABELS['amount'])],
                [format_text(f"↗️ {REPORT_LABELS['total_income']}"), f"{grand_total_income:,.2f} ₪"],
                [format_text(f"↘️ {REPORT_LABELS['total_expenses']}"), f"{grand_total_expense:,.2f} ₪"],
                [format_text(f"📈 {REPORT_LABELS['balance_profit']}"), f"{grand_net:,.2f} ₪"],
            ]
            
            profit_color = COLOR_ACCENT_EMERALD if grand_net >= 0 else COLOR_ACCENT_ROSE
            profit_bg = '#D1FAE5' if grand_net >= 0 else '#FFE4E6'
            
            final_summary_table = Table(final_summary_data, colWidths=[220, 160])
            final_summary_table.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLOR_PRIMARY_DARK)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#D1FAE5')),
                ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor(COLOR_ACCENT_EMERALD)),
                ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#FFE4E6')),
                ('TEXTCOLOR', (1, 2), (1, 2), colors.HexColor(COLOR_ACCENT_ROSE)),
                ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor(profit_bg)),
                ('TEXTCOLOR', (1, 3), (1, 3), colors.HexColor(profit_color)),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 12),
            ]))
            elements.append(final_summary_table)
            elements.append(Spacer(1, 25))

        # Charts
        if options.include_charts:
            CHART_TITLES = {
                "income_expense_pie": "עוגת הכנסות מול הוצאות",
                "expense_by_category_pie": "הוצאות לפי קטגוריה (גרף עוגה)",
                "expense_by_category_bar": "הוצאות לפי קטגוריה (גרף עמודות)",
                "trends_line": "מגמות הכנסות/הוצאות לאורך זמן",
                "expense_by_supplier_bar": "הוצאות לפי ספק (10 הגדולים)",
                "monthly_trends_line": "מגמות חודשיות - הכנסות והוצאות",
                "budget_vs_actual": "תקציב מול ביצוע לפי קטגוריה",
                "cumulative_expenses": "הוצאות מצטברות לאורך זמן"
            }
            
            # Check if there's any data to chart
            has_financial_data = summary and (summary.get('income', 0) > 0 or summary.get('expenses', 0) > 0)
            has_transactions = transactions and len(transactions) > 0
            
            if not has_financial_data and not has_transactions:
                # No data to chart - skip charts section entirely
                print("INFO: No financial data for charts, skipping charts section")
            else:
                elements.append(Spacer(1, 20))
                elements.append(Paragraph(format_text("גרפים"), style_h2))
                elements.append(Spacer(1, 10))

                charts_to_render = {}
                
                # פונקציה לבדיקת תקינות תמונת PNG
                def is_valid_png(data: bytes) -> bool:
                    if not data or len(data) < 100:
                        return False
                    # בדיקת חתימת PNG
                    png_signature = b'\x89PNG\r\n\x1a\n'
                    return data[:8] == png_signature

                # Track which chart types we already have from frontend
                frontend_chart_types = set()
                
                # Use provided images if available - but validate them first
                if chart_images:
                    for key, img_data in chart_images.items():
                        if is_valid_png(img_data):
                            hebrew_title = CHART_TITLES.get(key, key)
                            charts_to_render[hebrew_title] = img_data
                            frontend_chart_types.add(key)
                            print(f"INFO: Valid frontend chart image: {key} ({len(img_data)} bytes)")
                        else:
                            print(f"WARNING: Invalid frontend chart image: {key}, will regenerate")
                
                # Get requested chart types
                chart_types = options.chart_types or [
                    "income_expense_pie",
                    "expense_by_category_pie",
                    "expense_by_category_bar",
                    "trends_line",
                    "expense_by_supplier_bar",
                    "monthly_trends_line",
                    "budget_vs_actual",
                    "cumulative_expenses"
                ]
                
                # Generate server-side charts for types not provided by frontend
                for chart_type in chart_types:
                    # Skip if we already have this chart from frontend
                    if chart_type in frontend_chart_types:
                        print(f"INFO: Skipping {chart_type} - already have from frontend")
                        continue
                    
                    try:
                        print(f"INFO: Creating chart: {chart_type}")
                        chart_buffer = self._create_chart_image(chart_type, {}, summary, transactions, budgets)
                        if chart_buffer:
                            chart_buffer.seek(0)
                            chart_bytes = chart_buffer.read()
                            if chart_bytes and len(chart_bytes) > 100:  # Ensure we have actual image data
                                chart_name = CHART_TITLES.get(chart_type, chart_type)
                                charts_to_render[chart_name] = chart_bytes
                    except Exception as e:
                        print(f"אזהרה: שגיאה בהכנת גרף {chart_type}: {e}")

                if charts_to_render:
                    print(f"INFO: Rendering {len(charts_to_render)} charts to PDF")
                    for chart_name, image_bytes in charts_to_render.items():
                        try:
                            # Validate image data
                            if not image_bytes or len(image_bytes) < 100:
                                print(f"אזהרה: נתוני תמונה לא תקינים עבור {chart_name}")
                                continue
                            
                            # בדיקת חתימת PNG
                            png_signature = b'\x89PNG\r\n\x1a\n'
                            if image_bytes[:8] != png_signature:
                                print(f"אזהרה: התמונה {chart_name} אינה PNG תקינה, מדלג...")
                                continue
                            
                            print(f"INFO: Adding chart '{chart_name}' to PDF ({len(image_bytes)} bytes)")
                            
                            # יצירת buffer חדש עם הנתונים
                            img_buffer = BytesIO(image_bytes)
                            img_buffer.seek(0)
                            
                            # יצירת אובייקט תמונה עם יחס גובה-רוחב נכון
                            # A4 width = 595pt, margins = 60pt, usable width = 535pt
                            # Using 500pt width to have some padding
                            img = RLImage(img_buffer, width=500, height=375, kind='proportional')
                            img.hAlign = 'CENTER'
                            
                            # כותרת מקצועית עם רקע
                            chart_title_style = ParagraphStyle(
                                'ChartTitle',
                                parent=style_normal,
                                fontSize=12,
                                fontName=font_name,
                                textColor=colors.HexColor(COLOR_ACCENT_TEAL),
                                alignment=1,
                                spaceAfter=5,
                                spaceBefore=10,
                                fontStyle='bold'
                            )
                            elements.append(Paragraph(format_text(f"📊 {chart_name}"), chart_title_style))
                            elements.append(Spacer(1, 5))
                            elements.append(img)
                            elements.append(Spacer(1, 30))
                            print(f"INFO: Chart '{chart_name}' added successfully")
                        except Exception as e:
                            print(f"אזהרה: הוספת גרף {chart_name} ל-PDF נכשלה: {e}")
                            import traceback
                            traceback.print_exc()
                else:
                    # No charts were generated
                    print("WARNING: No charts were generated or all failed")
                    elements.append(Paragraph(format_text("אין נתונים להצגת גרפים"), style_normal))

        doc.build(elements)
        buffer.seek(0)
        return buffer.read()


    async def _generate_excel(self, project, options, transactions, budgets, fund, summary, chart_images=None, monthly_breakdown=None) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "דוח"
        ws.sheet_view.rightToLeft = True

        # ===== PROFESSIONAL MODERN STYLING =====
        # Color Palette - Sophisticated dark teal/slate theme
        PRIMARY_DARK = "0F172A"      # Slate-900 - Main dark background
        PRIMARY_MID = "1E293B"       # Slate-800 - Secondary dark
        ACCENT_TEAL = "0D9488"       # Teal-600 - Accent color
        ACCENT_EMERALD = "059669"    # Emerald-600 - Success/Income
        ACCENT_ROSE = "E11D48"       # Rose-600 - Expense/Alert
        ACCENT_AMBER = "D97706"      # Amber-600 - Warning/Category headers
        
        # Light backgrounds for data rows
        BG_LIGHT = "F8FAFC"          # Slate-50 - Light row
        BG_ALT = "F1F5F9"            # Slate-100 - Alternate row
        BG_TEAL_LIGHT = "CCFBF1"     # Teal-100 - Teal highlight
        BG_EMERALD_LIGHT = "D1FAE5"  # Emerald-100 - Green highlight
        BG_ROSE_LIGHT = "FFE4E6"     # Rose-100 - Red highlight
        BG_AMBER_LIGHT = "FEF3C7"    # Amber-100 - Amber highlight
        
        TEXT_DARK = "0F172A"         # Dark text
        TEXT_LIGHT = "FFFFFF"        # White text
        TEXT_MUTED = "64748B"        # Slate-500 - Muted text
        
        # Typography
        title_font = Font(name='Arial', bold=True, size=18, color=TEXT_LIGHT)
        subtitle_font = Font(name='Arial', size=11, color=TEXT_MUTED, italic=True)
        h2_font = Font(name='Arial', bold=True, size=13, color=TEXT_LIGHT)
        header_font = Font(name='Arial', bold=True, size=10, color=TEXT_LIGHT)
        data_font = Font(name='Arial', size=10, color=TEXT_DARK)
        data_bold_font = Font(name='Arial', bold=True, size=10, color=TEXT_DARK)
        money_positive_font = Font(name='Arial', bold=True, size=10, color=ACCENT_EMERALD)
        money_negative_font = Font(name='Arial', bold=True, size=10, color=ACCENT_ROSE)
        
        # Fills
        fill_title = PatternFill(start_color=PRIMARY_DARK, end_color=PRIMARY_DARK, fill_type="solid")
        fill_h2 = PatternFill(start_color=ACCENT_TEAL, end_color=ACCENT_TEAL, fill_type="solid")
        fill_header = PatternFill(start_color=PRIMARY_MID, end_color=PRIMARY_MID, fill_type="solid")
        fill_category_header = PatternFill(start_color=ACCENT_AMBER, end_color=ACCENT_AMBER, fill_type="solid")
        fill_light = PatternFill(start_color=BG_LIGHT, end_color=BG_LIGHT, fill_type="solid")
        fill_alt = PatternFill(start_color=BG_ALT, end_color=BG_ALT, fill_type="solid")
        fill_teal_light = PatternFill(start_color=BG_TEAL_LIGHT, end_color=BG_TEAL_LIGHT, fill_type="solid")
        fill_emerald_light = PatternFill(start_color=BG_EMERALD_LIGHT, end_color=BG_EMERALD_LIGHT, fill_type="solid")
        fill_rose_light = PatternFill(start_color=BG_ROSE_LIGHT, end_color=BG_ROSE_LIGHT, fill_type="solid")
        fill_amber_light = PatternFill(start_color=BG_AMBER_LIGHT, end_color=BG_AMBER_LIGHT, fill_type="solid")
        
        # Borders - Subtle and elegant
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        medium_border = Border(
            left=Side(style='medium', color=PRIMARY_MID),
            right=Side(style='medium', color=PRIMARY_MID),
            top=Side(style='medium', color=PRIMARY_MID),
            bottom=Side(style='medium', color=PRIMARY_MID)
        )
        bottom_accent = Border(
            bottom=Side(style='medium', color=ACCENT_TEAL)
        )
        
        # Alignment
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        current_row = 1
        
        # Set default column widths for professional look
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 30
        
        # ===== TITLE SECTION =====
        # Create a professional header spanning multiple columns
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws.row_dimensions[current_row].height = 40
        title_cell = ws[f'A{current_row}']
        title_cell.value = f"📊  {REPORT_LABELS['project_report']}: {project.name}"
        title_cell.font = title_font
        title_cell.fill = fill_title
        title_cell.alignment = center_align
        title_cell.border = medium_border
        current_row += 1

        # Date range subtitle
        date_range_text = ""
        if options.start_date and options.end_date:
            date_range_text = f"{options.start_date.strftime('%d/%m/%Y')} - {options.end_date.strftime('%d/%m/%Y')}"
        elif options.start_date:
            date_range_text = f"מ-{options.start_date.strftime('%d/%m/%Y')}"
        elif options.end_date:
            date_range_text = f"עד {options.end_date.strftime('%d/%m/%Y')}"
        else:
            date_range_text = "כל התקופות"
        
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws.row_dimensions[current_row].height = 25
        date_range_cell = ws[f'A{current_row}']
        date_range_cell.value = f"תקופה: {date_range_text}"
        date_range_cell.font = subtitle_font
        date_range_cell.fill = fill_light
        date_range_cell.alignment = center_align
        date_range_cell.border = bottom_accent
        current_row += 1
        
        # Production date
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws.row_dimensions[current_row].height = 25
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"{REPORT_LABELS['production_date']}: {date.today().strftime('%d/%m/%Y')}"
        date_cell.font = subtitle_font
        date_cell.fill = fill_light
        date_cell.alignment = center_align
        date_cell.border = bottom_accent
        current_row += 2  # Spacer

        # 1. Summary - Professional Financial Summary Card
        if options.include_summary and summary:
            # Section header with icon
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws.row_dimensions[current_row].height = 30
            summary_header = ws[f'A{current_row}']
            summary_header.value = f"💰  {REPORT_LABELS['financial_summary']}"
            summary_header.font = h2_font
            summary_header.fill = fill_h2
            summary_header.alignment = center_align
            summary_header.border = medium_border
            current_row += 1

            # Summary table headers - reversed for Hebrew RTL: amount on left, details on right
            ws.row_dimensions[current_row].height = 28
            ws[f'A{current_row}'] = REPORT_LABELS['amount']
            ws[f'B{current_row}'] = REPORT_LABELS['details']
            for col in ['A', 'B']:
                cell = ws[f'{col}{current_row}']
                cell.font = header_font
                cell.fill = fill_header
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1

            # Income row - with green highlight
            ws.row_dimensions[current_row].height = 25
            ws[f'A{current_row}'] = f"{summary['income']:,.2f} ₪"
            ws[f'A{current_row}'].font = money_positive_font
            ws[f'A{current_row}'].fill = fill_emerald_light
            ws[f'B{current_row}'] = f"↗️  {REPORT_LABELS['total_income']}"
            ws[f'B{current_row}'].font = data_bold_font
            ws[f'B{current_row}'].fill = fill_emerald_light
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'B{current_row}'].border = thin_border
            ws[f'B{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
            current_row += 1

            # Expenses row - with red highlight
            ws.row_dimensions[current_row].height = 25
            ws[f'A{current_row}'] = f"{summary['expenses']:,.2f} ₪"
            ws[f'A{current_row}'].font = money_negative_font
            ws[f'A{current_row}'].fill = fill_rose_light
            ws[f'B{current_row}'] = f"↘️  {REPORT_LABELS['total_expenses']}"
            ws[f'B{current_row}'].font = data_bold_font
            ws[f'B{current_row}'].fill = fill_rose_light
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'B{current_row}'].border = thin_border
            ws[f'B{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
            current_row += 1

            # Profit/Loss row - conditional coloring
            ws.row_dimensions[current_row].height = 28
            ws[f'A{current_row}'] = f"{summary['profit']:,.2f} ₪"
            ws[f'B{current_row}'] = f"📈  {REPORT_LABELS['balance_profit']}"
            ws[f'B{current_row}'].font = data_bold_font
            
            # Color based on profit/loss
            if summary['profit'] >= 0:
                ws[f'A{current_row}'].fill = fill_teal_light
                ws[f'B{current_row}'].fill = fill_teal_light
                ws[f'A{current_row}'].font = money_positive_font
            else:
                ws[f'A{current_row}'].fill = fill_rose_light
                ws[f'B{current_row}'].fill = fill_rose_light
                ws[f'A{current_row}'].font = money_negative_font
            
            ws[f'A{current_row}'].border = medium_border
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'B{current_row}'].border = medium_border
            ws[f'B{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
            current_row += 2  # Spacer

        # 2. Fund - Professional Fund Status Card
        if options.include_funds and fund:
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws.row_dimensions[current_row].height = 30
            fund_header = ws[f'A{current_row}']
            fund_header.value = f"🏦  {REPORT_LABELS['fund_status']}"
            fund_header.font = h2_font
            fund_header.fill = fill_h2
            fund_header.alignment = center_align
            fund_header.border = medium_border
            current_row += 1

            # Current balance row - reversed for Hebrew RTL: value on left, label on right
            ws.row_dimensions[current_row].height = 28
            ws[f'A{current_row}'] = f"{fund.current_balance:,.2f} ₪"
            ws[f'A{current_row}'].font = money_positive_font
            ws[f'A{current_row}'].fill = fill_teal_light
            ws[f'B{current_row}'] = f"💵  {REPORT_LABELS['current_balance']}"
            ws[f'B{current_row}'].font = data_bold_font
            ws[f'B{current_row}'].fill = fill_teal_light
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'B{current_row}'].border = thin_border
            ws[f'B{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
            current_row += 1

            # Monthly deposit row
            ws.row_dimensions[current_row].height = 25
            ws[f'A{current_row}'] = f"{fund.monthly_amount:,.2f} ₪"
            ws[f'A{current_row}'].font = data_bold_font
            ws[f'A{current_row}'].fill = fill_light
            ws[f'B{current_row}'] = f"📅  {REPORT_LABELS['monthly_deposit']}"
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = fill_light
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'B{current_row}'].border = thin_border
            ws[f'B{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
            current_row += 2  # Spacer

        # 3. Budgets - Professional Budget Table
        if options.include_budgets and budgets:
            ws.merge_cells(f'A{current_row}:D{current_row}')
            ws.row_dimensions[current_row].height = 30
            budget_header = ws[f'A{current_row}']
            budget_header.value = f"📋  {REPORT_LABELS['budget_vs_actual']}"
            budget_header.font = h2_font
            budget_header.fill = fill_h2
            budget_header.alignment = center_align
            budget_header.border = medium_border
            current_row += 1

            # Budget table headers - reversed for Hebrew RTL: category on right, calculations on left
            ws.row_dimensions[current_row].height = 28
            budget_headers = [
                REPORT_LABELS['remaining'],
                REPORT_LABELS['used'],
                REPORT_LABELS['budget'],
                REPORT_LABELS['category']
            ]
            for idx, header in enumerate(budget_headers):
                col = get_column_letter(idx + 1)
                cell = ws[f'{col}{current_row}']
                cell.value = header
                cell.font = header_font
                cell.fill = fill_header
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1

            # Budget data rows with alternating colors
            for row_idx, b in enumerate(budgets):
                ws.row_dimensions[current_row].height = 24
                cat_name = b['category'] if b['category'] else REPORT_LABELS['general']
                remaining = b['remaining_amount']
                usage_percent = (b['spent_amount'] / b['amount'] * 100) if b['amount'] > 0 else 0
                
                # Row fill - alternate colors
                row_fill = fill_light if row_idx % 2 == 0 else fill_alt
                
                # Column order: D=category, C=budget, B=used, A=remaining
                ws[f'D{current_row}'] = cat_name
                ws[f'D{current_row}'].font = data_bold_font
                ws[f'D{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
                
                ws[f'C{current_row}'] = f"{b['amount']:,.2f} ₪"
                ws[f'C{current_row}'].font = data_font
                ws[f'C{current_row}'].alignment = center_align
                
                ws[f'B{current_row}'] = f"{b['spent_amount']:,.2f} ₪"
                # Color based on usage
                if usage_percent > 100:
                    ws[f'B{current_row}'].font = money_negative_font
                elif usage_percent > 80:
                    ws[f'B{current_row}'].font = Font(name='Arial', bold=True, size=10, color=ACCENT_AMBER)
                else:
                    ws[f'B{current_row}'].font = data_font
                ws[f'B{current_row}'].alignment = center_align
                
                ws[f'A{current_row}'] = f"{remaining:,.2f} ₪"
                # Color remaining based on positive/negative
                if remaining < 0:
                    ws[f'A{current_row}'].font = money_negative_font
                    row_fill = fill_rose_light
                else:
                    ws[f'A{current_row}'].font = money_positive_font
                ws[f'A{current_row}'].alignment = center_align
                
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{current_row}'].fill = row_fill
                    ws[f'{col}{current_row}'].border = thin_border
                # Override alignment: D (category) right, others center
                ws[f'D{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
                current_row += 1

            current_row += 1  # Spacer

        # 4. Monthly Breakdown Table - Professional monthly breakdown by category and supplier
        # Always show the table if we have monthly breakdown data
        if monthly_breakdown and len(monthly_breakdown) > 0:
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws.row_dimensions[current_row].height = 30
            monthly_header = ws[f'A{current_row}']
            monthly_header.value = f"📅  דוח חודשי - הוצאות לפי קטגוריה וספק"
            monthly_header.font = h2_font
            monthly_header.fill = fill_h2
            monthly_header.alignment = center_align
            monthly_header.border = medium_border
            current_row += 1

            # Monthly breakdown table headers - reversed for Hebrew RTL: category/supplier/date on right, amount on left
            ws.row_dimensions[current_row].height = 28
            monthly_headers = [
                REPORT_LABELS['amount'],
                REPORT_LABELS['supplier'],
                REPORT_LABELS['category'],
                REPORT_LABELS['date'] if 'date' in REPORT_LABELS else 'חודש'
            ]
            for idx, header in enumerate(monthly_headers):
                col = get_column_letter(idx + 1)
                cell = ws[f'{col}{current_row}']
                cell.value = header
                cell.font = header_font
                cell.fill = fill_header
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1

            # Monthly breakdown data rows with alternating colors
            for row_idx, row_data in enumerate(monthly_breakdown):
                ws.row_dimensions[current_row].height = 24
                
                # Format month as Hebrew date (YYYY-MM -> MM/YYYY)
                month_str = row_data['month']
                try:
                    year, month = month_str.split('-')
                    # Convert to Hebrew month format
                    month_display = f"{month}/{year}"
                except:
                    month_display = month_str
                
                cat_name = row_data['category'] or REPORT_LABELS['general']
                supplier_name = row_data['supplier'] or "ללא ספק"
                amount = row_data['amount']
                
                # Row fill - alternate colors
                row_fill = fill_light if row_idx % 2 == 0 else fill_alt
                
                # Column order: A=amount, B=supplier, C=category, D=date
                ws[f'A{current_row}'] = f"{amount:,.2f} ₪"
                ws[f'A{current_row}'].font = data_font
                ws[f'A{current_row}'].fill = row_fill
                ws[f'A{current_row}'].border = thin_border
                ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                ws[f'B{current_row}'] = supplier_name
                ws[f'B{current_row}'].font = data_font
                ws[f'B{current_row}'].fill = row_fill
                ws[f'B{current_row}'].border = thin_border
                ws[f'B{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
                
                ws[f'C{current_row}'] = cat_name
                ws[f'C{current_row}'].font = data_bold_font
                ws[f'C{current_row}'].fill = row_fill
                ws[f'C{current_row}'].border = thin_border
                ws[f'C{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
                
                ws[f'D{current_row}'] = month_display
                ws[f'D{current_row}'].font = data_font
                ws[f'D{current_row}'].fill = row_fill
                ws[f'D{current_row}'].border = thin_border
                ws[f'D{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
                ws[f'D{current_row}'].font = money_negative_font
                ws[f'D{current_row}'].fill = row_fill
                ws[f'D{current_row}'].border = thin_border
                ws[f'D{current_row}'].alignment = center_align
                
                current_row += 1

            current_row += 1  # Spacer

        # 5. Transactions - Professional grouped transactions
        if options.include_transactions and transactions:
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws.row_dimensions[current_row].height = 32
            tx_header = ws[f'A{current_row}']
            tx_header.value = f"📝  {REPORT_LABELS['transaction_details']}"
            tx_header.font = h2_font
            tx_header.fill = fill_h2
            tx_header.alignment = center_align
            tx_header.border = medium_border
            current_row += 1

            # Group transactions by year first, then by category
            from collections import defaultdict
            transactions_by_year_and_category = defaultdict(lambda: defaultdict(list))
            selected_categories = set(options.categories) if options.categories and len(options.categories) > 0 else None

            for tx in transactions:
                if isinstance(tx, dict):
                    cat_name = tx.get('category') or REPORT_LABELS['general']
                    tx_date = tx.get('tx_date')
                else:
                    cat_name = tx.category.name if tx.category else REPORT_LABELS['general']
                    tx_date = tx.tx_date
                
                # Extract year from transaction date
                if isinstance(tx_date, str):
                    try:
                        if 'T' in tx_date:
                            tx_date = date.fromisoformat(tx_date.split('T')[0])
                        else:
                            tx_date = date.fromisoformat(tx_date)
                    except:
                        # If parsing fails, use current year
                        tx_date = date.today()
                elif not isinstance(tx_date, date):
                    tx_date = date.today()
                
                year = tx_date.year

                if selected_categories is None or cat_name in selected_categories:
                    transactions_by_year_and_category[year][cat_name].append(tx)

            # Sort years in descending order (newest first)
            sorted_years = sorted(transactions_by_year_and_category.keys(), reverse=True)

            # Create tables for each year, then each category within that year
            for year in sorted_years:
                # Year header
                current_row += 2 if year != sorted_years[0] else 0  # Add space before new year (except first)
                ws.merge_cells(f'A{current_row}:E{current_row}')
                ws.row_dimensions[current_row].height = 30
                year_header = ws[f'A{current_row}']
                year_header.value = f"שנת {year}"
                year_header.font = h2_font
                year_header.fill = fill_title
                year_header.alignment = center_align
                year_header.border = medium_border
                current_row += 1
                
                # Get categories for this year, sorted alphabetically
                year_categories = sorted(transactions_by_year_and_category[year].keys())
                
                for cat_idx, cat_name in enumerate(year_categories):
                    cat_transactions = transactions_by_year_and_category[year][cat_name]
                # Check which columns have data
                has_suppliers = any(
                    (tx.get('supplier_name') if isinstance(tx, dict) else (tx.supplier.name if tx.supplier else None))
                    for tx in cat_transactions
                ) if cat_transactions else False
                
                has_descriptions = any(
                    (tx.get('description') if isinstance(tx, dict) else tx.description)
                    for tx in cat_transactions
                ) if cat_transactions else False

                # Build dynamic columns list based on available data
                columns = ['date', 'type', 'amount']
                if has_suppliers:
                    columns.append('supplier')
                if has_descriptions:
                    columns.append('description')
                
                col_letters = ['A', 'B', 'C', 'D', 'E'][:len(columns)]
                max_col = col_letters[-1]

                # Category header - with amber accent
                ws.merge_cells(f'A{current_row}:{max_col}{current_row}')
                ws.row_dimensions[current_row].height = 28
                cat_header = ws[f'A{current_row}']
                cat_header.value = f"📁  {REPORT_LABELS['category']}: {cat_name}"
                cat_header.font = h2_font
                cat_header.fill = fill_category_header
                cat_header.alignment = center_align
                cat_header.border = medium_border
                current_row += 1

                # Build table headers dynamically
                col_to_label = {
                    'date': REPORT_LABELS['date'],
                    'type': REPORT_LABELS['type'],
                    'amount': REPORT_LABELS['amount'],
                    'supplier': REPORT_LABELS['supplier'],
                    'description': REPORT_LABELS['description']
                }
                
                ws.row_dimensions[current_row].height = 26
                for i, col_name in enumerate(columns):
                    cell = ws[f'{col_letters[i]}{current_row}']
                    cell.value = col_to_label[col_name]
                    cell.font = header_font
                    cell.fill = fill_header
                    cell.alignment = center_align
                    cell.border = thin_border
                current_row += 1

                # Calculate category total for summary
                cat_total_income = 0
                cat_total_expense = 0

                # Add transaction rows with alternating colors
                for row_idx, tx in enumerate(cat_transactions):
                    ws.row_dimensions[current_row].height = 22
                    
                    if isinstance(tx, dict):
                        is_income = tx.get('type') == "Income"
                        tx_type = REPORT_LABELS['income'] if is_income else REPORT_LABELS['expense']
                        supplier_name = tx.get('supplier_name') or ""
                        tx_desc = tx.get('description') or ""
                        tx_date = tx.get('tx_date')
                        tx_amount = tx.get('amount', 0)
                    else:
                        is_income = tx.type == "Income"
                        tx_type = REPORT_LABELS['income'] if is_income else REPORT_LABELS['expense']
                        supplier_name = tx.supplier.name if tx.supplier else ""
                        tx_desc = tx.description or ""
                        tx_date = tx.tx_date
                        tx_amount = tx.amount

                    # Track totals
                    if is_income:
                        cat_total_income += tx_amount
                    else:
                        cat_total_expense += tx_amount

                    col_to_value = {
                        'date': format_date_hebrew(tx_date),
                        'type': tx_type,
                        'amount': f"{tx_amount:,.2f} ₪",
                        'supplier': supplier_name,
                        'description': tx_desc
                    }
                    
                    # Alternate row colors
                    row_fill = fill_light if row_idx % 2 == 0 else fill_alt
                    
                    for i, col_name in enumerate(columns):
                        cell = ws[f'{col_letters[i]}{current_row}']
                        cell.value = col_to_value[col_name]
                        cell.font = data_font
                        cell.fill = row_fill
                        cell.border = thin_border
                        cell.alignment = center_align
                        
                        # Special styling for type and amount columns
                        if col_name == 'type':
                            if is_income:
                                cell.font = money_positive_font
                            else:
                                cell.font = money_negative_font
                        elif col_name == 'amount':
                            if is_income:
                                cell.font = money_positive_font
                            else:
                                cell.font = money_negative_font
                    
                    current_row += 1

                # Category summary row
                ws.row_dimensions[current_row].height = 26
                ws.merge_cells(f'A{current_row}:B{current_row}')
                summary_cell = ws[f'A{current_row}']
                summary_cell.value = f"סה״כ {cat_name}"
                summary_cell.font = data_bold_font
                summary_cell.fill = fill_amber_light
                summary_cell.alignment = center_align
                summary_cell.border = thin_border
                
                cat_net = cat_total_income - cat_total_expense
                amount_cell = ws[f'C{current_row}']
                amount_cell.value = f"{cat_net:,.2f} ₪"
                amount_cell.font = money_positive_font if cat_net >= 0 else money_negative_font
                amount_cell.fill = fill_amber_light
                amount_cell.alignment = center_align
                amount_cell.border = thin_border
                
                # Fill remaining cells in summary row
                for col in col_letters[3:]:
                    ws[f'{col}{current_row}'].fill = fill_amber_light
                    ws[f'{col}{current_row}'].border = thin_border
                
                current_row += 2  # Spacer between categories

            # ========== COMPREHENSIVE PERIOD SUMMARY ==========
            current_row += 3  # Extra spacing before summary
            
            # Period Summary Header
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws.row_dimensions[current_row].height = 35
            period_summary_header = ws[f'A{current_row}']
            period_summary_header.value = "📊  סיכום כולל לתקופה"
            period_summary_header.font = title_font
            period_summary_header.fill = fill_title
            period_summary_header.alignment = center_align
            period_summary_header.border = medium_border
            current_row += 2
            
            # Calculate totals by category across all years
            category_totals = {}
            for year in sorted_years:
                for cat_name in transactions_by_year_and_category[year].keys():
                    if cat_name not in category_totals:
                        category_totals[cat_name] = {'income': 0, 'expense': 0}
                    
                    for tx in transactions_by_year_and_category[year][cat_name]:
                        if isinstance(tx, dict):
                            tx_type = tx.get('type')
                            tx_amount = tx.get('amount', 0)
                        else:
                            tx_type = tx.type
                            tx_amount = tx.amount
                        
                        if tx_type == "Income":
                            category_totals[cat_name]['income'] += float(tx_amount)
                        else:
                            category_totals[cat_name]['expense'] += float(tx_amount)
            
            # Category Expense Summary Section
            ws.merge_cells(f'A{current_row}:D{current_row}')
            ws.row_dimensions[current_row].height = 30
            cat_summary_header = ws[f'A{current_row}']
            cat_summary_header.value = "📁  הוצאות לפי קטגוריה"
            cat_summary_header.font = h2_font
            cat_summary_header.fill = fill_category_header
            cat_summary_header.alignment = center_align
            cat_summary_header.border = medium_border
            current_row += 1
            
            # Headers for category expense summary
            ws.row_dimensions[current_row].height = 28
            cat_headers = [REPORT_LABELS['category'], REPORT_LABELS['expenses'], REPORT_LABELS['income'], "נטו"]
            for idx, header in enumerate(cat_headers):
                col = get_column_letter(idx + 1)
                cell = ws[f'{col}{current_row}']
                cell.value = header
                cell.font = header_font
                cell.fill = fill_header
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1
            
            grand_total_income = 0
            grand_total_expense = 0
            
            for row_idx, (cat_name, totals) in enumerate(sorted(category_totals.items())):
                ws.row_dimensions[current_row].height = 24
                cat_income = totals['income']
                cat_expense = totals['expense']
                cat_net = cat_income - cat_expense
                
                grand_total_income += cat_income
                grand_total_expense += cat_expense
                
                row_fill = fill_light if row_idx % 2 == 0 else fill_alt
                
                ws[f'A{current_row}'] = cat_name
                ws[f'A{current_row}'].font = data_bold_font
                ws[f'A{current_row}'].fill = row_fill
                ws[f'A{current_row}'].border = thin_border
                ws[f'A{current_row}'].alignment = center_align
                
                ws[f'B{current_row}'] = f"{cat_expense:,.2f} ₪"
                ws[f'B{current_row}'].font = money_negative_font
                ws[f'B{current_row}'].fill = row_fill
                ws[f'B{current_row}'].border = thin_border
                ws[f'B{current_row}'].alignment = center_align
                
                ws[f'C{current_row}'] = f"{cat_income:,.2f} ₪"
                ws[f'C{current_row}'].font = money_positive_font
                ws[f'C{current_row}'].fill = row_fill
                ws[f'C{current_row}'].border = thin_border
                ws[f'C{current_row}'].alignment = center_align
                
                ws[f'D{current_row}'] = f"{cat_net:,.2f} ₪"
                ws[f'D{current_row}'].font = money_positive_font if cat_net >= 0 else money_negative_font
                ws[f'D{current_row}'].fill = row_fill
                ws[f'D{current_row}'].border = thin_border
                ws[f'D{current_row}'].alignment = center_align
                
                current_row += 1
            
            # Grand total row for categories
            grand_net = grand_total_income - grand_total_expense
            ws.row_dimensions[current_row].height = 28
            
            ws[f'A{current_row}'] = "סה״כ"
            ws[f'A{current_row}'].font = data_bold_font
            ws[f'A{current_row}'].fill = fill_teal_light
            ws[f'A{current_row}'].border = medium_border
            ws[f'A{current_row}'].alignment = center_align
            
            ws[f'B{current_row}'] = f"{grand_total_expense:,.2f} ₪"
            ws[f'B{current_row}'].font = money_negative_font
            ws[f'B{current_row}'].fill = fill_teal_light
            ws[f'B{current_row}'].border = medium_border
            ws[f'B{current_row}'].alignment = center_align
            
            ws[f'C{current_row}'] = f"{grand_total_income:,.2f} ₪"
            ws[f'C{current_row}'].font = money_positive_font
            ws[f'C{current_row}'].fill = fill_teal_light
            ws[f'C{current_row}'].border = medium_border
            ws[f'C{current_row}'].alignment = center_align
            
            ws[f'D{current_row}'] = f"{grand_net:,.2f} ₪"
            ws[f'D{current_row}'].font = money_positive_font if grand_net >= 0 else money_negative_font
            ws[f'D{current_row}'].fill = fill_teal_light
            ws[f'D{current_row}'].border = medium_border
            ws[f'D{current_row}'].alignment = center_align
            
            current_row += 3
            
            # Budget vs Actual Summary (if budgets available)
            if budgets and len(budgets) > 0:
                ws.merge_cells(f'A{current_row}:E{current_row}')
                ws.row_dimensions[current_row].height = 30
                budget_summary_header = ws[f'A{current_row}']
                budget_summary_header.value = "📋  תקציב מול ביצוע לתקופה"
                budget_summary_header.font = h2_font
                budget_summary_header.fill = fill_h2
                budget_summary_header.alignment = center_align
                budget_summary_header.border = medium_border
                current_row += 1
                
                # Budget headers - reversed for Hebrew RTL: category on right, calculations on left
                ws.row_dimensions[current_row].height = 28
                budget_headers = [
                    "ניצול %",
                    REPORT_LABELS['remaining'],
                    REPORT_LABELS['used'],
                    REPORT_LABELS['budget'],
                    REPORT_LABELS['category']
                ]
                for idx, header in enumerate(budget_headers):
                    col = get_column_letter(idx + 1)
                    cell = ws[f'{col}{current_row}']
                    cell.value = header
                    cell.font = header_font
                    cell.fill = fill_header
                    cell.alignment = center_align
                    cell.border = thin_border
                current_row += 1
                
                total_budget = 0
                total_spent = 0
                
                for row_idx, b in enumerate(budgets):
                    ws.row_dimensions[current_row].height = 24
                    cat_name = b['category'] if b['category'] else REPORT_LABELS['general']
                    budget_amount = b['amount']
                    spent_amount = b['spent_amount']
                    remaining = b['remaining_amount']
                    usage_pct = (spent_amount / budget_amount * 100) if budget_amount > 0 else 0
                    
                    total_budget += budget_amount
                    total_spent += spent_amount
                    
                    row_fill = fill_light if row_idx % 2 == 0 else fill_alt
                    if usage_pct > 100:
                        row_fill = fill_rose_light
                    
                    # Column order: E=category, D=budget, C=used, B=remaining, A=usage%
                    ws[f'E{current_row}'] = cat_name
                    ws[f'E{current_row}'].font = data_bold_font
                    ws[f'E{current_row}'].fill = row_fill
                    ws[f'E{current_row}'].border = thin_border
                    ws[f'E{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
                    
                    ws[f'D{current_row}'] = f"{budget_amount:,.2f} ₪"
                    ws[f'D{current_row}'].font = data_font
                    ws[f'D{current_row}'].fill = row_fill
                    ws[f'D{current_row}'].border = thin_border
                    ws[f'D{current_row}'].alignment = center_align
                    
                    ws[f'C{current_row}'] = f"{spent_amount:,.2f} ₪"
                    if usage_pct > 100:
                        ws[f'C{current_row}'].font = money_negative_font
                    elif usage_pct > 80:
                        ws[f'C{current_row}'].font = Font(name='Arial', bold=True, size=10, color=ACCENT_AMBER)
                    else:
                        ws[f'C{current_row}'].font = data_font
                    ws[f'C{current_row}'].fill = row_fill
                    ws[f'C{current_row}'].border = thin_border
                    ws[f'C{current_row}'].alignment = center_align
                    
                    ws[f'B{current_row}'] = f"{remaining:,.2f} ₪"
                    ws[f'B{current_row}'].font = money_positive_font if remaining >= 0 else money_negative_font
                    ws[f'B{current_row}'].fill = row_fill
                    ws[f'B{current_row}'].border = thin_border
                    ws[f'B{current_row}'].alignment = center_align
                    
                    ws[f'A{current_row}'] = f"{usage_pct:.1f}%"
                    ws[f'A{current_row}'].font = data_font
                    ws[f'A{current_row}'].fill = row_fill
                    ws[f'A{current_row}'].border = thin_border
                    ws[f'A{current_row}'].alignment = center_align
                    
                    current_row += 1
                
                # Budget total row
                total_remaining = total_budget - total_spent
                total_usage = (total_spent / total_budget * 100) if total_budget > 0 else 0
                ws.row_dimensions[current_row].height = 28
                
                # Total row - reversed: E=category, D=budget, C=used, B=remaining, A=usage%
                ws[f'E{current_row}'] = "סה״כ"
                ws[f'E{current_row}'].font = data_bold_font
                ws[f'E{current_row}'].fill = fill_teal_light
                ws[f'E{current_row}'].border = medium_border
                ws[f'E{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
                
                ws[f'D{current_row}'] = f"{total_budget:,.2f} ₪"
                ws[f'D{current_row}'].font = data_bold_font
                ws[f'D{current_row}'].fill = fill_teal_light
                ws[f'D{current_row}'].border = medium_border
                ws[f'D{current_row}'].alignment = center_align
                
                ws[f'C{current_row}'] = f"{total_spent:,.2f} ₪"
                ws[f'C{current_row}'].font = data_bold_font
                ws[f'C{current_row}'].fill = fill_teal_light
                ws[f'C{current_row}'].border = medium_border
                ws[f'C{current_row}'].alignment = center_align
                
                ws[f'B{current_row}'] = f"{total_remaining:,.2f} ₪"
                ws[f'B{current_row}'].font = money_positive_font if total_remaining >= 0 else money_negative_font
                ws[f'B{current_row}'].fill = fill_teal_light
                ws[f'B{current_row}'].border = medium_border
                ws[f'B{current_row}'].alignment = center_align
                
                ws[f'A{current_row}'] = f"{total_usage:.1f}%"
                ws[f'A{current_row}'].font = data_bold_font
                ws[f'A{current_row}'].fill = fill_teal_light
                ws[f'A{current_row}'].border = medium_border
                ws[f'A{current_row}'].alignment = center_align
                ws[f'E{current_row}'].alignment = center_align
                
                current_row += 3
            
            # Final Overall Summary
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws.row_dimensions[current_row].height = 30
            final_summary_header = ws[f'A{current_row}']
            final_summary_header.value = "💰  סיכום פיננסי כולל"
            final_summary_header.font = h2_font
            final_summary_header.fill = fill_h2
            final_summary_header.alignment = center_align
            final_summary_header.border = medium_border
            current_row += 1
            
            # Final summary rows
            ws.row_dimensions[current_row].height = 28
            ws[f'A{current_row}'] = "פרט"
            ws[f'B{current_row}'] = REPORT_LABELS['amount']
            for col in ['A', 'B']:
                ws[f'{col}{current_row}'].font = header_font
                ws[f'{col}{current_row}'].fill = fill_header
                ws[f'{col}{current_row}'].alignment = center_align
                ws[f'{col}{current_row}'].border = thin_border
            current_row += 1
            
            # Income row
            ws.row_dimensions[current_row].height = 26
            ws[f'A{current_row}'] = f"↗️  {REPORT_LABELS['total_income']}"
            ws[f'A{current_row}'].font = data_bold_font
            ws[f'A{current_row}'].fill = fill_emerald_light
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = center_align
            ws[f'B{current_row}'] = f"{grand_total_income:,.2f} ₪"
            ws[f'B{current_row}'].font = money_positive_font
            ws[f'B{current_row}'].fill = fill_emerald_light
            ws[f'B{current_row}'].border = thin_border
            ws[f'B{current_row}'].alignment = center_align
            current_row += 1
            
            # Expense row
            ws.row_dimensions[current_row].height = 26
            ws[f'A{current_row}'] = f"↘️  {REPORT_LABELS['total_expenses']}"
            ws[f'A{current_row}'].font = data_bold_font
            ws[f'A{current_row}'].fill = fill_rose_light
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = center_align
            ws[f'B{current_row}'] = f"{grand_total_expense:,.2f} ₪"
            ws[f'B{current_row}'].font = money_negative_font
            ws[f'B{current_row}'].fill = fill_rose_light
            ws[f'B{current_row}'].border = thin_border
            ws[f'B{current_row}'].alignment = center_align
            current_row += 1
            
            # Profit/Loss row
            profit_fill = fill_teal_light if grand_net >= 0 else fill_rose_light
            profit_font = money_positive_font if grand_net >= 0 else money_negative_font
            ws.row_dimensions[current_row].height = 28
            ws[f'A{current_row}'] = f"📈  {REPORT_LABELS['balance_profit']}"
            ws[f'A{current_row}'].font = data_bold_font
            ws[f'A{current_row}'].fill = profit_fill
            ws[f'A{current_row}'].border = medium_border
            ws[f'A{current_row}'].alignment = center_align
            ws[f'B{current_row}'] = f"{grand_net:,.2f} ₪"
            ws[f'B{current_row}'].font = profit_font
            ws[f'B{current_row}'].fill = profit_fill
            ws[f'B{current_row}'].border = medium_border
            ws[f'B{current_row}'].alignment = center_align
            current_row += 2

        # Charts - Add native Excel charts for better quality
        if options.include_charts:
            try:
                # Check if there's any data to chart
                has_financial_data = summary and (summary.get('income', 0) > 0 or summary.get('expenses', 0) > 0)
                has_transactions = transactions and len(transactions) > 0
                
                if not has_financial_data and not has_transactions:
                    # No data to chart - skip charts section entirely
                    print("INFO: No financial data for Excel charts, skipping charts section")
                else:
                    current_row += 2  # Spacer

                    # Get chart types to render - default to all if not specified
                    chart_types_to_render = options.chart_types or [
                        "income_expense_pie",
                        "expense_by_category_pie",
                        "expense_by_category_bar",
                        "trends_line",
                        "expense_by_supplier_bar",
                        "monthly_trends_line",
                        "budget_vs_actual",
                        "cumulative_expenses"
                    ]
                    
                    charts_added = 0
                    
                    # Try to use native Excel charts first (only if no images provided from frontend)
                    if CHARTS_AVAILABLE and (not chart_images or len(chart_images) == 0):
                        # Add charts section header first
                        ws.merge_cells(f'A{current_row}:E{current_row}')
                        ws.row_dimensions[current_row].height = 30
                        charts_header = ws[f'A{current_row}']
                        charts_header.value = "📊  גרפים"
                        charts_header.font = h2_font
                        charts_header.fill = fill_h2
                        charts_header.alignment = center_align
                        charts_header.border = medium_border
                        current_row += 2
                        
                        charts_added = self._add_native_excel_charts(
                            wb, ws, current_row, summary, transactions, chart_types_to_render
                        )
                        if charts_added > 0:
                            current_row += charts_added * 18
                    else:
                        # Use image-based charts from frontend
                        from openpyxl.drawing.image import Image as XLImage

                        CHART_TITLES = {
                            "income_expense_pie": "הכנסות מול הוצאות",
                            "expense_by_category_pie": "הוצאות לפי קטגוריה (עוגה)",
                            "expense_by_category_bar": "הוצאות לפי קטגוריה (עמודות)",
                            "trends_line": "מגמות לאורך זמן",
                            "expense_by_supplier_bar": "הוצאות לפי ספק",
                            "monthly_trends_line": "מגמות חודשיות",
                            "budget_vs_actual": "תקציב מול ביצוע",
                            "cumulative_expenses": "הוצאות מצטברות"
                        }

                        charts_to_render = {}

                        # Use provided images - translate keys to Hebrew
                        if chart_images:
                            for key, img_data in chart_images.items():
                                hebrew_title = CHART_TITLES.get(key, key)
                                charts_to_render[hebrew_title] = img_data
                        # Otherwise generate them server-side
                        else:
                            for chart_type in chart_types_to_render:
                                try:
                                    print(f"INFO: Creating chart for Excel: {chart_type}")
                                    chart_buffer = self._create_chart_image(chart_type, {}, summary, transactions, budgets)
                                    if chart_buffer:
                                        chart_buffer.seek(0)
                                        chart_bytes = chart_buffer.read()
                                        if chart_bytes and len(chart_bytes) > 100:
                                            chart_name = CHART_TITLES.get(chart_type, chart_type)
                                            charts_to_render[chart_name] = chart_bytes
                                except Exception as e:
                                    print(f"WARNING: Error preparing chart {chart_type}: {e}")

                        if charts_to_render:
                            # Add charts section header
                            ws.merge_cells(f'A{current_row}:E{current_row}')
                            ws.row_dimensions[current_row].height = 30
                            charts_header = ws[f'A{current_row}']
                            charts_header.value = "📊  גרפים"
                            charts_header.font = h2_font
                            charts_header.fill = fill_h2
                            charts_header.alignment = center_align
                            charts_header.border = medium_border
                            current_row += 2
                            
                            row = current_row
                            for chart_name, image_bytes in charts_to_render.items():
                                try:
                                    if not image_bytes or len(image_bytes) < 100:
                                        continue
                                    img_buffer = BytesIO(image_bytes)
                                    img = XLImage(img_buffer)
                                    # הגדלה לאיכות טובה יותר ב-Excel
                                    img.width = 600
                                    img.height = 400
                                    ws.add_image(img, f'A{row}')
                                    ws.row_dimensions[row].height = 300
                                    row += 16
                                    charts_added += 1
                                except Exception as e:
                                    print(f"אזהרה: הוספת גרף {chart_name} ל-Excel נכשלה: {e}")
                            current_row = row
                    
                    if charts_added == 0:
                        print("INFO: No charts were added to Excel")

            except Exception as e:
                import traceback
                print(f"אזהרה: שגיאה בסעיף גרפים ב-Excel: {e}")
                traceback.print_exc()

        # Limit the used range to only the rows and columns we actually used
        # This ensures the rest of the sheet is empty
        max_col = 'E'  # Maximum column we might use
        if current_row > 1:
            # Set the print area to only the used range
            ws.print_area = f'A1:{max_col}{current_row - 1}'
            # Delete any rows/columns beyond what we used (optional, but ensures clean sheet)
            # Note: openpyxl doesn't have a direct way to delete unused rows/columns,
            # but limiting print_area and not setting values beyond current_row achieves the goal

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()


    async def _generate_zip(self, project, options, report_content, transactions) -> bytes:
        from backend.services.s3_service import S3Service
        try:
            s3_service = S3Service()
            has_s3 = True
        except Exception:
            has_s3 = False

        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
            # Report File - Build filename with project name and date range
            ext = "xlsx" if options.format == "zip" else "pdf"  # Default to excel inside zip if zip requested directly?
            # Actually, if options.format is zip, we generated excel above.
            
            # Sanitize project name for filename
            safe_project_name = "".join([c for c in project.name if
                                         c.isalnum() or c in (' ', '-', '_')]).strip() if project.name else f"project_{project.id}"
            
            # Add date range to filename
            if options.start_date and options.end_date:
                date_range_str = f"{options.start_date.strftime('%Y-%m-%d')}_{options.end_date.strftime('%Y-%m-%d')}"
            elif options.start_date:
                date_range_str = f"מ-{options.start_date.strftime('%Y-%m-%d')}"
            elif options.end_date:
                date_range_str = f"עד-{options.end_date.strftime('%Y-%m-%d')}"
            else:
                date_range_str = "כל-התקופות"
            
            report_filename = f"{safe_project_name}_{date_range_str}.{ext}"
            zf.writestr(report_filename, report_content)

            if has_s3:
                # Add Contract if requested
                if options.include_project_contract and project.contract_file_url:
                    try:
                        contract_content = s3_service.get_file_content(project.contract_file_url)
                        if contract_content:
                            fname = project.contract_file_url.split('/')[-1]
                            zf.writestr(f"contract_{fname}", contract_content)
                    except Exception as e:
                        print(f"שגיאה בהוספת חוזה ל-ZIP: {e}")

                # Add Image if requested
                if options.include_project_image and project.image_url:
                    try:
                        # Assuming image_url is a path or key relative to bucket/base
                        image_content = s3_service.get_file_content(project.image_url)
                        if image_content:
                            fname = project.image_url.split('/')[-1]
                            zf.writestr(f"project_image_{fname}", image_content)
                    except Exception as e:
                        print(f"שגיאה בהוספת תמונה ל-ZIP: {e}")

                # Documents
                if options.include_transactions:
                    for tx in transactions:
                        file_path = tx.get('file_path') if isinstance(tx, dict) else getattr(tx, 'file_path', None)
                        if file_path:
                            try:
                                content = s3_service.get_file_content(file_path)
                                if content:
                                    # Safe filename
                                    tx_date = tx.get('tx_date') if isinstance(tx, dict) else tx.tx_date
                                    tx_id = tx.get('id') if isinstance(tx, dict) else tx.id
                                    fname = f"{tx_date}_{tx_id}.{file_path.split('.')[-1]}"
                                    zf.writestr(f"documents/{fname}", content)
                            except Exception:
                                pass

        output.seek(0)
        return output.read()


    async def generate_excel_report(self, project_id: int) -> bytes:
        """Generate Excel report for a project"""
        # Fetch data
        project_data = await self.project_profitability(project_id)
        transactions = await self.get_project_transactions(project_id)
        expense_categories = await self.get_project_expense_categories(project_id)

        # Get Project details
        proj = (await self.db.execute(select(Project).where(Project.id == project_id))).scalar_one()

        wb = Workbook()

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        fill_blue = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        fill_orange = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
        fill_green = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")

        # 1. Summary Sheet
        ws_summary = wb.active
        ws_summary.title = REPORT_LABELS['financial_summary'][:30]
        ws_summary.sheet_view.rightToLeft = True

        summary_data = [
            [f"{REPORT_LABELS['project_report']}: {proj.name}"],
            [f"{REPORT_LABELS['production_date']}: {date.today().strftime('%d/%m/%Y')}"],
            [],
            [REPORT_LABELS['financial_summary']],
            [REPORT_LABELS['total_income'], project_data["income"]],
            [REPORT_LABELS['total_expenses'], project_data["expenses"]],
            [REPORT_LABELS['profit'], project_data["profit"]],
            [REPORT_LABELS['monthly_budget'], project_data["budget_monthly"]],
            [REPORT_LABELS['annual_budget'], project_data["budget_annual"]],
        ]

        for row in summary_data:
            ws_summary.append(row)

        ws_summary.column_dimensions['A'].width = 20
        ws_summary['D1'].fill = fill_blue  # Just an example if we had headers properly

        # 2. Transactions Sheet
        ws_tx = wb.create_sheet(REPORT_LABELS['transaction_details'][:30])
        ws_tx.sheet_view.rightToLeft = True
        
        # Check which columns have data
        has_category = any(tx.get("category") for tx in transactions) if transactions else False
        has_description = any(tx.get("description") for tx in transactions) if transactions else False
        has_payment_method = any(tx.get("payment_method") for tx in transactions) if transactions else False
        has_notes = any(tx.get("notes") for tx in transactions) if transactions else False
        has_file = any(tx.get("file_path") for tx in transactions) if transactions else False
        
        # Build dynamic column list
        columns = ['date', 'type', 'amount']  # Always include these
        if has_category:
            columns.append('category')
        if has_description:
            columns.append('description')
        if has_payment_method:
            columns.append('payment_method')
        if has_notes:
            columns.append('notes')
        if has_file:
            columns.append('file')
        
        col_to_label = {
            'date': REPORT_LABELS['date'],
            'type': REPORT_LABELS['type'],
            'amount': REPORT_LABELS['amount'],
            'category': REPORT_LABELS['category'],
            'description': REPORT_LABELS['description'],
            'payment_method': REPORT_LABELS['payment_method'],
            'notes': REPORT_LABELS['notes'],
            'file': REPORT_LABELS['file']
        }
        
        col_widths = {
            'date': 12,
            'type': 10,
            'amount': 12,
            'category': 15,
            'description': 30,
            'payment_method': 15,
            'notes': 20,
            'file': 8
        }
        
        # Add headers
        headers = [col_to_label[col] for col in columns]
        ws_tx.append(headers)
        for cell in ws_tx[1]:
            cell.font = header_font
            cell.fill = fill_orange
            cell.alignment = Alignment(horizontal='center')

        # Add transaction rows
        for tx in transactions:
            tx_type = REPORT_LABELS['income'] if tx["type"] == "Income" else REPORT_LABELS['expense']
            col_to_value = {
                'date': format_date_hebrew(tx["tx_date"]),
                'type': tx_type,
                'amount': tx["amount"],
                'category': tx["category"] or "",
                'description': tx["description"] or "",
                'payment_method': translate_payment_method(tx.get("payment_method")),
                'notes': tx["notes"] or "",
                'file': REPORT_LABELS['yes'] if tx.get("file_path") else REPORT_LABELS['no']
            }
            row = [col_to_value[col] for col in columns]
            ws_tx.append(row)

        # Set column widths
        col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        for i, col in enumerate(columns):
            if i < len(col_letters):
                ws_tx.column_dimensions[col_letters[i]].width = col_widths[col]

        # 3. Categories Breakdown - reversed for Hebrew RTL: category on right, amount on left
        ws_cat = wb.create_sheet(REPORT_LABELS['categories'][:30])
        ws_cat.sheet_view.rightToLeft = True
        ws_cat.append([REPORT_LABELS['amount'], REPORT_LABELS['category']])
        for cell in ws_cat[1]:
            cell.font = header_font
            cell.fill = fill_green
            cell.alignment = Alignment(horizontal='center')
        # Set category column (B) to right alignment
        ws_cat['B1'].alignment = Alignment(horizontal='right')

        for cat in expense_categories:
            ws_cat.append([cat["amount"], cat["category"]])
            # Set category column (B) to right alignment for data rows
            last_row = ws_cat.max_row
            ws_cat[f'B{last_row}'].alignment = Alignment(horizontal='right')
            ws_cat[f'A{last_row}'].alignment = Alignment(horizontal='left')

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()


async def generate_zip_export(self, project_id: int) -> bytes:
    """Generate ZIP export with Excel report and transaction documents"""
    # Generate Excel
    excel_data = await self.generate_excel_report(project_id)

    # Get transactions for documents
    transactions = await self.get_project_transactions(project_id)

    # Initialize S3 Service
    from backend.services.s3_service import S3Service
    try:
        s3_service = S3Service()
        has_s3 = True
    except Exception:
        has_s3 = False
        print("אזהרה: שירות S3 לא זמין לייצוא ZIP")

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        # Add Excel report
        zf.writestr(f"project_{project_id}_report.xlsx", excel_data)

        # Add documents
        if has_s3:
            for tx in transactions:
                file_path = tx.get("file_path")
                if file_path:
                    try:
                        # Extract filename from path or URL
                        original_filename = file_path.split("/")[-1]
                        # Use transaction ID and date to make filename unique and meaningful
                        ext = original_filename.split(".")[-1] if "." in original_filename else "bin"
                        filename = f"{tx['tx_date']}_{tx['type']}_{tx['id']}.{ext}"

                        content = s3_service.get_file_content(file_path)
                        if content:
                            zf.writestr(f"documents/{filename}", content)
                    except Exception as e:
                        print(f"שגיאה בהוספת קובץ {file_path} ל-ZIP: {e}")

    output.seek(0)
    return output.read()
